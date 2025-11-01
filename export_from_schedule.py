# -*- coding: utf-8 -*-
"""
Простой экспорт расписания в Excel
"""
import datetime as dt
from tkinter import messagebox

def export_schedule_to_excel(data):
    """
    Простой экспорт данных расписания в Excel с группировкой по сменам
    
    Args:
        data: список словарей с данными расписания
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from tkinter import filedialog
        import datetime as dt
        
        if not data:
            messagebox.showwarning("Экспорт", "Нет данных для экспорта.")
            return
        
        # Группируем данные по сменам
        from schedule_tab import _group_by_shifts
        grouped_by_shifts = _group_by_shifts(data)
        
        # Создаем рабочую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Расписание по сменам"
        
        # Стили
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="CCCCCC")
        shift_header_font = Font(bold=True, size=10)
        shift_header_fill = PatternFill("solid", fgColor="E6E6E6")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Заголовки
        headers = [
            "Дата", "Линия", "Смена", "ID задания", "Наименование", "Длительность",
            "Объём", "Кол-во", "Вкус", "Бренд", "Тип", "Начало", "Окончание", "Примечание"
        ]
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        # Записываем данные
        row = 2
        for line in sorted(grouped_by_shifts.keys()):
            line_shifts = grouped_by_shifts[line]
            
            # Группируем смены по дате и сортируем
            sorted_shifts = sorted(line_shifts.items(), key=lambda x: (x[0].split("_")[0], x[0].split("_")[1]))
            
            for shift_key, events in sorted_shifts:
                shift_date, shift_name = shift_key.split("_", 1)
                
                # Сортируем события в смене по времени
                events = sorted(events, key=lambda r: (
                    r.get("date", ""),
                    r.get("start", "")
                ))
                
                for event in events:
                    values = [
                        event.get("date", ""),
                        event.get("line", ""),
                        shift_name,
                        event.get("job_id", ""),
                        event.get("name", ""),
                        event.get("duration", ""),
                        event.get("volume", ""),
                        event.get("qty", ""),
                        event.get("flavor", ""),
                        event.get("brand", ""),
                        event.get("type", ""),
                        event.get("start", ""),
                        event.get("end", ""),
                        event.get("note", "")
                    ]
                    
                    for col, value in enumerate(values, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.border = border
                        # Выравнивание по центру для всех ячеек данных
                        cell.alignment = center_align
                        
                        # Выделяем CIP события
                        if event.get("_auto_cip"):
                            cell.fill = PatternFill("solid", fgColor="FFE6E6")
                        elif event.get("_transition"):
                            cell.fill = PatternFill("solid", fgColor="E6F3FF")
                    
                    row += 1
        
        # Автоширина колонок (как в Excel)
        for col in range(1, len(headers) + 1):
            max_width = 0
            for row_idx in range(1, row):
                cell = ws.cell(row=row_idx, column=col)
                if not cell.value:
                    continue
                
                # Вычисляем ширину текста с учетом шрифта
                cell_value = str(cell.value)
                
                # Базовая ширина на основе количества символов
                text_width = 0
                for char in cell_value:
                    if char.isupper() or char in 'БВГДЖЗКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ':
                        text_width += 1.2
                    elif char in 'ijl!|:;,.':
                        text_width += 0.5
                    elif char in 'mwMWАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ':
                        text_width += 1.3
                    else:
                        text_width += 1.0
                
                # Учитываем размер шрифта
                font_size = 11
                if cell.font and cell.font.size:
                    font_size = cell.font.size
                adjusted_width = text_width * (font_size / 11.0)
                
                # Учитываем жирный шрифт
                if cell.font and cell.font.bold:
                    adjusted_width *= 1.1
                
                max_width = max(max_width, adjusted_width)
            
            # Добавляем больше padding для гарантии, что текст влезет
            final_width = min(max(max_width + 3.5, 10), 80)  # Увеличен padding с 2.5 до 3.5
            ws.column_dimensions[get_column_letter(col)].width = final_width
        
        # Сохраняем файл
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"schedule_shifts_{timestamp}.xlsx"
        
        file_path = filedialog.asksaveasfilename(
            title="Сохранить расписание в Excel",
            defaultextension=".xlsx",
            initialfile=filename,
            filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")]
        )
        
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Экспорт", f"Расписание сохранено в файл:\n{file_path}")
            
    except Exception as e:
        messagebox.showerror("Ошибка экспорта", f"Ошибка при экспорте: {e}")

# Для обратной совместимости
def export_from_schedule(path_json: str = "schedule_data.json"):
    """Старая функция для обратной совместимости"""
    try:
        import json
        with open(path_json, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        schedule_data = data.get("data", data) if isinstance(data, dict) else data
        export_schedule_to_excel(schedule_data)
        
    except Exception as e:
        messagebox.showerror("Ошибка экспорта", f"Не удалось прочитать файл {path_json}: {e}")