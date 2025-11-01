# -*- coding: utf-8 -*-
"""
schedule_excel_export.py — Красивый экспорт расписания в Excel
-------------------------------------------------------------------------------
• Современное форматирование с цветами и стилями
• Группировка по линиям и сменам
• Автоподгонка ширины столбцов
• Заголовки с логотипом и датой
• Условное форматирование для статусов
• Отдельные листы для каждой линии
"""

from __future__ import annotations
import os
import re
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------
# КОНСТАНТЫ И НАСТРОЙКИ
# ---------------------------------------------------------------------

_COLORS = {
    'header': 'FF2E86AB',      # Синий заголовок
    'subheader': 'FFA8DADC',  # Светло-синий подзаголовок
    'production': 'FF90EE90', # Светло-зеленый для производства
    'cip': 'FFFFB6C1',        # Светло-розовый для CIP
    'eviction': 'FFFFD700',   # Золотой для вытеснений
    'format_change': 'FFB0C4DE', # Светло-стальной для переналадки формата
    'border': 'FF000000',     # Черная граница
    'text': 'FF000000',       # Черный текст
    'white': 'FFFFFFFF'       # Белый фон
}

_STYLES = {
    'title': Font(name='Arial', size=16, bold=True, color=_COLORS['white']),
    'subtitle': Font(name='Arial', size=12, bold=True, color=_COLORS['text']),
    'header': Font(name='Arial', size=11, bold=True, color=_COLORS['white']),
    'data': Font(name='Arial', size=10, color=_COLORS['text']),
    'time': Font(name='Arial', size=10, bold=True, color=_COLORS['text'])
}

# ---------------------------------------------------------------------
# ОСНОВНЫЕ ФУНКЦИИ
# ---------------------------------------------------------------------

def export_schedule_to_excel(schedule_data: List[Dict[str, Any]], 
                           output_path: Optional[str] = None) -> bool:
    """
    Экспортирует расписание в красивый Excel файл
    
    Args:
        schedule_data: Данные расписания
        output_path: Путь для сохранения (если None - диалог выбора)
    
    Returns:
        True если экспорт успешен, False иначе
    """
    if not HAS_OPENPYXL:
        messagebox.showerror("Ошибка", 
            "Модуль openpyxl не установлен!\n"
            "Установите: pip install openpyxl")
        return False
    
    if not schedule_data:
        messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
        return False
    
    # Выбор файла для сохранения
    if not output_path:
        output_path = filedialog.asksaveasfilename(
            title="Сохранить расписание в Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")]
        )
        if not output_path:
            return False
    
    try:
        # Создаем рабочую книгу
        wb = openpyxl.Workbook()
        
        # Удаляем стандартный лист
        wb.remove(wb.active)
        
        # Группируем данные по линиям
        grouped_data = _group_schedule_by_lines(schedule_data)
        
        # Создаем листы для каждой линии
        for line_name, line_data in grouped_data.items():
            _create_line_sheet(wb, line_name, line_data)
        
        # Сохраняем файл
        wb.save(output_path)
        
        messagebox.showinfo("Успех", 
            f"Расписание успешно экспортировано в:\n{output_path}")
        return True
        
    except Exception as e:
        messagebox.showerror("Ошибка экспорта", 
            f"Не удалось экспортировать расписание:\n{e}")
        return False

def _group_schedule_by_lines(schedule_data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """Группирует данные расписания по линиям"""
    grouped = {}
    
    for record in schedule_data:
        line = record.get('line', 'Без линии')
        if line not in grouped:
            grouped[line] = []
        grouped[line].append(record)
    
    return grouped

def _create_line_sheet(wb, line_name: str, line_data: List[Dict[str, Any]]):
    """Создает лист для конкретной линии с разделением на смены"""
    ws = wb.create_sheet(title=line_name[:31])  # Excel ограничение длины имени листа
    
    # Заголовок листа
    _add_sheet_header(ws, line_name, len(line_data))
    
    # Группируем данные по сменам
    shifts_data = _group_data_by_shifts(line_data)
    
    # Данные расписания с разделением на смены
    _add_schedule_data_with_shifts(ws, shifts_data)
    
    # Форматирование
    _format_line_sheet(ws, len(line_data))
    
    # Скрываем неиспользуемые столбцы справа
    _hide_unused_columns(ws)

def _add_sheet_header(ws, line_name: str, record_count: int):
    """Добавляет заголовок листа"""
    # Основной заголовок
    ws.merge_cells('A1:E1')
    ws['A1'] = f"📋 РАСПИСАНИЕ ПРОИЗВОДСТВА - {line_name.upper()}"
    ws['A1'].font = _STYLES['title']
    ws['A1'].fill = PatternFill(start_color=_COLORS['header'], end_color=_COLORS['header'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Подзаголовок с датой и количеством записей
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Дата создания: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Записей: {record_count}"
    ws['A2'].font = _STYLES['subtitle']
    ws['A2'].fill = PatternFill(start_color=_COLORS['subheader'], end_color=_COLORS['subheader'], fill_type='solid')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

def _group_data_by_shifts(line_data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """Группирует данные по сменам, используя ту же логику, что и интерфейс расписания"""
    # Импортируем функции из schedule_tab
    try:
        from schedule_tab import _split_jobs_across_shifts, _group_schedule_by_shifts
        # Разбиваем работы по сменам как в интерфейсе
        split_data = _split_jobs_across_shifts(line_data)
        # Группируем по сменам как в интерфейсе
        shifts = _group_schedule_by_shifts(split_data)
        
        # Преобразуем формат для экспорта
        sorted_shifts = {}
        for shift_key, shift_records in shifts.items():
            date_str, shift_name = shift_key.split("_")
            # Преобразуем дату в формат DD.MM
            from datetime import datetime
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            formatted_date = f"{date_obj.day:02d}.{date_obj.month:02d}"
            
            if shift_name == "Дневная":
                shift_display_name = "Дневная смена (8:00-20:00)"
            else:
                shift_display_name = "Ночная смена (20:00-8:00)"
            
            key = f"{formatted_date} - {shift_display_name}"
            sorted_shifts[key] = shift_records
        
        return sorted_shifts
        
    except ImportError:
        # Если не удалось импортировать, используем старую логику
        shifts_by_date = {}
        
        for record in line_data:
            start_time = record.get('start', '')
            date_str, shift_name = _determine_shift_and_date(start_time)
            
            if date_str not in shifts_by_date:
                shifts_by_date[date_str] = {}
            
            if shift_name not in shifts_by_date[date_str]:
                shifts_by_date[date_str][shift_name] = []
            
            shifts_by_date[date_str][shift_name].append(record)
        
        # Сортируем по датам и сменам
        sorted_shifts = {}
        for date_str in sorted(shifts_by_date.keys()):
            date_shifts = shifts_by_date[date_str]
            # Сортируем смены: сначала дневная, потом ночная
            shift_order = ['Дневная смена (8:00-20:00)', 'Ночная смена (20:00-8:00)']
            for shift_name in shift_order:
                if shift_name in date_shifts:
                    key = f"{date_str} - {shift_name}"
                    sorted_shifts[key] = date_shifts[shift_name]
        
        return sorted_shifts


def _determine_shift_and_date(start_time: str) -> tuple[str, str]:
    """Определяет смену и дату по времени начала"""
    try:
        # Извлекаем дату и время (формат: "28.10 08:00")
        if ' ' in start_time:
            date_part, time_part = start_time.split(' ', 1)
            hour = int(time_part.split(':')[0])
        else:
            date_part = ""
            hour = int(start_time.split(':')[0])
        
        # Определяем смену (8:00-8:00 как в расписании)
        if 8 <= hour < 20:
            shift_name = 'Дневная смена (8:00-20:00)'
            # Для дневной смены дата остается той же
            date_str = date_part if date_part else "Неопределенная дата"
        else:
            shift_name = 'Ночная смена (20:00-8:00)'
            # Для ночной смены дата может быть предыдущего дня
            if hour < 8:
                # Если время до 8:00, это ночная смена предыдущего дня
                date_str = date_part if date_part else "Неопределенная дата"
            else:
                # Если время после 20:00, это ночная смена текущего дня
                date_str = date_part if date_part else "Неопределенная дата"
        
        return date_str, shift_name
    except:
        return "Неопределенная дата", "Неопределенная смена"

def _add_schedule_data_with_shifts(ws, shifts_data: Dict[str, List[Dict[str, Any]]]):
    """Добавляет данные расписания с разделением на смены и даты"""
    current_row = 3
    
    for shift_key, shift_records in shifts_data.items():
        # Заголовок смены с датой
        ws.merge_cells(f'A{current_row}:E{current_row}')
        shift_cell = ws.cell(row=current_row, column=1, value=f"📅 {shift_key}")
        shift_cell.font = Font(name='Arial', size=12, bold=True, color=_COLORS['white'])
        shift_cell.fill = PatternFill(start_color=_COLORS['subheader'], end_color=_COLORS['subheader'], fill_type='solid')
        shift_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Заголовки столбцов для смены
        headers = [
            "Время начала", "Время окончания", "Длительность", 
            "Продукт", "Количество"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = _STYLES['header']
            cell.fill = PatternFill(start_color=_COLORS['header'], end_color=_COLORS['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Данные смены
        for record in shift_records:
            # Определяем тип записи и цвет
            record_type = _get_record_type(record)
            fill_color = _get_record_color(record_type)
            
            # Заполняем данные
            start_time = record.get('start', '')
            end_time = record.get('end', '')
            
            # Убираем дату из времени, если она есть (дата уже в заголовке смены)
            if ' ' in start_time:
                start_time = start_time.split(' ')[1]  # Берем только время
            if ' ' in end_time:
                end_time = end_time.split(' ')[1]  # Берем только время
            
            ws.cell(row=current_row, column=1, value=start_time)
            ws.cell(row=current_row, column=2, value=end_time)
            ws.cell(row=current_row, column=3, value=f"{record.get('duration', '')} мин")
            
            # Для продукта: если это переход, показываем тип события вместо длинного описания
            product_name = record.get('name', '')
            if record_type in ['CIP', 'ВЫТЕСНЕНИЕ', 'ПЕРЕНАЛАДКА']:
                # Для автоматических CIP используем полное название из name
                if record.get('job_id', '').startswith('AUTO-CIP-'):
                    product_name = record.get('name', record.get('type', 'CIP'))
                # Для обычных CIP используем значение из поля type (CIP1, CIP2, etc.)
                elif record_type == 'CIP':
                    product_name = record.get('type', 'CIP')
                elif record_type == 'ВЫТЕСНЕНИЕ' or 'ВЫТ' in product_name or 'ВЫТЕСНЕНИЕ' in product_name:
                    product_name = 'ВЫТЕСНЕНИЕ'
                elif record_type == 'ПЕРЕНАЛАДКА':
                    product_name = 'Переналадка формата'
                else:
                    product_name = record_type
            ws.cell(row=current_row, column=4, value=product_name)
            
            # Для количества: если пустое, ставим "-"
            qty = record.get('qty', '')
            if not qty or qty == '':
                qty = '-'
            ws.cell(row=current_row, column=5, value=qty)
            
            # Форматирование строки
            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.font = _STYLES['data']
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                # Выравнивание по центру для всех ячеек данных
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Особое форматирование для времени
                if col in [1, 2]:
                    cell.font = _STYLES['time']
            
            current_row += 1

def _get_record_type(record: Dict[str, Any]) -> str:
    """Определяет тип записи"""
    job_id = record.get('job_id', '')
    if job_id.startswith('CIP-') or job_id.startswith('AUTO-CIP-'):
        return 'CIP'
    elif job_id.startswith('ВЫТ-'):
        return 'ВЫТЕСНЕНИЕ'
    elif job_id.startswith('П-'):
        return 'ПЕРЕНАЛАДКА'
    else:
        return 'ПРОИЗВОДСТВО'

def _get_record_color(record_type: str) -> str:
    """Возвращает цвет для типа записи"""
    color_map = {
        'ПРОИЗВОДСТВО': _COLORS['production'],
        'CIP': _COLORS['cip'],
        'ВЫТЕСНЕНИЕ': _COLORS['eviction'],
        'ПЕРЕНАЛАДКА': _COLORS['format_change']
    }
    return color_map.get(record_type, _COLORS['white'])

def _format_line_sheet(ws, record_count: int):
    """Форматирует лист линии с улучшенным автофитом"""
    print(f"Форматируем лист: {ws.title}, записей: {record_count}")
    
    # Границы
    thin_border = Border(
        left=Side(style='thin', color=_COLORS['border']),
        right=Side(style='thin', color=_COLORS['border']),
        top=Side(style='thin', color=_COLORS['border']),
        bottom=Side(style='thin', color=_COLORS['border'])
    )
    
    # Применяем границы ко всем ячейкам с данными
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border
    
    # Улучшенный автофит столбцов
    _autofit_columns(ws)
    
    # Высота строк
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    
    # Автоподгонка высоты строк с данными
    for row in range(4, max_row + 1):
        ws.row_dimensions[row].height = 18

def _autofit_columns(ws):
    """Автофит столбцов по содержимому, как в Excel"""
    # Словарь для хранения максимальной ширины каждой колонки
    column_widths = {}
    
    print(f"Автофит: обрабатываем лист с {ws.max_row} строками и {ws.max_column} колонками")
    
    for row in ws.iter_rows():
        for cell in row:
            # Пропускаем пустые ячейки и объединенные ячейки
            if not cell.value:
                continue
            
            # Проверяем, является ли ячейка частью объединенной области
            if hasattr(ws, 'merged_cells') and ws.merged_cells:
                is_merged = False
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        is_merged = True
                        break
                if is_merged:
                    continue
            
            col_letter = cell.column_letter
            if col_letter not in column_widths:
                column_widths[col_letter] = 0
            
            # Вычисляем ширину текста с учетом шрифта
            cell_value = str(cell.value)
            
            # Базовая ширина на основе количества символов
            # Учитываем, что разные символы имеют разную ширину
            text_width = 0
            for char in cell_value:
                if char.isupper() or char in 'БВГДЖЗКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ':
                    # Заглавные буквы и кириллица шире
                    text_width += 1.2
                elif char in 'ijl!|:;,.':
                    # Узкие символы
                    text_width += 0.5
                elif char in 'mwMWАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ':
                    # Широкие символы
                    text_width += 1.3
                else:
                    text_width += 1.0
            
            # Учитываем размер шрифта
            font_size = 11  # по умолчанию
            if cell.font and cell.font.size:
                font_size = cell.font.size
            
            # Корректируем ширину на размер шрифта
            # Базовая ширина для шрифта 11
            adjusted_width = text_width * (font_size / 11.0)
            
            # Учитываем жирный шрифт (он шире)
            if cell.font and cell.font.bold:
                adjusted_width *= 1.1
            
            # Обновляем максимум для колонки
            if adjusted_width > column_widths[col_letter]:
                column_widths[col_letter] = adjusted_width
                print(f"Новый максимум для колонки {col_letter}: {adjusted_width:.1f} ('{cell_value[:30]}...')")
    
    # Применяем ширины колонок
    for col_letter, max_width in column_widths.items():
        # Добавляем больше padding для гарантии, что текст влезет
        # Excel использует единицы измерения, где 1 ≈ ширина символа в шрифте Calibri 11
        final_width = min(max_width + 3.5, 80)  # Увеличен padding с 2.5 до 3.5
        ws.column_dimensions[col_letter].width = final_width
        print(f"Устанавливаем ширину колонки {col_letter}: {final_width:.1f}")

def _hide_unused_columns(ws):
    """Скрывает неиспользуемые столбцы справа"""
    # Скрываем столбцы F и далее (индексы 6+)
    for col in range(6, 27):  # F до Z
        ws.column_dimensions[get_column_letter(col)].hidden = True


# ---------------------------------------------------------------------
# GUI ДЛЯ ЭКСПОРТА
# ---------------------------------------------------------------------

class ScheduleExportDialog:
    """Диалог для экспорта расписания"""
    
    def __init__(self, parent, schedule_data: List[Dict[str, Any]]):
        self.parent = parent
        self.schedule_data = schedule_data
        
        # Создаем диалог
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Экспорт расписания в Excel")
        self.dialog.geometry("500x400")
        self.dialog.resizable(False, False)
        
        # Центрируем диалог
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        self._create_widgets()
        self._center_dialog()
    
    def _create_widgets(self):
        """Создает виджеты диалога"""
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.pack(fill="both", expand=True)
        
        # Заголовок
        title_label = ttk.Label(main_frame, text="📊 Экспорт расписания в Excel", 
                                font=("Arial", 14, "bold"))
        title_label.pack(pady=(0, 20))
        
        # Информация о данных
        info_frame = ttk.LabelFrame(main_frame, text="Информация о расписании", padding="10")
        info_frame.pack(fill="x", pady=(0, 20))
        
        total_records = len(self.schedule_data)
        lines = set(record.get('line', 'Без линии') for record in self.schedule_data)
        
        ttk.Label(info_frame, text=f"Всего записей: {total_records}").pack(anchor="w")
        ttk.Label(info_frame, text=f"Линий: {len(lines)}").pack(anchor="w")
        ttk.Label(info_frame, text=f"Линии: {', '.join(sorted(lines))}").pack(anchor="w")
        
        # Настройки экспорта
        settings_frame = ttk.LabelFrame(main_frame, text="Настройки экспорта", padding="10")
        settings_frame.pack(fill="x", pady=(0, 20))
        
        # Автоматическое имя файла
        default_filename = f"Расписание_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        self.filename_var = tk.StringVar(value=default_filename)
        
        ttk.Label(settings_frame, text="Имя файла:").pack(anchor="w")
        filename_frame = ttk.Frame(settings_frame)
        filename_frame.pack(fill="x", pady=(5, 0))
        
        self.filename_entry = ttk.Entry(filename_frame, textvariable=self.filename_var, width=40)
        self.filename_entry.pack(side="left", fill="x", expand=True)
        
        ttk.Button(filename_frame, text="Обзор...", 
                  command=self._browse_file).pack(side="right", padx=(10, 0))
        
        # Кнопки
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(20, 0))
        
        ttk.Button(button_frame, text="Экспорт", 
                  command=self._export).pack(side="right", padx=(10, 0))
        ttk.Button(button_frame, text="Отмена", 
                  command=self.dialog.destroy).pack(side="right")
    
    def _center_dialog(self):
        """Центрирует диалог на экране"""
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (self.dialog.winfo_width() // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (self.dialog.winfo_height() // 2)
        self.dialog.geometry(f"+{x}+{y}")
    
    def _browse_file(self):
        """Выбор файла для сохранения"""
        filename = filedialog.asksaveasfilename(
            title="Сохранить расписание в Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")],
            initialvalue=self.filename_var.get()
        )
        if filename:
            self.filename_var.set(filename)
    
    def _export(self):
        """Выполняет экспорт"""
        filename = self.filename_var.get().strip()
        if not filename:
            messagebox.showwarning("Предупреждение", "Введите имя файла")
            return
        
        # Добавляем расширение если его нет
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # Экспортируем
        success = export_schedule_to_excel(self.schedule_data, filename)
        if success:
            self.dialog.destroy()

def show_export_dialog(parent, schedule_data: List[Dict[str, Any]]):
    """Показывает диалог экспорта"""
    if not HAS_OPENPYXL:
        messagebox.showerror("Ошибка", 
            "Модуль openpyxl не установлен!\n"
            "Установите: pip install openpyxl")
        return
    
    ScheduleExportDialog(parent, schedule_data)

# ---------------------------------------------------------------------
# ТОЧКА ВХОДА
# ---------------------------------------------------------------------

if __name__ == "__main__":
    # Тестовые данные
    test_data = [
        {
            "line": "линия 5",
            "job_id": "J-251028-L05-001",
            "name": "Сироп Имбирный Пряник 1,0 л ТМ «Баринофф»",
            "start": "28.10 08:00",
            "end": "28.10 10:30",
            "duration": "150",
            "qty": "6000",
            "note": ""
        },
        {
            "line": "линия 5", 
            "job_id": "CIP-J-251028-L05-001",
            "name": "CIP",
            "start": "28.10 10:30",
            "end": "28.10 11:00",
            "duration": "30",
            "qty": "",
            "note": "Мойка линии"
        }
    ]
    
    # Тест экспорта
    root = tk.Tk()
    root.withdraw()  # Скрываем главное окно
    
    success = export_schedule_to_excel(test_data)
    print(f"Экспорт {'успешен' if success else 'неудачен'}")
    
    root.destroy()
