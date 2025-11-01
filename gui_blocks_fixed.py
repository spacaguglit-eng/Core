# -*- coding: utf-8 -*-
"""
gui_blocks.py — читаем 2 прямоугольных блока из Excel и показываем в таблице
+ неблокирующая загрузка (поток), прогресс-бар и отмена
+ ускоренное чтение диапазонов (usecols/skiprows/nrows)
"""

import warnings

warnings.simplefilter("ignore")  # убираем шумные UserWarning от openpyxl

import os
import re
import math
import numbers
import threading
import queue
import zipfile  # проверка, что .xlsx — это реально zip (OOXML)
import json
from typing import Dict, List, Tuple, Optional
from openpyxl.formatting.rule import ColorScaleRule
import pandas as pd
# tk_fix313.py — шим для Python 3.13.x, чтобы заработал tkinter.ttk
import sys, tkinter as _tk

if sys.version_info >= (3, 13):
    if not hasattr(_tk, "_flatten"):
        def _flatten(seq):
            for item in seq:
                if isinstance(item, (list, tuple)):
                    yield from _flatten(item)
                else:
                    yield item
        _tk._flatten = _flatten

    if not hasattr(_tk, "_join"):
        def _join(seq, sep=" "):
            return sep.join(map(str, seq))
        _tk._join = _join

    if not hasattr(_tk, "_splitdict"):
        def _splitdict(tk, string):
            # безопасная упрощённая версия
            return dict(item.split("=", 1) for item in string.split() if "=" in item)
        _tk._splitdict = _splitdict

    if not hasattr(_tk, "_stringify"):
        def _stringify(value):
            return str(value)
        _tk._stringify = _stringify

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
# === export to Excel
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from planning_tab import show_planning_tab

from report_core import (
    FilterOpts,
    build_downtime_index as core_build_downtime_index,
    build_summary_rows as core_build_summary_rows,
    build_report_rows as core_build_report_rows,
    compute_oee_matrix as core_compute_oee_matrix,
    top3_for as core_top3_for,
    fmt_top_item as core_fmt_top_item,
)

# === Каталог продуктов (нормализация имён, скорости) =========================
from catalog import (
    Catalog,
    make_default_catalog as make_catalog_default,
)
# === СОРТИРОВКА В Treeview ===================================================
_SORT_STATE: dict[tuple[int, str], bool] = {}

def _sortkey_nat(v: str):
    """Естественная сортировка для смешанных строк"""
    s = str(v).strip().replace("\xa0", " ")
    s_num = s.replace(" ", "").replace(",", ".")
    try:
        return (0, int(float(s_num)))
    except Exception:
        pass
    try:
        return (1, float(s_num))
    except Exception:
        pass
    parts = re.findall(r"\d+|\D+", s.lower())
    return (2, tuple(int(p) if p.isdigit() else p for p in parts))

def _tree_sort_by(tree: ttk.Treeview, col: str):
    """Сортировка Treeview по колонке"""
    rev = _SORT_STATE.get((id(tree), col), False)
    items = list(tree.get_children(""))
    rows = [(tree.set(i, col), i) for i in items]
    rows.sort(key=lambda t: _sortkey_nat(t[0]), reverse=rev)
    for idx, (_, iid) in enumerate(rows):
        tree.move(iid, "", idx)
    _SORT_STATE[(id(tree), col)] = not rev

def enable_tree_sort(tree: ttk.Treeview):
    """Включение сортировки для всех колонок Treeview"""
    for col in tree["columns"]:
        old = tree.heading(col).get("text", "")
        tree.heading(col, text=old, command=lambda c=col: _tree_sort_by(tree, c))

CATALOG: Catalog = make_catalog_default()
CATALOG_JSON_PATH = os.path.join(os.path.dirname(__file__), "catalog_data.json")







# ===== ФИЛЬТРЫ (мультивыбор линий/дней) ======================================
SELECTED_LINES: set[str] = set()
SELECTED_DAYS: set[str] = set()
ALL_LINES: list[str] = []
ALL_DAYS: list[str] = []
VAR_LINES: dict[str, tk.BooleanVar] = {}
VAR_DAYS: dict[str, tk.BooleanVar] = {}

# ===== КОНФИГ ДИАПАЗОНОВ =====================================================
RANGES = {
    "products": [
        {"range": "A21:K31", "headers": False, "round_int": True},  # День
        {"range": "A136:K146", "headers": False, "round_int": True},  # Ночь
    ],
    "downtimes": [
        {
            "range": "A47:Q91",
            "headers": False,
            "round_int": False,
            "drop_cols": ["B", "C", "D", "E", "G", "M", "N", "P", "Q"],
        },
        {
            "range": "A162:Q228",
            "headers": False,
            "round_int": False,
            "drop_cols": ["B", "C", "D", "E", "G", "M", "N", "P", "Q"],
        },
    ],
}

# ============================================================================
DATA: Dict[str, Dict[str, List[List]]] = (
    {}
)  # {block_name: {"array": [[..]], "headers": [..], "meta": {...}}}

# --- утилита: парсинг A1-диапазона ------------------------------------------
_A1_RE = re.compile(r"^\s*([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)\s*$")


def col_letters_to_index(letters: str) -> int:
    """A->1, B->2, ... AA->27"""
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def index_to_col_letters(n: int) -> str:
    """1->A, 2->B, ... 27->AA"""
    s = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s

def parse_a1_range(a1: str) -> Tuple[int, int, int, int]:
    """
    'B7:F38' -> (row1, col1, row2, col2), 1-based
    """
    m = _A1_RE.match(a1)
    if not m:
        raise ValueError(f"Неверный формат диапазона: {a1!r}")
    c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    j1 = col_letters_to_index(c1)
    j2 = col_letters_to_index(c2)
    if r2 < r1 or j2 < j1:
        raise ValueError(f"Диапазон перевёрнут: {a1!r}")
    return r1, j1, r2, j2


def _natural_key(s: str):
    parts = re.findall(r"\d+|\D+", str(s))
    out = []
    for p in parts:
        if p.isdigit():
            out.append((0, int(p)))
        else:
            out.append((1, p.lower()))
    return tuple(out)


def _sel_to_human():
    def _fmt(name, items):
        if not items:
            return f"{name}=Все"
        v = sorted(items, key=_natural_key)
        return f"{name}=" + ",".join(map(str, v[:8])) + ("…" if len(v) > 8 else "")

    return _fmt("Линии", SELECTED_LINES) + " • " + _fmt("Дни", SELECTED_DAYS)


# --- GUI ---------------------------------------------------------------------
root = tk.Tk()
import traceback, sys

def _tk_ex_hook(exc, val, tb):
    # Печатаем в консоль и в лог-файл
    traceback.print_exception(exc, val, tb)
    with open("tk_errors.log", "a", encoding="utf-8") as f:
        traceback.print_exception(exc, val, tb, file=f)

root.report_callback_exception = _tk_ex_hook

root.title("Чтение блоков из Excel (файл или папка)")
root.geometry("1000x600")
# верхняя панель
frm_top = ttk.Frame(root)
frm_top.pack(fill="x", padx=8, pady=6)

btn_file = ttk.Button(frm_top, text="Файл…")
btn_file.pack(side="left")

btn_folder = ttk.Button(frm_top, text="Папка…")
btn_folder.pack(side="left", padx=(6, 0))

lbl_file = ttk.Label(frm_top, text="Источник не выбран", width=60, anchor="w")
lbl_file.pack(side="left", padx=8)
# Кнопка: открыть последний JSON из памяти
def _open_last_json():
    p = _get_last_json()
    if not p:
        messagebox.showinfo("Нет сохранённого JSON", "Путь к последнему JSON не найден.")
        return
    state["path"] = p
    state["paths"] = []
    lbl_file.config(text=p)
    log(f"Последний JSON: {p}")
    start_load([p])

btn_last_json = ttk.Button(frm_top, text="Последний JSON", command=_open_last_json)
btn_last_json.pack(side="left", padx=(6, 0))


# выбор блока
frm_sel = ttk.Frame(root)
frm_sel.pack(fill="x", padx=8, pady=4)

ttk.Label(frm_sel, text="Блок:").pack(side="left")
_summary_option = "Сводка"
combo_block = ttk.Combobox(
    frm_sel, values=[_summary_option], state="readonly", width=60
)
combo_block.set(_summary_option)
combo_block.pack(side="left", padx=6)
# показывать выбранный блок сразу при смене значения
combo_block.bind("<<ComboboxSelected>>", lambda _e: show_block(combo_block.get()))



ttk.Label(frm_sel, text="События:").pack(side="left", padx=(16, 6))
combo_events = ttk.Combobox(frm_sel, state="disabled", width=80)
combo_events.pack(side="left")
lbl_ev_count = ttk.Label(frm_sel, text="")
lbl_ev_count.pack(side="left", padx=6)



# ===== ПАНЕЛЬ ФИЛЬТРОВ (мультивыбор) =========================================
frm_filters = ttk.Frame(root)
frm_filters.pack(fill="x", padx=8, pady=(0, 6))

btn_lines = ttk.Button(frm_filters, text="Линии ▾")
btn_lines.pack(side="left", padx=(0, 6))

btn_days = ttk.Button(frm_filters, text="Дни ▾")
btn_days.pack(side="left", padx=(0, 6))


def _reset_filters():
    SELECTED_LINES.clear()
    SELECTED_DAYS.clear()
    try:
        show_block(_summary_option)
    except Exception:
        pass


ttk.Button(frm_filters, text="Сброс фильтров", command=_reset_filters).pack(
    side="left", padx=6
)
lbl_filters = ttk.Label(frm_filters, text="")
lbl_filters.pack(side="left", padx=12)

def _compute_export_tables_from_summary(headers_sum, rows_sum):
    """Строит набор таблиц для экспорта и предпросмотра из сводных данных:
       возвращает (hdr_lines, rows_lines, hdr_prod, rows_prod)."""
    idx = {h: i for i, h in enumerate(headers_sum)}

    def _to_int(x) -> int:
        if x is None or x == "":
            return 0
        s = str(x).replace(" ", "").replace("\u00A0", "")
        try:
            return int(s.replace(",", ""))
        except Exception:
            try:
                return int(float(s.replace(",", ".")))
            except Exception:
                return 0

    # По продуктам
    hdr_prod = ["Линия", "Смена", "Продукт", "План", "Факт", "%", "Информация"]
    rows_prod = []
    for r in rows_sum:
        line    = str(r[idx["Линия"]]).strip()
        shift   = str(r[idx["Смена"]]).strip()
        product = str(r[idx["Продукт"]]).strip()
        plan    = _to_int(r[idx["Потолок (шт)"]])
        fact    = _to_int(r[idx["Факт (шт)"]])
        pct     = (fact / plan * 100.0) if plan else 0.0
        top3 = _top3_for(r[idx["Продукт"]], r[idx["День"]], r[idx["Смена"]])
        info = "\n".join(_fmt_top_item(t) for t in top3[:3])
        rows_prod.append([line, shift, product, plan, fact, round(pct, 1), info])
    rows_prod.sort(key=lambda x: (x[0], x[1], x[2]))

    # По линиям
    hdr_lines = ["Линия", "План", "Факт", "%", "Информация"]
    totals = {}
    for r in rows_sum:
        line = str(r[idx["Линия"]]).strip()
        plan = _to_int(r[idx["Потолок (шт)"]])
        fact = _to_int(r[idx["Факт (шт)"]])
        d = totals.setdefault(line, {"plan": 0, "fact": 0})
        d["plan"] += plan
        d["fact"] += fact

    def _numkey(s):
        ss = str(s)
        return (0, int(ss)) if ss.isdigit() else (1, ss)

    rows_lines = []
    for line in sorted(totals.keys(), key=_numkey):
        plan = totals[line]["plan"]
        fact = totals[line]["fact"]
        pct  = (fact / plan * 100.0) if plan else 0.0
        info = "\n".join(_top3_for_line(rows_sum, idx, line))
        rows_lines.append([line, plan, fact, round(pct, 1), info])

    return hdr_lines, rows_lines, hdr_prod, rows_prod

def _open_multi_panel(kind: str, anchor_widget):
    items = ALL_LINES if kind == "lines" else ALL_DAYS
    selected = SELECTED_LINES if kind == "lines" else SELECTED_DAYS
    var_map = VAR_LINES if kind == "lines" else VAR_DAYS
    title = "Выбор линий" if kind == "lines" else "Выбор дней"
    if not items:
        return
    win = tk.Toplevel(root)
    win.title(title)
    win.transient(root)
    win.resizable(False, True)
    win.attributes("-topmost", True)
    try:
        x = anchor_widget.winfo_rootx()
        y = anchor_widget.winfo_rooty() + anchor_widget.winfo_height()
        win.geometry(f"+{x}+{y}")
    except Exception:
        pass

    frm_top_btns = ttk.Frame(win)
    frm_top_btns.pack(fill="x", padx=8, pady=(8, 4))

    def _apply_all():
        selected.clear()
        selected.update(items)
        for it in items:
            var_map.setdefault(it, tk.BooleanVar()).set(True)

    def _apply_none():
        selected.clear()
        for it in items:
            var_map.setdefault(it, tk.BooleanVar()).set(False)

    def _apply_invert():
        new = set(items) - set(selected)
        selected.clear()
        selected.update(new)
        for it in items:
            var_map.setdefault(it, tk.BooleanVar()).set(it in new)

    ttk.Button(frm_top_btns, text="Все", command=_apply_all).pack(side="left")
    ttk.Button(frm_top_btns, text="Снять", command=_apply_none).pack(
        side="left", padx=6
    )
    ttk.Button(frm_top_btns, text="Инвертировать", command=_apply_invert).pack(
        side="left"
    )

    frm_list = ttk.Frame(win)
    frm_list.pack(fill="both", expand=True, padx=8, pady=4)

    canvas = tk.Canvas(frm_list, borderwidth=0, highlightthickness=0, width=200)
    vs = ttk.Scrollbar(frm_list, orient="vertical", command=canvas.yview)
    inner = ttk.Frame(canvas)
    inner.bind(
        "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    canvas.create_window((0, 0), window=inner, anchor="nw")
    canvas.configure(yscrollcommand=vs.set)
    canvas.pack(side="left", fill="both", expand=True)
    vs.pack(side="left", fill="y")

    for it in items:
        var = var_map.setdefault(it, tk.BooleanVar(value=(it in selected)))
        ttk.Checkbutton(inner, text=str(it), variable=var).pack(anchor="w")

    frm_bottom = ttk.Frame(win)
    frm_bottom.pack(fill="x", padx=8, pady=(4, 8))

    def _do_apply_and_close():
        chosen = {it for it in items if var_map.get(it) and var_map[it].get()}
        selected.clear()
        selected.update(chosen)
        _rebuild_filter_menus()   # подпись + перерисовка таблиц/матрицы
        win.destroy()

    ttk.Button(frm_bottom, text="Применить", command=_do_apply_and_close).pack(
        side="right"
    )
    ttk.Button(frm_bottom, text="Отмена", command=win.destroy).pack(
        side="right", padx=6
    )
    win.grab_set()


btn_lines.configure(command=lambda: _open_multi_panel("lines", btn_lines))
btn_days.configure(command=lambda: _open_multi_panel("days", btn_days))
# === NOTEBOOK (две вкладки: Таблица и OEE-матрица) ==========================
# === NOTEBOOK ===============================================================
nb = ttk.Notebook(root)

from planning_tab import show_planning_tab
from json_import_tab import show_json_import_tab

# 1) СНАЧАЛА Планирование (чтобы появился tree_plan и вкладка Импорт JSON внутри Расписания)
show_planning_tab(nb)

# 2) Импорт JSON теперь внутри вкладки Расписание, не нужен отдельный вызов
# show_json_import_tab(nb)

# Остальные вкладки уже как угодно
tab_report = ttk.Frame(nb)
nb.add(tab_report, text="Отчёт (Факт/План)")

nb.pack(fill="both", expand=True, padx=8, pady=8)



# Таблица отчёта
frm_report = ttk.Frame(tab_report)
frm_report.pack(fill="both", expand=True, padx=8, pady=(0, 8))

report_tree = ttk.Treeview(frm_report, show="headings")
report_vsb = ttk.Scrollbar(frm_report, orient="vertical", command=report_tree.yview)
report_hsb = ttk.Scrollbar(frm_report, orient="horizontal", command=report_tree.xview)
report_tree.configure(yscrollcommand=report_vsb.set, xscrollcommand=report_hsb.set)

report_tree.grid(row=0, column=0, sticky="nsew")
report_vsb.grid(row=0, column=1, sticky="ns")
report_hsb.grid(row=1, column=0, sticky="ew")
frm_report.rowconfigure(0, weight=1)
frm_report.columnconfigure(0, weight=1)


# --- вкладка "Таблица" и "OEE-матрица" -------------------------------------
tab_table = ttk.Frame(nb)  # сюда перенесём текущую таблицу Treeview
tab_oee = ttk.Frame(nb)    # здесь будем рисовать матрицу
nb.add(tab_table, text="Таблица")
nb.add(tab_oee, text="OEE-матрица")

# --- вкладка "Отчёт (Экспорт)" ---------------------------------------------
tab_report_export = ttk.Frame(nb)
nb.add(tab_report_export, text="Отчёт (Экспорт)")

rep_toolbar = ttk.Frame(tab_report_export)
rep_toolbar.pack(fill="x", padx=8, pady=(8, 4))

btn_make_report = ttk.Button(rep_toolbar, text="Сформировать отчёт")
btn_make_report.pack(side="left")
# рядом с btn_make_report:


# экспорт — из текущей таблицы Treeview:
btn_make_report.configure(command=lambda: export_report_to_excel(tree))

lbl_report_info = ttk.Label(rep_toolbar, text="", foreground="#555")
lbl_report_info.pack(side="left", padx=12)
# --- Предпросмотр отчёта ---
frm_preview = ttk.Frame(tab_report_export)
frm_preview.pack(fill="both", expand=True, padx=8, pady=(4, 8))

nb_preview = ttk.Notebook(frm_preview)
nb_preview.pack(fill="both", expand=True)

# По линиям
tab_prev_lines = ttk.Frame(nb_preview)
nb_preview.add(tab_prev_lines, text="По линиям")
tv_prev_lines = ttk.Treeview(tab_prev_lines, show="headings")
vsb_pl = ttk.Scrollbar(tab_prev_lines, orient="vertical", command=tv_prev_lines.yview)
hsb_pl = ttk.Scrollbar(tab_prev_lines, orient="horizontal", command=tv_prev_lines.xview)
tv_prev_lines.configure(yscrollcommand=vsb_pl.set, xscrollcommand=hsb_pl.set)
tv_prev_lines.grid(row=0, column=0, sticky="nsew")
vsb_pl.grid(row=0, column=1, sticky="ns")
hsb_pl.grid(row=1, column=0, sticky="ew")
tab_prev_lines.rowconfigure(0, weight=1)
tab_prev_lines.columnconfigure(0, weight=1)

# По продуктам
tab_prev_prod = ttk.Frame(nb_preview)
nb_preview.add(tab_prev_prod, text="По продуктам")
tv_prev_prod = ttk.Treeview(tab_prev_prod, show="headings")
vsb_pp = ttk.Scrollbar(tab_prev_prod, orient="vertical", command=tv_prev_prod.yview)
hsb_pp = ttk.Scrollbar(tab_prev_prod, orient="horizontal", command=tv_prev_prod.xview)
tv_prev_prod.configure(yscrollcommand=vsb_pp.set, xscrollcommand=hsb_pp.set)
tv_prev_prod.grid(row=0, column=0, sticky="nsew")
vsb_pp.grid(row=0, column=1, sticky="ns")
hsb_pp.grid(row=1, column=0, sticky="ew")
tab_prev_prod.rowconfigure(0, weight=1)
tab_prev_prod.columnconfigure(0, weight=1)
def render_export_preview():
    headers_sum, rows_sum = build_summary_rows()
    if not rows_sum:
        for tv in (tv_prev_lines, tv_prev_prod):
            tv.delete(*tv.get_children("")); tv["columns"] = ()
        try:
            lbl_report_info.config(text="Нет данных для предпросмотра")
        except Exception:
            pass
        return

    hdrL, rowsL, hdrP, rowsP = _compute_export_tables_from_summary(headers_sum, rows_sum)

    # Линии
    tv_prev_lines.delete(*tv_prev_lines.get_children(""))
    tv_prev_lines["columns"] = [f"c{i}" for i in range(len(hdrL))]
    for i, h in enumerate(hdrL):
        tv_prev_lines.heading(f"c{i}", text=h)
        tv_prev_lines.column(f"c{i}", width=(90 if i in (0, 3) else 120), anchor="w")
    for r in rowsL:
        tv_prev_lines.insert("", "end", values=r)
    enable_tree_sort(tv_prev_lines)

    # Продукты
    tv_prev_prod.delete(*tv_prev_prod.get_children(""))
    tv_prev_prod["columns"] = [f"c{i}" for i in range(len(hdrP))]
    widths = {0: 80, 1: 70, 2: 340, 3: 110, 4: 110, 5: 70, 6: 420}
    for i, h in enumerate(hdrP):
        tv_prev_prod.heading(f"c{i}", text=h)
        tv_prev_prod.column(f"c{i}", width=widths.get(i, 120), anchor="w")
    for r in rowsP:
        tv_prev_prod.insert("", "end", values=r)
    enable_tree_sort(tv_prev_prod)

    try:
        lbl_report_info.config(text="Предпросмотр обновлён")
    except Exception:
        pass

# === Вкладка «Каталог» ======================================================
tab_catalog = ttk.Frame(nb)
nb.add(tab_catalog, text="Каталог")

# Основной контейнер
frm_cat = ttk.Frame(tab_catalog)
frm_cat.pack(fill="both", expand=True, padx=10, pady=10)

# Панель инструментов каталога
frm_cat_toolbar = ttk.Frame(frm_cat)
frm_cat_toolbar.pack(fill="x", pady=(0, 10))

# Группа: Управление записями
left_group = ttk.LabelFrame(frm_cat_toolbar, text="Управление записями", padding=8)
left_group.pack(side="left", padx=(0, 10))

btn_add_cat = ttk.Button(left_group, text="➕ Добавить строку")
btn_add_cat.pack(side="left", padx=(0, 6))

btn_del_cat = ttk.Button(left_group, text="🗑️ Удалить выбранные")
btn_del_cat.pack(side="left", padx=(0, 6))

# Группа: Файл
right_group = ttk.LabelFrame(frm_cat_toolbar, text="Файл", padding=8)
right_group.pack(side="left", padx=(0, 10))

btn_load_cat = ttk.Button(right_group, text="📂 Загрузить каталог")
btn_load_cat.pack(side="left", padx=(0, 6))

btn_save_cat = ttk.Button(right_group, text="💾 Сохранить каталог")
btn_save_cat.pack(side="left", padx=(0, 6))

# Группа: Парсинг
parse_group = ttk.LabelFrame(frm_cat_toolbar, text="Парсинг", padding=8)
parse_group.pack(side="left")

CATALOG.add_parsing_button(parse_group)

# Панель фильтров
filter_frame = ttk.LabelFrame(frm_cat, text="Фильтры", padding=8)
filter_frame.pack(fill="x", pady=(0, 10))

# Все фильтры в одну строку
filters_row = ttk.Frame(filter_frame)
filters_row.pack(fill="x")

ttk.Label(filters_row, text="Наименование:").pack(side="left", padx=(0, 5))
filter_name_var = tk.StringVar()
filter_name_entry = ttk.Entry(filters_row, textvariable=filter_name_var, width=20)
filter_name_entry.pack(side="left", padx=(0, 15))

ttk.Label(filters_row, text="Линия:").pack(side="left", padx=(0, 5))
filter_line_var = tk.StringVar()
filter_line_entry = ttk.Entry(filters_row, textvariable=filter_line_var, width=10)
filter_line_entry.pack(side="left", padx=(0, 15))

ttk.Label(filters_row, text="Тара:").pack(side="left", padx=(0, 5))
filter_container_var = tk.StringVar()
filter_container_entry = ttk.Entry(filters_row, textvariable=filter_container_var, width=10)
filter_container_entry.pack(side="left", padx=(0, 15))

ttk.Label(filters_row, text="Скорость:").pack(side="left", padx=(0, 5))
filter_speed_var = tk.StringVar()
filter_speed_entry = ttk.Entry(filters_row, textvariable=filter_speed_var, width=10)
filter_speed_entry.pack(side="left", padx=(0, 15))

ttk.Label(filters_row, text="Действие:").pack(side="left", padx=(0, 5))
filter_action_var = tk.StringVar()
filter_action_entry = ttk.Entry(filters_row, textvariable=filter_action_var, width=15)
filter_action_entry.pack(side="left", padx=(0, 15))

# Кнопка сброса фильтров
btn_clear_filters = ttk.Button(filter_frame, text="Очистить фильтры")
btn_clear_filters.pack(side="right", padx=(10, 0))

# Информационная метка о количестве записей
info_frame = ttk.Frame(frm_cat)
info_frame.pack(fill="x", pady=(0, 5))
cat_info_label = ttk.Label(info_frame, text="Всего: 0 | Показано: 0", foreground="#666")
cat_info_label.pack(side="left")

# Таблица каталога
table_frame = ttk.Frame(frm_cat)
table_frame.pack(fill="both", expand=True)

cat_cols = ("name", "line", "container", "speed", "limit", "action")

tree_cat = ttk.Treeview(table_frame, columns=cat_cols, show="headings", selectmode="extended")
tree_cat.heading("name", text="Наименование")
tree_cat.heading("line", text="Линия")
tree_cat.heading("container", text="Тара")
tree_cat.heading("speed", text="Скорость, шт/ч")
tree_cat.heading("limit", text="Предельный объём")
tree_cat.heading("action", text="Действие")

tree_cat.column("name", width=360, anchor="w")
tree_cat.column("line", width=80, anchor="center")
tree_cat.column("container", width=120, anchor="center")
tree_cat.column("speed", width=120, anchor="e")
tree_cat.column("limit", width=140, anchor="e")
tree_cat.column("action", width=180, anchor="w")

vsb_cat = ttk.Scrollbar(table_frame, orient="vertical", command=tree_cat.yview)
hsb_cat = ttk.Scrollbar(table_frame, orient="horizontal", command=tree_cat.xview)
tree_cat.configure(yscrollcommand=vsb_cat.set, xscrollcommand=hsb_cat.set)

tree_cat.grid(row=0, column=0, sticky="nsew")
vsb_cat.grid(row=0, column=1, sticky="ns")
hsb_cat.grid(row=1, column=0, sticky="ew")
table_frame.rowconfigure(0, weight=1)
table_frame.columnconfigure(0, weight=1)
# ===== Функции работы с каталогом ===========================================

def save_catalog_json(path: Optional[str] = None, *, silent: bool = False):
    """Сохраняем таблицу каталога в JSON (список строк, как rows())."""
    _path = path or CATALOG_JSON_PATH
    try:
        rows = CATALOG.rows()
        with open(_path, "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False, indent=2)
        if not silent:
            try:
                log(f"[Каталог] Сохранено: {_path}")
            except Exception:
                pass
    except Exception as e:
        if not silent:
            try:
                log(f"[Каталог] Ошибка сохранения: {e}")
            except Exception:
                pass


def load_catalog_json(path: Optional[str] = None, *, silent: bool = False):
    """Грузим JSON и полностью заменяем содержимое каталога."""
    _path = path or CATALOG_JSON_PATH
    try:
        if not os.path.isfile(_path):
            return
        with open(_path, "r", encoding="utf-8") as f:
            rows = json.load(f)
        if isinstance(rows, list):
            CATALOG.import_rows(rows)
        if not silent:
            try:
                log(f"[Каталог] Загружено: {_path}")
            except Exception:
                pass
    except Exception as e:
        if not silent:
            try:
                log(f"[Каталог] Ошибка загрузки: {e}")
            except Exception:
                pass


def _apply_filters(row):
    """Применить фильтры к строке каталога"""
    # Получаем значения фильтров
    name_filter = filter_name_var.get().strip().lower()
    line_filter = filter_line_var.get().strip().lower()
    container_filter = filter_container_var.get().strip().lower()
    speed_filter = filter_speed_var.get().strip().lower()
    action_filter = filter_action_var.get().strip().lower()
    
    # Проверяем каждый фильтр
    if name_filter:
        name = str(row.get("name", "")).lower()
        if name_filter not in name:
            return False
    
    if line_filter:
        line = str(row.get("line", "")).lower()
        if line_filter not in line:
            return False
    
    if container_filter:
        container = str(row.get("container", "")).lower()
        if container_filter not in container:
            return False
    
    if speed_filter:
        speed = "" if row.get("speed") is None else str(row.get("speed"))
        if speed_filter not in speed.lower():
            return False
    
    if action_filter:
        action = str(row.get("action", "")).lower()
        if action_filter not in action:
            return False
    
    return True


def _cat_refresh():
    """Обновить отображение таблицы каталога с учетом фильтров"""
    for i in tree_cat.get_children():
        tree_cat.delete(i)
    try:
        rows = CATALOG.rows()
    except Exception:
        rows = []
    
    filtered_count = 0
    for r in rows:
        # Применяем фильтры
        if not _apply_filters(r):
            continue
        
        vals = (
            r.get("name", ""),
            r.get("line", ""),
            r.get("container", ""),
            ("" if r.get("speed") is None else str(r.get("speed"))),
            ("" if r.get("limit") is None else str(r.get("limit"))),
            r.get("action", ""),
        )
        tree_cat.insert("", "end", values=vals)
        filtered_count += 1
    
    enable_tree_sort(tree_cat)
    
    # Обновляем информацию о количестве записей
    try:
        total = len(rows)
        cat_info_label.config(text=f"Всего: {total} | Показано: {filtered_count}")
    except Exception:
        pass


def _clear_filters():
    """Очистить все фильтры"""
    filter_name_var.set("")
    filter_line_var.set("")
    filter_container_var.set("")
    filter_speed_var.set("")
    filter_action_var.set("")
    _cat_refresh()


# Загрузка каталога при старте
try:
    load_catalog_json(silent=True)
except Exception:
    pass
_cat_refresh()



# ===== Редактирование по двойному клику/Enter ================================

_edit_entry = None
_edit_item = None
_edit_col = None


def _cell_bbox_cat(item, col):
    """Получить координаты ячейки для редактирования"""
    try:
        bx = tree_cat.bbox(item, col)
        return bx if bx else None
    except Exception:
        return None


def _start_edit_cat(event):
    """Начать редактирование ячейки по двойному клику"""
    global _edit_entry, _edit_item, _edit_col
    region = tree_cat.identify("region", event.x, event.y)
    if region != "cell":
        return
    col = tree_cat.identify_column(event.x)   # "#1"…
    row = tree_cat.identify_row(event.y)
    if not row or not col:
        return
    bbox = _cell_bbox_cat(row, col)
    if not bbox:
        return
    x, y, w, h = bbox
    value = tree_cat.set(row, tree_cat["columns"][int(col[1:])-1])
    _edit_item, _edit_col = row, col
    _edit_entry = tk.Entry(tree_cat)
    _edit_entry.insert(0, value)
    _edit_entry.select_range(0, "end")
    _edit_entry.focus_set()
    _edit_entry.place(x=x, y=y, width=w, height=h)
    _edit_entry.bind("<Return>", _commit_edit_cat)
    _edit_entry.bind("<Escape>", _cancel_edit_cat)
    _edit_entry.bind("<FocusOut>", _commit_edit_cat)


def _cancel_edit_cat(event=None):
    """Отменить редактирование"""
    global _edit_entry, _edit_item, _edit_col
    if _edit_entry:
        _edit_entry.destroy()
    _edit_entry = _edit_item = _edit_col = None


def _float_or_none(x: str):
    """Преобразовать строку в float или None"""
    x = str(x).strip()
    if x == "":
        return None
    x = x.replace(" ", "").replace(",", ".")
    try:
        return float(x)
    except Exception:
        return None


def _apply_row_to_catalog(item_id):
    """Применить изменения строки к каталогу"""
    values = tree_cat.item(item_id, "values")
    if not values:
        return
    name, line, container, speed, limit, action = (values + ("", "", "", "", "", ""))[:6]
    try:
        CATALOG.upsert(
            name=name,
            line=line,
            container=container,
            speed=_float_or_none(speed),
            limit=_float_or_none(limit),
            action=action,
        )
        save_catalog_json(silent=True)
    except Exception:
        pass

def _commit_edit_cat(event=None):
    """Сохранить изменения ячейки"""
    global _edit_entry, _edit_item, _edit_col
    if not _edit_entry:
        return
    val = _edit_entry.get()
    col_idx = int(_edit_col[1:]) - 1
    col_name = tree_cat["columns"][col_idx]
    tree_cat.set(_edit_item, col_name, val)
    _apply_row_to_catalog(_edit_item)
    _cancel_edit_cat()
    # Обновить таблицу с учетом фильтров
    _cat_refresh()


def _add_row_cat(event=None):
    """Добавить пустую строку и сразу перейти в редактирование первой ячейки"""
    iid = tree_cat.insert("", "end", values=("", "", "", "", "", ""))
    tree_cat.see(iid)
    tree_cat.selection_set(iid)
    # Синхронно создаём пустую запись в каталоге — по факту заполнится при коммите
    _apply_row_to_catalog(iid)
    # Старт редактирования первой колонки
    bbox = _cell_bbox_cat(iid, "#1")
    if bbox:
        x, y, w, h = bbox
        global _edit_entry, _edit_item, _edit_col
        _edit_item, _edit_col = iid, "#1"
        _edit_entry = tk.Entry(tree_cat)
        _edit_entry.insert(0, "")
        _edit_entry.select_range(0, "end")
        _edit_entry.focus_set()
        _edit_entry.place(x=x, y=y, width=w, height=h)
        _edit_entry.bind("<Return>", _commit_edit_cat)
        _edit_entry.bind("<Escape>", _cancel_edit_cat)
        _edit_entry.bind("<FocusOut>", _commit_edit_cat)


def _delete_selected_cat(event=None):
    """Удалить выбранные строки из грида и каталога"""
    sels = tree_cat.selection()
    if not sels:
        return
    for iid in sels:
        vals = tree_cat.item(iid, "values") or ()
        # Удаляем из каталога, если есть ключи name/line
        if len(vals) >= 2:
            name, line = vals[0], vals[1]
            nm = CATALOG.normalize_name(name or "")
            ln = CATALOG._canon_line(line or "")
            # Убрать точечные скорости
            try:
                if (nm, ln) in CATALOG.product_speeds:
                    del CATALOG.product_speeds[(nm, ln)]
            except Exception:
                pass
            # Убрать мета
            try:
                if (nm, ln) in CATALOG.product_meta:
                    del CATALOG.product_meta[(nm, ln)]
            except Exception:
                pass
        tree_cat.delete(iid)
    save_catalog_json(silent=True)
    _cat_refresh()


def _paste_catalog(event=None):
    """Вставка из буфера обмена (TSV/CSV с заголовком или без)"""
    try:
        raw = tree_cat.clipboard_get()
    except Exception:
        return "break"
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")
    lines = [ln for ln in raw.split("\n") if ln.strip() != ""]
    if not lines:
        return "break"

    header_like = ["наименование", "линия", "тара", "скорость", "предельный", "действие"]
    has_header = any(h in lines[0].lower() for h in header_like)
    rows = lines[1:] if has_header else lines

    import re as _re
    for ln in rows:
        parts = [p.strip() for p in _re.split(r"\t|;", ln)]
        if not parts or all(p == "" for p in parts):
            continue
        name = parts[0] if len(parts) > 0 else ""
        line = parts[1] if len(parts) > 1 else ""
        container = parts[2] if len(parts) > 2 else ""
        speed = parts[3] if len(parts) > 3 else ""
        limit = parts[4] if len(parts) > 4 else ""
        action = parts[5] if len(parts) > 5 else ""
        iid = tree_cat.insert("", "end", values=(name, line, container, speed, limit, action))
        _apply_row_to_catalog(iid)
    save_catalog_json(silent=True)
    _cat_refresh()
    return "break"


# ===== Привязка событий ======================================================

tree_cat.bind("<Double-1>", _start_edit_cat)      # Двойной клик — редактировать
tree_cat.bind("<Return>", _start_edit_cat)        # Enter — редактировать
tree_cat.bind("<Insert>", _add_row_cat)           # Insert — добавить строку
tree_cat.bind("<Delete>", _delete_selected_cat)   # Delete — удалить
tree_cat.bind("<Control-n>", _add_row_cat)        # Ctrl+N — добавить
tree_cat.bind("<Control-v>", _paste_catalog)      # Ctrl+V — вставить

# Настройка команд кнопок
btn_add_cat.configure(command=_add_row_cat)
btn_del_cat.configure(command=_delete_selected_cat)
btn_save_cat.configure(command=lambda: (save_catalog_json(), _cat_refresh()))
btn_load_cat.configure(command=lambda: (load_catalog_json(), _cat_refresh()))
btn_clear_filters.configure(command=_clear_filters)

# Привязка событий фильтров (обновление при изменении)
def _on_filter_change(*args):
    """Обработчик изменения фильтров"""
    _cat_refresh()

filter_name_var.trace('w', _on_filter_change)
filter_line_var.trace('w', _on_filter_change)
filter_container_var.trace('w', _on_filter_change)
filter_speed_var.trace('w', _on_filter_change)
filter_action_var.trace('w', _on_filter_change)
# === Вкладка «Матрицы» ====================================================
try:
    from gui_matrix import show_matrix_tab
except Exception:
    show_matrix_tab = None

try:
    if show_matrix_tab is not None:
        show_matrix_tab(nb, catalog=CATALOG)
except Exception as e:
    print(f"[Матрицы] Ошибка при создании вкладки: {e}")


# (можно будет добавить превью/таблицу, но пока не нужно)



# (можно будет добавить превью/таблицу, но пока не нужно)



def _rebuild_filter_menus():
    try:
        lbl_filters.config(text=_sel_to_human())
    except Exception:
        pass
    try:
        render_oee_matrix()
    except Exception:
        pass
    try:
        render_report_table()
    except Exception:
        pass
    # новое:
    try:
        show_block(_summary_option)
    except Exception:
        pass





# таблица + скроллы
frm_table = ttk.Frame(tab_table)
frm_table.pack(fill="both", expand=True, padx=8, pady=8)

tree = ttk.Treeview(frm_table, show="headings")
vsb = ttk.Scrollbar(frm_table, orient="vertical", command=tree.yview)
hsb = ttk.Scrollbar(frm_table, orient="horizontal", command=tree.xview)

tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)


def on_row_select(_evt=None):
    if combo_block.get() != _summary_option:
        combo_events.configure(state="disabled", values=[])
        lbl_ev_count.config(text="")
        return
    sel = tree.selection()
    if not sel:
        combo_events.configure(state="disabled", values=[])
        lbl_ev_count.config(text="")
        return
    vals = tree.item(sel[0], "values")
    if not vals:
        combo_events.configure(state="disabled", values=[])
        lbl_ev_count.config(text="")
        return
    prod_name = str(vals[0]).strip()
    day_label = str(vals[2]).strip() if len(vals) > 2 else ""
    shift_label = str(vals[3]).strip() if len(vals) > 3 else ""
    key = (_norm_name(prod_name), day_label, shift_label)
    events = DOWNTIME_BY.get(key, [])
    if not events:
        combo_events.configure(state="disabled", values=["— нет событий —"])
        combo_events.set("— нет событий —")
        lbl_ev_count.config(text="0")
        return
    items = [_fmt_event_row(ev) for ev in events]
    combo_events.configure(state="readonly", values=items)
    combo_events.set(items[0])
    lbl_ev_count.config(text=str(len(items)))
    update_quick_subtotal()

def update_quick_subtotal():
    """Считает подитог по выделению в основной таблице `tree`."""
    sels = tree.selection()
    if not sels:
        try:
            lbl_quick.config(text="Подитог: ничего не выбрано")
        except Exception:
            pass
        return

    cols = list(tree["columns"])
    # Ищем нужные поля по именам заголовков (работает и в «Сводке», и в «Отчётной»)
    hdr_plan  = _col_idx_by_header(tree, "Потолок (шт)") or _col_idx_by_header(tree, "План, шт")
    hdr_fact  = _col_idx_by_header(tree, "Факт (шт)")
    hdr_eff   = _col_idx_by_header(tree, "EffMin (мин)")
    hdr_plndt = _col_idx_by_header(tree, "План. простой (мин)")
    hdr_downt = _col_idx_by_header(tree, "Σ простоев (мин)")
    hdr_oee   = _col_idx_by_header(tree, "OEE, %")

    n = len(sels)
    sum_plan = sum_fact = sum_eff = sum_plndt = sum_downt = 0.0
    oee_sum = 0.0
    oee_wsum = 0.0  # взвешиваем по Плану, если он есть

    for iid in sels:
        vals = list(tree.item(iid, "values") or [])
        # суммируем то, что нашли
        if hdr_plan is not None and hdr_plan < len(vals):
            v = _as_float(vals[hdr_plan]);   sum_plan  += (v or 0.0)
        if hdr_fact is not None and hdr_fact < len(vals):
            v = _as_float(vals[hdr_fact]);   sum_fact  += (v or 0.0)
        if hdr_eff is not None and hdr_eff < len(vals):
            v = _as_float(vals[hdr_eff]);    sum_eff   += (v or 0.0)
        if hdr_plndt is not None and hdr_plndt < len(vals):
            v = _as_float(vals[hdr_plndt]);  sum_plndt += (v or 0.0)
        if hdr_downt is not None and hdr_downt < len(vals):
            v = _as_float(vals[hdr_downt]);  sum_downt += (v or 0.0)

        if hdr_oee is not None and hdr_oee < len(vals):
            o = _as_float(vals[hdr_oee])
            if o is not None:
                oee_sum += o
                # если есть План — используем его как вес
                w = 0.0
                if hdr_plan is not None and hdr_plan < len(vals):
                    wv = _as_float(vals[hdr_plan]); w = (wv or 0.0)
                oee_wsum += (o * (w if w > 0 else 1.0))

    # средний OEE: взвешенно по Плану (если суммарный план > 0), иначе простое среднее
    oee_avg = None
    if hdr_oee is not None:
        if sum_plan > 0:
            oee_avg = oee_wsum / (sum_plan if sum_plan > 0 else 1.0)
        elif n > 0:
            oee_avg = oee_sum / n

    parts = [f"строк: {n}"]
    if hdr_plan  is not None:  parts.append(f"План: {int(round(sum_plan))}")
    if hdr_fact  is not None:  parts.append(f"Факт: {int(round(sum_fact))}")
    if hdr_eff   is not None:  parts.append(f"EffMin: {int(round(sum_eff))}")
    if hdr_plndt is not None:  parts.append(f"План.простой: {int(round(sum_plndt))}")
    if hdr_downt is not None:  parts.append(f"Σ простоев: {int(round(sum_downt))}")
    if oee_avg   is not None:  parts.append(f"OEE≈ {oee_avg:.1f}%")

    try:
        lbl_quick.config(text="Подитог: " + " • ".join(parts))
    except Exception:
        pass

tree.bind("<<TreeviewSelect>>", lambda e: (on_row_select(e), update_quick_subtotal()))


tree.grid(row=0, column=0, sticky="nsew")
vsb.grid(row=0, column=1, sticky="ns")
hsb.grid(row=1, column=0, sticky="ew")
frm_table.rowconfigure(0, weight=1)
frm_table.columnconfigure(0, weight=1)
# === Быстрый подитог (для выделенных строк) =================================
frm_quick = ttk.Frame(tab_table)
frm_quick.pack(fill="x", padx=8, pady=(0, 8))

lbl_quick = ttk.Label(frm_quick, text="Подитог: ничего не выбрано", anchor="w")
lbl_quick.pack(side="left", fill="x", expand=True)


# === OEE-матрица: UI (вкладка) ==============================================
# Панель кнопок
oee_toolbar = ttk.Frame(tab_oee)
oee_toolbar.pack(fill="x", padx=8, pady=(8, 4))


def _oee_refresh():
    try:
        render_oee_matrix()
    except Exception as e:
        log(f"[OEE] Ошибка отрисовки матрицы: {e}")


ttk.Button(oee_toolbar, text="Обновить", command=_oee_refresh).pack(side="left")

# Прокручиваемая область с сеткой
oee_wrap = ttk.Frame(tab_oee)
oee_wrap.pack(fill="both", expand=True, padx=8, pady=(0, 8))

oee_canvas = tk.Canvas(oee_wrap, highlightthickness=0)
oee_vsb = ttk.Scrollbar(oee_wrap, orient="vertical", command=oee_canvas.yview)
oee_hsb = ttk.Scrollbar(oee_wrap, orient="horizontal", command=oee_canvas.xview)
oee_inner = ttk.Frame(oee_canvas)

oee_inner.bind(
    "<Configure>", lambda e: oee_canvas.configure(scrollregion=oee_canvas.bbox("all"))
)
oee_canvas.create_window((0, 0), window=oee_inner, anchor="nw")
oee_canvas.configure(yscrollcommand=oee_vsb.set, xscrollcommand=oee_hsb.set)

oee_canvas.grid(row=0, column=0, sticky="nsew")
oee_vsb.grid(row=0, column=1, sticky="ns")
oee_hsb.grid(row=1, column=0, sticky="ew")
oee_wrap.rowconfigure(0, weight=1)
oee_wrap.columnconfigure(0, weight=1)


# ... конец настройки вкладок/таблиц/OEE ...

# --- ЛОГ (внизу, над прогресс-баром)
from log_ui import create_log_panel, log
frm_log, _ = create_log_panel(root, height=6)
frm_log.pack(fill="both", expand=False, padx=8, pady=(0, 8))

# ===== ПРОГРЕСС-БАР и ОТМЕНА ================================================
frm_prog = ttk.Frame(root)

frm_prog = ttk.Frame(root)
frm_prog.pack(fill="x", padx=8, pady=(0, 8))
prg = ttk.Progressbar(frm_prog, mode="determinate", maximum=100)
prg.pack(side="left", fill="x", expand=True)
lbl_prog = ttk.Label(frm_prog, text="")
lbl_prog.pack(side="left", padx=8)
btn_cancel = ttk.Button(frm_prog, text="Отмена", state="disabled")
btn_cancel.pack(side="left")
# ===== НАСТРОЙКИ/ПАМЯТЬ (последний JSON) ====================================
SETTINGS_PATH = os.path.join(os.path.dirname(__file__), "gui_blocks.settings.json")
_SETTINGS = {}

def _settings_load():
    global _SETTINGS
    try:
        if os.path.isfile(SETTINGS_PATH):
            with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
                _SETTINGS = json.load(f) or {}
        else:
            _SETTINGS = {}
    except Exception:
        _SETTINGS = {}

def _settings_save():
    try:
        with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
            json.dump(_SETTINGS, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _remember_last_json(path: str):
    if not path or not str(path).lower().endswith(".json"):
        return
    _SETTINGS["last_json_path"] = path
    _settings_save()

def _get_last_json() -> str:
    p = _SETTINGS.get("last_json_path", "")
    return p if (isinstance(p, str) and os.path.isfile(p)) else ""


# ===== ПОМОЩНИКИ ФОРМАТИРОВАНИЯ =============================================
def _fmt_event_row(ev: List) -> str:
    try:
        name = str(ev[D2_COL_NAME]).strip() if len(ev) > D2_COL_NAME else ""
        reason = str(ev[D2_COL_REASON]).strip() if len(ev) > D2_COL_REASON else ""
        kind = str(ev[D2_COL_KIND]).strip() if len(ev) > D2_COL_KIND else ""
        beg = str(ev[D2_COL_BEG]).strip() if len(ev) > D2_COL_BEG else ""
        end = str(ev[D2_COL_END]).strip() if len(ev) > D2_COL_END else ""
        mins = _safe_minutes(ev[D2_COL_MIN] if len(ev) > D2_COL_MIN else 0)
        desc = ""
        if len(ev) > D2_COL_DESC:
            desc = str(ev[D2_COL_DESC]).strip()
        tag = f" [{kind}]" if kind else ""
        # при наличии расшифровки — добавляем её в конец
        base = f"{beg}–{end} • {mins} мин"
        if reason:
            base += f" • {reason}{tag}"
        if desc:
            base += f" — {desc}"
        return base
    except Exception:
        return str(ev)
    
def _top3_for(name: str, day_label: str, shift_label: str):
    return core_top3_for(DOWNTIME_BY, name, day_label, shift_label)


def _fmt_top_item(item: dict) -> str:
    return core_fmt_top_item(item)




def _to_float(x):
    """Безопасное преобразование в float"""
    try:
        return float(str(x).replace(",", "."))
    except Exception:
        return None

def _as_float(s):
    """Безопасное преобразование в float с очисткой"""
    try:
        return float(str(s).replace(" ", "").replace("\u00A0", "").replace(",", "."))
    except Exception:
        return None

def _row_speed_from_products(row: list) -> Optional[float]:
    """Получение скорости из строки продуктов"""
    if len(row) > 4:
        f = _to_float(row[4])
        if f is not None and f > 0:
            return f
    return None

def _fmt_cell(v):
    """Форматирование ячейки"""
    try:
        if isinstance(v, numbers.Number):
            return str(int(round(float(v))))
    except Exception:
        pass
    return str(v)

def _col_idx_by_header(tv, header_name: str) -> Optional[int]:
    """Возвращает индекс колонки по видимому заголовку (или None)."""
    cols = list(tv["columns"])
    for i, cid in enumerate(cols):
        if str(tv.heading(cid).get("text", "")).strip().lower() == header_name.strip().lower():
            return i
    return None


def _round_if_needed(v, tol=1e-9):
    try:
        if isinstance(v, float) and math.isclose(v, round(v), abs_tol=tol):
            return int(round(v))
    except Exception:
        pass
    return v

def _read_from_treeview(tv):
    """Возвращает (headers, rows) из ttk.Treeview.
       headers — подписи колонок (tv.heading(col)['text']),
       rows — значения (в том же порядке)."""
    cols = list(tv["columns"])
    headers = [tv.heading(c)["text"] for c in cols]
    rows = [tv.item(i, "values") for i in tv.get_children("")]
    return headers, rows

def _compute_export_tables_from_summary(headers_sum, rows_sum):
    """Строит набор таблиц для экспорта и предпросмотра из сводных данных"""
    idx = {h: i for i, h in enumerate(headers_sum)}

    def _to_int(x) -> int:
        if x is None or x == "":
            return 0
        s = str(x).replace(" ", "").replace("\u00A0", "")
        try:
            return int(s.replace(",", ""))
        except Exception:
            try:
                return int(float(s.replace(",", ".")))
            except Exception:
                return 0

    # По продуктам
    hdr_prod = ["Линия", "Смена", "Продукт", "План", "Факт", "%", "Информация"]
    rows_prod = []
    for r in rows_sum:
        line = str(r[idx["Линия"]]).strip()
        shift = str(r[idx["Смена"]]).strip()
        product = str(r[idx["Продукт"]]).strip()
        plan = _to_int(r[idx["Потолок (шт)"]])
        fact = _to_int(r[idx["Факт (шт)"]])
        pct = (fact / plan * 100.0) if plan else 0.0
        top3 = _top3_for(r[idx["Продукт"]], r[idx["День"]], r[idx["Смена"]])
        info = "\n".join(_fmt_top_item(t) for t in top3[:3])
        rows_prod.append([line, shift, product, plan, fact, round(pct, 1), info])
    rows_prod.sort(key=lambda x: (x[0], x[1], x[2]))

    # По линиям
    hdr_lines = ["Линия", "План", "Факт", "%", "Информация"]
    totals = {}
    for r in rows_sum:
        line = str(r[idx["Линия"]]).strip()
        plan = _to_int(r[idx["Потолок (шт)"]])
        fact = _to_int(r[idx["Факт (шт)"]])
        d = totals.setdefault(line, {"plan": 0, "fact": 0})
        d["plan"] += plan
        d["fact"] += fact

    def _numkey(s):
        ss = str(s)
        return (0, int(ss)) if ss.isdigit() else (1, ss)

    rows_lines = []
    for line in sorted(totals.keys(), key=_numkey):
        plan = totals[line]["plan"]
        fact = totals[line]["fact"]
        pct  = (fact / plan * 100.0) if plan else 0.0
        info = "\n".join(_top3_for_line(rows_sum, idx, line))
        rows_lines.append([line, plan, fact, round(pct, 1), info])

    return hdr_lines, rows_lines, hdr_prod, rows_prod

# ====== Индексация простоев + сводка / OEE ==================================
B1_COL_NAME = 0
B1_COL_BEG = 1
B1_COL_END = 2
B1_COL_DUR = 3

D2_COL_NAME = 0
D2_COL_REASON = 1
D2_COL_KIND = 2
D2_COL_BEG = 3
D2_COL_END = 4
D2_COL_MIN = 5
D2_COL_DESC = 6  # будет добавляться программно в конец
UNACC_TAG = "__SYNTH_UNACCOUNTED__"  # маркер синтетического события «неучтёнка»



def _is_planned(kind: str, reason: str = "") -> bool:
    def norm(s: str) -> str:
        return str(s or "").lower().replace("ё", "е")

    s = norm(kind) + " " + norm(reason)
    if "неплан" in s:
        return False
    return "план" in s


DEFAULT_SPEED_BY_LINE: dict[str, float] = {}


def get_nominal_speed(line: str, product_name: str) -> Optional[float]:
    """
    1) пытаемся взять скорость из каталога (product+line)
    2) fallback — дефолт по линии
    """
    # синхронизируем дефолты каталога (если вы где-то заполнили DEFAULT_SPEED_BY_LINE)
    try:
        CATALOG.set_line_defaults(DEFAULT_SPEED_BY_LINE)
    except Exception:
        pass

    ln = (line or "").strip()
    nm = (product_name or "").strip()
    s = CATALOG.speed(ln, nm)
    if s is not None:
        return s
    return DEFAULT_SPEED_BY_LINE.get(ln)



def _extract_fact_qty(row: list) -> Optional[int]:
    try:
        for v in reversed(row):
            try:
                f = float(v)
                if math.isfinite(f):
                    return int(round(f))
            except Exception:
                continue
    except Exception:
        pass
    return None
def inject_unaccounted_time():
    """
    Считает «неучтёнку» для каждого (продукт, день, смена) и добавляет её в DOWNTIME_BY.
    Формула (мин):
      T_unacc = max( RunMin - AllDowntimes - FactMin , 0 )
    где:
      RunMin      — длит. запуска из «Продуктов» (B1_COL_DUR, мин)
      AllDowntimes— сумма ВСЕХ простоев из DOWNTIME_BY (мин)
      FactMin     — (Факт / скорость) * 60
    """
    if not PRODUCT_BLOCKS:
        return

    # подчистим прежние синтетические записи
    for key in list(DOWNTIME_BY.keys()):
        events = DOWNTIME_BY.get(key, [])
        DOWNTIME_BY[key] = [
            ev for ev in events
            if not (len(ev) > D2_COL_DESC and "Неучтенное время" in str(ev[D2_COL_DESC]))
        ]

    injected = 0

    for blk_name in PRODUCT_BLOCKS:
        blk = DATA.get(blk_name)
        if not blk:
            continue

        meta = blk.get("meta", {})
        day_label   = str(meta.get("sheet", ""))
        shift_label = str(meta.get("shift", ""))
        line_label  = str(meta.get("line", ""))

        for row in blk["array"]:
            if not row or len(row) <= B1_COL_NAME:
                continue
            raw_name = str(row[B1_COL_NAME]).strip()
            name = CATALOG.normalize_name(raw_name)
            if not name or name == "0":
                continue

            run_min = _safe_minutes(row[B1_COL_DUR] if len(row) > B1_COL_DUR else 0)
            if run_min <= 0:
                continue

            fact_qty = _extract_fact_qty(row)
            if fact_qty is None:
                continue

            speed = _row_speed_from_products(row)
            if speed is None:
                speed = get_nominal_speed(line_label, name)
            if speed is None or speed <= 0:
                continue

            key = (_norm_name(name), day_label, shift_label)

            # БЕРЁМ ВСЕ простои без фильтра «план/не план»
            all_dt = 0
            for ev in DOWNTIME_BY.get(key, []):
                all_dt += _safe_minutes(ev[D2_COL_MIN] if len(ev) > D2_COL_MIN else 0)

            fact_min = (float(fact_qty) / float(speed)) * 60.0

            unacc = int(round(run_min - all_dt - fact_min))
            if unacc > 0:
                synth = [
                    name,                 # D2_COL_NAME
                    "Неучтенное время",   # D2_COL_REASON (категория)
                    "",                   # D2_COL_KIND — пусто, чтобы не было [..]
                    "",                   # D2_COL_BEG
                    "",                   # D2_COL_END
                    unacc,                # D2_COL_MIN
                    "",                   # D2_COL_DESC — пусто, чтобы не было "— …"
                ]
                DOWNTIME_BY.setdefault(key, []).append(synth)
                injected += 1


    try:
        log(f"[Неучтёнка] Добавлено синтетических событий: {injected}")
    except Exception:
        pass



DOWNTIME_BY: Dict[Tuple[str, str, str], List[List]] = {}
AGG_BY: Dict[Tuple[str, str, str], Dict] = {}


def _norm_name(s: str) -> str:
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s.replace("\xa0", " "))
    s = s.replace("«", '"').replace("»", '"').replace("“", '"').replace("”", '"')
    return s


PRODUCT_BLOCKS: List[str] = []
DOWNTIME_BLOCKS: List[str] = []
ALL_BLOCKS: List[str] = []


def _safe_minutes(x) -> int:
    try:
        if isinstance(x, str) and x.strip() == "":
            return 0
        return int(round(float(x)))
    except Exception:
        return 0
def _top3_for_line(rows_sum, idx, line_label: str) -> list[str]:
    """Агрегированный топ-3 причин по всей линии, с разрезом по сменам."""
    agg: dict[str, dict] = {}   # reason -> {total, day, night, kind, desc}

    for r in rows_sum:
        if str(r[idx["Линия"]]).strip() != str(line_label).strip():
            continue
        name_norm  = _norm_name(r[idx["Продукт"]])
        day_label  = str(r[idx["День"]]).strip()
        shift_lbl  = str(r[idx["Смена"]]).strip()  # "День"/"Ночь"
        key = (name_norm, day_label, shift_lbl)

        for ev in DOWNTIME_BY.get(key, []):
            mins   = _safe_minutes(ev[D2_COL_MIN] if len(ev) > D2_COL_MIN else 0)
            reason = str(ev[D2_COL_REASON]).strip() if len(ev) > D2_COL_REASON else ""
            kind   = str(ev[D2_COL_KIND]).strip()   if len(ev) > D2_COL_KIND   else ""
            desc   = str(ev[D2_COL_DESC]).strip()   if len(ev) > D2_COL_DESC   else ""

            # Берём только неплановые, как и раньше
            if _is_planned(kind, reason):
                continue
            if not reason and mins <= 0 and not desc:
                continue

            d = agg.setdefault(reason, {"total": 0, "day": 0, "night": 0, "kind": "", "desc": ""})
            d["total"] += mins
            if shift_lbl == "День":
                d["day"]   += mins
            elif shift_lbl == "Ночь":
                d["night"] += mins
            if not d["kind"] and kind:
                d["kind"] = kind
            if not d["desc"] and desc:
                d["desc"] = desc

    # топ-3 по сумме минут
    top = sorted(agg.items(), key=lambda kv: kv[1]["total"], reverse=True)[:3]
    out = []
    for reason, d in top:
        total = d["total"]
        if total <= 0 and not reason:
            continue
        tag = f" [{d['kind']}]" if d["kind"] else ""
        # хвост со сменами — только ненулевые части
        parts = []
        if d["day"]   > 0: parts.append(f"День {d['day']}")
        if d["night"] > 0: parts.append(f"Ночь {d['night']}")
        shifts = f" ({' / '.join(parts)})" if parts else ""
        s = f"{total} мин • {reason}{tag}{shifts}"
        if d["desc"]:
            s += f" — {d['desc']}"
        out.append(s)
    return out



def _is_blank_time(x) -> bool:
    s = str(x).strip()
    return s in ("", "0", "00:00:00", "0:00:00")


def build_downtime_index():
    global DOWNTIME_BY, AGG_BY
    DOWNTIME_BY, AGG_BY = core_build_downtime_index(DATA, DOWNTIME_BLOCKS)



def build_summary_rows() -> Tuple[List[str], List[List]]:
    flt = FilterOpts(
        selected_lines=set(SELECTED_LINES),
        selected_days=set(SELECTED_DAYS),
        current_line="Все",
    )
    return core_build_summary_rows(DATA, PRODUCT_BLOCKS, DOWNTIME_BY, DEFAULT_SPEED_BY_LINE, flt)



def _top3_reasons_with_desc(name_norm: str, day_label: str, shift_label: str) -> list[str]:
    """
    Берём все события простоев для ключа (name_norm, day, shift),
    агрегируем ПО НЕПЛАНОВЫМ простоям: сумма минут по 'reason' (B),
    запоминаем 'kind' (C) и первую непустую расшифровку (G),
    сортируем по убыванию минут и возвращаем три строки.
    Формат: 'NN мин • {reason} [kind] — desc'
    """
    key = (name_norm, day_label, shift_label)
    events = DOWNTIME_BY.get(key, [])
    if not events:
        return []

    agg = {}       # reason -> minutes
    kind_map = {}  # reason -> kind
    desc_map = {}  # reason -> first non-empty desc

    for ev in events:
        mins = _safe_minutes(ev[D2_COL_MIN] if len(ev) > D2_COL_MIN else 0)
        reason = str(ev[D2_COL_REASON]).strip() if len(ev) > D2_COL_REASON else ""
        kind = str(ev[D2_COL_KIND]).strip() if len(ev) > D2_COL_KIND else ""
        desc = str(ev[D2_COL_DESC]).strip() if len(ev) > D2_COL_DESC else ""

        # только НЕплановые в топ-3
        if _is_planned(kind, reason):
            continue

        if not reason and mins <= 0 and not desc:
            continue

        agg[reason] = agg.get(reason, 0) + mins
        if reason not in kind_map and kind:
            kind_map[reason] = kind
        if reason not in desc_map and desc:
            desc_map[reason] = desc

    top = sorted(agg.items(), key=lambda kv: kv[1], reverse=True)[:3]
    out = []
    for reason, total_min in top:
        if total_min <= 0 and not reason:
            continue
        k = kind_map.get(reason, "")
        d = desc_map.get(reason, "")
        tag = f" [{k}]" if k else ""
        s = f"{total_min} мин • {reason}{tag}"
        if d:
            s += f" — {d}"
        out.append(s)
    return out


def build_report_rows() -> tuple[list[str], list[list]]:
    """
    Формирует строки отчёта:
      Продукт | Линия | День | Смена | План, шт | Факт, шт | OEE, % | Топ-1 | Топ-2 | Топ-3
    План = EffMin * Номинальная_скорость / 60, где EffMin = Длит - Плановые простои.
    Скорость — из 'Продуктов' (колонка E), иначе дефолт по линии.
    """
    headers = [
        "Продукт", "Линия", "День", "Смена",
        "План, шт", "Факт, шт", "OEE, %",
        "Топ-1", "Топ-2", "Топ-3",
    ]
    rows: list[list] = []

    

    for blk_name in PRODUCT_BLOCKS:
        blk = DATA.get(blk_name)
        if not blk:
            continue

        meta = blk.get("meta", {})
        day_label = str(meta.get("sheet", ""))
        shift_label = str(meta.get("shift", ""))
        line_label = str(meta.get("line", ""))

        # фильтры
        if SELECTED_LINES and line_label not in SELECTED_LINES:
            continue
        if SELECTED_DAYS and day_label not in SELECTED_DAYS:
            continue

        for r in blk["array"]:
            if not r or len(r) <= B1_COL_NAME:
                continue
            raw_name = str(r[B1_COL_NAME]).strip()
            name = CATALOG.normalize_name(raw_name)
            if not name or name == "0":
                continue


            beg = r[B1_COL_BEG] if len(r) > B1_COL_BEG else ""
            end = r[B1_COL_END] if len(r) > B1_COL_END else ""
            if _is_blank_time(beg) or _is_blank_time(end):
                continue

            # Длит и EffMin
            run_min = _safe_minutes(r[B1_COL_DUR] if len(r) > B1_COL_DUR else 0)
            key_ev = (_norm_name(name), day_label, shift_label)
            events = DOWNTIME_BY.get(key_ev, [])
            planned_dt = 0
            for ev in events:
                m = _safe_minutes(ev[D2_COL_MIN] if len(ev) > D2_COL_MIN else 0)
                reason = str(ev[D2_COL_REASON]).strip() if len(ev) > D2_COL_REASON else ""
                kind = str(ev[D2_COL_KIND]).strip() if len(ev) > D2_COL_KIND else ""
                if _is_planned(kind, reason):
                    planned_dt += m
            eff_min = max(run_min - planned_dt, 0)
            if eff_min <= 0:
                continue

            # Скорость
            speed = _row_speed_from_products(r)
            if speed is None:
                speed = get_nominal_speed(line_label, name)
            if speed is None or speed <= 0:
                continue

            plan_qty = int(round(eff_min * (speed / 60.0)))
            fact_qty = _extract_fact_qty(r)
            if fact_qty is None:
                continue

            oee_pct = (fact_qty / plan_qty * 100.0) if plan_qty > 0 else None

            # Топ-3 неплановых с расшифровкой
            top3 = _top3_reasons_with_desc(_norm_name(name), day_label, shift_label)
            top1 = top3[0] if len(top3) > 0 else ""
            top2 = top3[1] if len(top3) > 1 else ""
            top3s = top3[2] if len(top3) > 2 else ""

            rows.append([
                name, line_label, day_label, shift_label,
                plan_qty, int(fact_qty),
                (f"{oee_pct:.1f}" if (oee_pct is not None) else ""),
                top1, top2, top3s,
            ])
    return headers, rows


def render_report_table():
    """Рисуем таблицу отчёта во вкладке."""
    # очистка колонок/строк
    for col in report_tree["columns"]:
        report_tree.heading(col, text="")
    report_tree.delete(*report_tree.get_children())

    headers, rows = build_report_rows()
    report_tree["columns"] = [f"r{i}" for i in range(len(headers))]

    # ширины под содержимое
    col_widths = [240, 60, 70, 70, 110, 110, 80, 360, 360, 360]
    for i, h in enumerate(headers):
        report_tree.heading(f"r{i}", text=str(h))
        w = col_widths[i] if i < len(col_widths) else 120
        report_tree.column(f"r{i}", width=w, anchor="w")

    for r in rows:
        report_tree.insert("", "end", values=[_fmt_cell(x) for x in r])
            # включаем клик-сортировку колонок отчёта
    enable_tree_sort(report_tree)
        # редактирование + коп/вставка в отчётной таблице (делаем один раз)
    if not getattr(report_tree, "_editing_enabled", False):
        enable_treeview_editing(
            tree,
            readonly_cols=(),
            on_commit=lambda *_: update_quick_subtotal()
        )
        report_tree._editing_enabled = True
# === OEE-матрица: расчёт и отрисовка ========================================
def _color_for_oee(pct: float) -> str:
    """Подбор цвета по порогам."""
    if pct is None:
        return "#f0f0f0"  # пусто
    if pct < 70:
        return "#cc3d3d"  # красный
    if pct < 85:
        return "#e98a2b"  # оранжевый
    if pct < 100:
        return "#1e9d52"  # зелёный
    if pct < 115:
        return "#21b39c"  # бирюзовый
    return "#2d6cdf"  # синий


def _compute_oee_matrix():
    flt = FilterOpts(
        selected_lines=set(SELECTED_LINES),
        selected_days=set(SELECTED_DAYS),
        current_line="Все",
    )
    return core_compute_oee_matrix(DATA, PRODUCT_BLOCKS, DOWNTIME_BY, DEFAULT_SPEED_BY_LINE, flt)




def _cell(oee_pct):
    """Текст ячейки."""
    return "" if (oee_pct is None) else f"{oee_pct:.1f}%"


def render_oee_matrix():
    """Строим матрицу во вкладке."""
    # очистка
    for w in oee_inner.winfo_children():
        w.destroy()

    days, lines, cell, totals_shift, totals_line = _compute_oee_matrix()

    # Параметры сетки
    cw = 90  # ширина ячейки
    ch = 28  # высота
    pad = 2

    # Заголовок слева: «О, % / День»
    hdr0 = ttk.Label(oee_inner, text="O, %\nДень", anchor="center")
    hdr0.grid(row=0, column=0, sticky="nsew", padx=pad, pady=pad)
    oee_inner.grid_columnconfigure(0, minsize=60)

    # Шапка по линиям (2 подколонки: День/Ночь)
    col = 1
    for line in lines:
        span = 2
        lbl = ttk.Label(oee_inner, text=f"Линия {line}", anchor="center")
        lbl.grid(row=0, column=col, columnspan=span, sticky="nsew", padx=pad, pady=pad)
        # Подзаголовки смен
        for j, shift in enumerate(("День", "Ночь")):
            sub = ttk.Label(oee_inner, text=shift, anchor="center")
            sub.grid(row=1, column=col + j, sticky="nsew", padx=pad, pady=pad)
            oee_inner.grid_columnconfigure(col + j, minsize=cw)
        col += span

    # Строки по дням
    row = 2
    for d in days:
        # метка дня слева
        day_lbl = ttk.Label(oee_inner, text=str(d), anchor="center")
        day_lbl.grid(row=row, column=0, sticky="nsew", padx=pad, pady=pad)
        # клетки по линиям/сменам
        col = 1
        for line in lines:
            for shift in ("День", "Ночь"):
                val = cell.get((d, line, shift))
                txt = _cell(val)
                bg = _color_for_oee(val)
                lab = tk.Label(
                    oee_inner, text=txt, bg=bg, fg="white", justify="center", width=10
                )
                lab.grid(
                    row=row, column=col, sticky="nsew", padx=pad, pady=pad, ipady=4
                )
                col += 1
        row += 1

    # Итоговая строка по сменам
    if days and lines:
        # подпись
        ttk.Label(oee_inner, text="Смена", anchor="center").grid(
            row=row, column=0, sticky="nsew", padx=pad, pady=pad
        )
        col = 1
        for line in lines:
            for shift in ("День", "Ночь"):
                val = totals_shift.get((line, shift))
                bg = _color_for_oee(val)
                tk.Label(oee_inner, text=_cell(val), bg=bg, fg="white").grid(
                    row=row, column=col, sticky="nsew", padx=pad, pady=pad, ipady=4
                )
                col += 1
        row += 1
        # Итоговая строка по линии (обе смены)
        ttk.Label(oee_inner, text="Линия", anchor="center").grid(
            row=row, column=0, sticky="nsew", padx=pad, pady=pad
        )
        col = 1
        for line in lines:
            # объединяем 2 колонки под одну «линию»
            val = totals_line.get(line)
            bg = _color_for_oee(val)
            tk.Label(oee_inner, text=_cell(val), bg=bg, fg="white").grid(
                row=row,
                column=col,
                columnspan=2,
                sticky="nsew",
                padx=pad,
                pady=pad,
                ipady=4,
            )
            col += 2
def _oee_color_fill(pct: float | None) -> PatternFill | None:
    if pct is None:
        return None
    if pct < 70:
        return PatternFill("solid", fgColor="CC3D3D")
    if pct < 85:
        return PatternFill("solid", fgColor="E98A2B")
    if pct < 100:
        return PatternFill("solid", fgColor="1E9D52")
    if pct < 115:
        return PatternFill("solid", fgColor="21B39C")
    return PatternFill("solid", fgColor="2D6CDF")

# === Treeview: редактирование ячеек + коп/вставка ============================
def enable_treeview_editing(tv, readonly_cols=(), on_commit=None):
    """
    Делает ячейки Treeview редактируемыми (даблклик/Enter),
    добавляет копирование/вставку (Ctrl+C / Ctrl+V) табличными блоками.
    - readonly_cols: имена столбцов (ID из tv["columns"]) ИЛИ заголовки (heading["text"]),
      которые нельзя редактировать.
    - on_commit: необязательный колбэк (item_id, col_id, old_value, new_value).
    """
    tv._tv_last_clicked_col = None  # куда вставлять при Ctrl+V
    tv._tv_editor = None
    tv._tv_readonly = set()
    col_ids = list(tv["columns"])
    # Нормализуем список readonly: принимаем и id, и заголовок
    heading_by_id = {cid: tv.heading(cid).get("text", cid) for cid in col_ids}
    id_by_heading = {v: k for k, v in heading_by_id.items()}
    for rc in readonly_cols:
        tv._tv_readonly.add(rc if rc in col_ids else id_by_heading.get(rc, rc))

    def _cell_under_cursor(event):
        col_id = tv.identify_column(event.x)  # '#1'..'#N'
        row_id = tv.identify_row(event.y)
        if not col_id or not row_id:
            return None, None
        idx = int(col_id[1:]) - 1
        if idx < 0 or idx >= len(col_ids):
            return None, None
        return row_id, col_ids[idx]

    def _start_edit(event=None, row_id=None, col_id=None):
        # определить ячейку, если не передали
        if row_id is None or col_id is None:
            row_id, col_id = _cell_under_cursor(event)
        if not row_id or not col_id:
            return
        if col_id in tv._tv_readonly:
            return
        # координаты ячейки
        bbox = tv.bbox(row_id, col_id)
        if not bbox:
            return
        x, y, w, h = bbox
        value = tv.set(row_id, col_id)

        # один редактор за раз
        if tv._tv_editor is not None:
            try:
                tv._tv_editor.destroy()
            except Exception:
                pass
            tv._tv_editor = None

        import tkinter as _tk
        entry = _tk.Entry(tv)
        entry.insert(0, value if value is not None else "")
        entry.select_range(0, 'end')
        entry.focus_set()
        entry.place(x=x, y=y, width=w, height=h)

        def _commit(e=None):
            new_val = entry.get()
            old_val = tv.set(row_id, col_id)
            entry.destroy()
            tv._tv_editor = None
            if new_val != old_val:
                tv.set(row_id, col_id, new_val)
                if callable(on_commit):
                    on_commit(row_id, col_id, old_val, new_val)

        def _cancel(e=None):
            entry.destroy()
            tv._tv_editor = None

        entry.bind("<Return>", _commit)
        entry.bind("<KP_Enter>", _commit)
        entry.bind("<Escape>", _cancel)
        entry.bind("<FocusOut>", _commit)

        tv._tv_editor = entry

    def _remember_col(event):
        # чтобы Ctrl+V знал, куда вставлять (стартовая колонка)
        col_id = tv.identify_column(event.x)
        if col_id:
            idx = int(col_id[1:]) - 1
            if 0 <= idx < len(col_ids):
                tv._tv_last_clicked_col = col_ids[idx]

    def _copy(event=None):
        # копируем выбранные строки (или одну фокусную) как TSV
        rows = tv.selection()
        if not rows:
            f = tv.focus()
            rows = (f,) if f else ()
        if not rows:
            return "break"
        data = []
        for iid in rows:
            vals = [str(tv.set(iid, c) or "") for c in col_ids]
            data.append("\t".join(vals))
        tsv = "\n".join(data)
        tv.clipboard_clear()
        tv.clipboard_append(tsv)
        return "break"

    def _paste(event=None):
        try:
            raw = tv.clipboard_get()
        except Exception:
            return "break"
        if not raw:
            return "break"

        # старт: фокусная строка и последняя кликнутая колонка
        start_row = tv.focus() or (tv.selection()[0] if tv.selection() else None)
        if not start_row:
            return "break"
        start_col = tv._tv_last_clicked_col or (col_ids[0] if col_ids else None)
        if start_col is None:
            return "break"
        start_col_idx = col_ids.index(start_col)

        # разложить TSV
        lines = raw.splitlines()
        grid = [ln.split("\t") for ln in lines]

        # список всех строк в порядке отображения
        all_items = list(tv.get_children(""))
        start_idx = all_items.index(start_row)

        # вставляем блоком
        r_i = 0
        for ridx in range(start_idx, min(start_idx + len(grid), len(all_items))):
            iid = all_items[ridx]
            row_vals = list(tv.item(iid, "values"))
            row_vals = row_vals + [""] * max(0, len(col_ids) - len(row_vals))
            cells = grid[r_i]
            for c_rel, val in enumerate(cells):
                c_abs = start_col_idx + c_rel
                if c_abs >= len(col_ids):
                    break
                col_id = col_ids[c_abs]
                if col_id in tv._tv_readonly:
                    continue
                old = row_vals[c_abs]
                if old != val:
                    row_vals[c_abs] = val
                    if callable(on_commit):
                        on_commit(iid, col_id, old, val)
            tv.item(iid, values=row_vals)
            r_i += 1
            if r_i >= len(grid):
                break
        return "break"

    def _select_all(event=None):
        tv.selection_set(tv.get_children(""))
        return "break"

    # биндим
    tv.bind("<Double-1>", _start_edit)        # редактирование
    tv.bind("<Button-1>", _remember_col, add="+")  # помнить колонку
    tv.bind("<Control-c>", _copy)
    tv.bind("<Control-C>", _copy)
    tv.bind("<Control-Insert>", _copy)
    tv.bind("<Control-v>", _paste)
    tv.bind("<Control-V>", _paste)
    tv.bind("<Shift-Insert>", _paste)
    tv.bind("<Control-a>", _select_all)
    tv.bind("<Control-A>", _select_all)

    # Возвращаем внутренние хендлеры на всякий
    return {"start_edit": _start_edit, "copy": _copy, "paste": _paste}
# ============================================================================#

def export_report_to_excel(tv_source=None):
    log("[Отчёт] Запуск формирования…")

    # Если передали Treeview — читаем из него, иначе собираем сводку
    if tv_source is not None:
        headers_sum, rows_sum = _read_from_treeview(tv_source)
    else:
        headers_sum, rows_sum = build_summary_rows()

    if not rows_sum:
        messagebox.showinfo("Пусто", "Нет данных для отчёта.")
        return

    # Формируем те же структуры, что пойдут в Excel/предпросмотр
    rep_headers_lines, rep_rows_lines, rep_headers_products, rep_rows_products = \
        _compute_export_tables_from_summary(headers_sum, rows_sum)

    # Диалог сохранения
    import datetime as _dt, os as _os
    ts = _dt.datetime.now().strftime("%Y%m%d_%H%M")
    default_name = f"Отчёт_OEE_{ts}.xlsx"
    path = filedialog.asksaveasfilename(
        title="Сохранить отчёт",
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Excel", "*.xlsx"), ("Все файлы", "*.*")]
    )
    if not path:
        log("[Отчёт] Отменено пользователем")
        return

    # Оформление Excel
    wb = Workbook()
    filters_text = _sel_to_human()

    def write_sheet(ws, title: str, headers: list[str], rows: list[list],
                widths: dict[int, int], oee_col_idx: int,
                merge_dup_col_idx: int | None = None,
                # ↓ НОВОЕ:
                group_col_idx: int | None = None,          # по этой колонке группируем (1-based)
                merge_cols_within_group: list[int] | None = None,  # какие колонки мерджить внутри группы (например [1,2])
                add_group_separators: bool = False):        # рисовать жирный разделитель между группами
        """Оформление таблицы + проценты + (опц.) группировка и разделители."""
        from datetime import datetime as _dt

        # Заголовок и фильтры
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=16, color="004578")
        ws["A2"] = filters_text
        ws["A2"].font = Font(italic=True, color="666666")
        ws["A3"] = ""
        start_row = 4

        # Шапка
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=start_row, column=j, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.fill = PatternFill("solid", fgColor="EAEAEA")
        ws.freeze_panes = f"A{start_row+1}"
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row}"

        thin = Side(style="thin", color="DADADA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        zebra2 = PatternFill("solid", fgColor="F4F6F8")
        sep_bottom = Side(style="medium", color="C0C0C0")  # << добавь это

        # Данные
        data_first = start_row + 1
        data_row = data_first
        for r_idx, row in enumerate(rows, start=0):
            fill = zebra2 if (r_idx % 2) else None
            for j, v in enumerate(row, 1):
                c = ws.cell(row=data_row, column=j, value=v)
                c.border = border
                c.alignment = Alignment(vertical="center", wrap_text=True)
                if fill:
                    c.fill = fill

                # ← центрируем номер линии (1-й столбец)
                if j == 1:
                    c.alignment = Alignment(horizontal="center", vertical="center")

                # План / Факт
                if j in (2, 3) and len(headers) in (5, 7):
                    c.number_format = "#,##0"
                    c.alignment = Alignment(horizontal="right", vertical="center")

                # Процент
                if j == oee_col_idx:
                    try:
                        num = float(str(v).replace(" ", "").replace("\u00A0", "").replace(",", "."))
                    except Exception:
                        num = 0.0
                    c.value = num / 100.0
                    c.number_format = "0.0%"
                    c.alignment = Alignment(horizontal="right", vertical="center")

            

            data_row += 1


        data_last = data_row - 1

        # Ширины
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Цветовая шкала для OEE
        if data_last >= data_first:
            col = get_column_letter(oee_col_idx)
            ws.conditional_formatting.add(
                f"{col}{data_first}:{col}{data_last}",
                ColorScaleRule(
                    start_type="num", start_value=0.70, start_color="F8696B",
                    mid_type="num",   mid_value=1.00, mid_color="FFEB84",
                    end_type="num",   end_value=1.15, end_color="63BE7B",
                )
            )

        # ===== ГРУППИРОВАНИЕ И РАЗДЕЛИТЕЛИ ===================================
        if group_col_idx is not None and rows:
            # 1) мерджим указанные колонки внутри каждой группы
            merge_cols = set(merge_cols_within_group or [])
            # прогоним по диапазону и найдём участки одинакового значения group_col
            gcol = get_column_letter(group_col_idx)
            run_start = data_first
            prev = ws[f"{gcol}{data_first}"].value

            thick_top = Side(style="medium", color="B0B0B0")  # жирный разделитель
            for rr in range(data_first + 1, data_last + 1 + 1):
                cur = ws[f"{gcol}{rr}"].value if rr <= data_last else None
                if cur != prev:
                    # мерджим внутри диапазона [run_start, rr-1] для всех колонок из merge_cols
                    for col_idx in sorted(merge_cols):
                        col_letter = get_column_letter(col_idx)
                        # делаем подгрупповые мерджи для подряд одинаковых значений
                        sub_start = run_start
                        sub_prev = ws[f"{col_letter}{run_start}"].value
                        for r2 in range(run_start + 1, rr):
                            v2 = ws[f"{col_letter}{r2}"].value
                            if v2 != sub_prev:
                                if sub_prev not in (None, "", " ") and r2 - 1 > sub_start:
                                    ws.merge_cells(f"{col_letter}{sub_start}:{col_letter}{r2-1}")
                                    ws[f"{col_letter}{sub_start}"].alignment = Alignment(horizontal="center", vertical="center")
                                sub_start = r2
                                sub_prev = v2
                        if sub_prev not in (None, "", " ") and rr - 1 > sub_start:
                            ws.merge_cells(f"{col_letter}{sub_start}:{col_letter}{rr-1}")
                            ws[f"{col_letter}{sub_start}"].alignment = Alignment(horizontal="center", vertical="center")

                    # жирный верхний бордер у первой строки НОВОЙ группы
                    if add_group_separators and rr <= data_last:
                        for j in range(1, len(headers) + 1):
                            cell = ws.cell(row=rr, column=j)
                            cell.border = Border(
                                left=cell.border.left, right=cell.border.right,
                                top=thick_top, bottom=cell.border.bottom
                            )
                    run_start = rr
                    prev = cur

        # Подпись
        stamp_row = (data_last if data_last >= data_first else start_row) + 2
        ws.cell(row=stamp_row, column=len(headers),
                value=_dt.now().strftime("Сформировано: %Y-%m-%d %H:%M")
            ).font = Font(italic=True, color="777777")


    # Лист 1 — По линиям
    ws_lines = wb.active
    ws_lines.title = "По линиям"
    write_sheet(
        ws_lines,
        "Отчёт по выпуску и OEE (по линиям)",
        rep_headers_lines,
        rep_rows_lines,
        widths={1: 10, 2: 14, 3: 14, 4: 8, 5: 70},
        oee_col_idx=4,
    )


    # Лист 2 — По продуктам
    ws_prod = wb.create_sheet("По продуктам")
    write_sheet(
    ws_prod,
    "Отчёт по выпуску и OEE (по продуктам)",
    rep_headers_products,
    rep_rows_products,
    widths={1: 8, 2: 8, 3: 34, 4: 14, 5: 14, 6: 8, 7: 70},
    oee_col_idx=6,
    # то, что ты добавлял
    group_col_idx=1,
    merge_cols_within_group=[1, 2],
    add_group_separators=True
    )

    # Сохранение и пост-акции
    wb.save(path)
    try:
        lbl_report_info.config(text=f"Сохранено: {path}")
    except Exception:
        pass
    log(f"[Отчёт] Файл сохранён: {path}")

    try:
        _os.startfile(path)
    except Exception:
        pass







# --- кнопки выбора источника -------------------------------------------------
state = {"path": None, "paths": [], "xls": None}

# фильтры файлов папки
ACCEPT_EXT = (".xlsx", ".xlsm")
ACCEPT_JSON = (".json",)          # ← ДОБАВЛЕНО: поддержка JSON
SKIP_PREFIXES = ("~$",)
# === КОМПАКТНЫЙ КЛАСС ДЛЯ ЧТЕНИЯ EXCEL ОТЧЕТОВ ============================
class ExcelReportReader:
    """Компактный класс для чтения отчетов из Excel файлов"""
    
    def __init__(self, ranges_config: dict = None):
        """Инициализация с конфигурацией диапазонов"""
        self.ranges = ranges_config or RANGES
        self.accept_ext = (".xlsx", ".xlsm")
        self.skip_prefixes = ("~$",)
    
    def read_range_fast(self, xls: pd.ExcelFile, sheet: str, a1_range: str, 
                       headers: bool = False, round_int: bool = False,
                       drop_letters: Optional[List[str]] = None, 
                       add_desc: bool = False) -> Dict[str, List[List]]:
        """Быстрое чтение диапазона Excel"""
        r1, c1, r2, c2 = parse_a1_range(a1_range)
        usecols = f"{index_to_col_letters(c1)}:{index_to_col_letters(c2)}"
        skip = r1 - 1
        nrows = r2 - r1 + 1

        with pd.option_context("mode.chained_assignment", None):
            df = xls.parse(
                sheet_name=sheet, header=None, engine="openpyxl",
                usecols=usecols, skiprows=skip, nrows=nrows,
                dtype=str, keep_default_na=False,
            )
            
            # Обработка описания для простоев
            desc_series = None
            if add_desc:
                try:
                    desc_rel = col_letters_to_index("G") - c1
                    if 0 <= desc_rel < df.shape[1]:
                        desc_series = df.iloc[:, desc_rel].astype(str)
                except Exception:
                    desc_series = None

            # Удаление ненужных колонок
            if drop_letters:
                start_abs = c1
                width = df.shape[1]
                to_drop_idx = []
                for L in drop_letters:
                    j_abs = col_letters_to_index(str(L))
                    rel_idx = j_abs - start_abs
                    if 0 <= rel_idx < width:
                        to_drop_idx.append(rel_idx)
                if to_drop_idx:
                    df.drop(df.columns[to_drop_idx], axis=1, inplace=True)

            # Добавление описания
            if add_desc and desc_series is not None:
                df = pd.concat([df, desc_series.rename("__DESC__")], axis=1)

        arr = df.values.tolist()
        if not arr:
            return {"array": [], "headers": []}

        # Формирование заголовков и данных
        if headers:
            hdrs = [str(x) if x != "" else f"col{j+1}" for j, x in enumerate(arr[0])]
            rows = arr[1:]
        else:
            width = len(arr[0])
            hdrs = [index_to_col_letters(j + 1) for j in range(width)]
            rows = arr

        # Округление чисел
        if round_int:
            def to_num(v):
                try:
                    fv = float(str(v).replace(",", "."))
                    if math.isfinite(fv) and abs(fv - round(fv)) < 1e-9:
                        return int(round(fv))
                    return fv
                except Exception:
                    return v
            rows = [[to_num(v) for v in r] for r in rows]

        return {"array": rows, "headers": hdrs}
    
    def validate_file(self, file_path: str) -> bool:
        """Проверка валидности файла"""
        base = os.path.basename(file_path)
        
        # Проверка расширения
        if not base.lower().endswith(self.accept_ext):
            return False
        
        # Проверка временных файлов
        if base.startswith(self.skip_prefixes):
            return False
        
        # Проверка OOXML zip
        try:
            if not zipfile.is_zipfile(file_path):
                return False
        except Exception:
            return False
        
        return True
    
    def get_day_sheets(self, xls: pd.ExcelFile) -> List[str]:
        """Получение листов с числовыми названиями (дни)"""
        return [s for s in xls.sheet_names if str(s).strip().isdigit()]
    
    def extract_line_number(self, filename: str) -> str:
        """Извлечение номера линии из имени файла"""
        m_line = re.search(r"линия\s*№?\s*(\d+)", filename, re.IGNORECASE)
        return m_line.group(1) if m_line else ""
    
    def read_products(self, xls: pd.ExcelFile, sheet_name: str, 
                     base_name: str, line_num: str) -> List[Dict]:
        """Чтение данных продуктов"""
        results = []
        
        for idx, pr in enumerate(self.ranges["products"], start=1):
            try:
                res = self.read_range_fast(
                    xls=xls, sheet=sheet_name, a1_range=pr["range"],
                    headers=pr.get("headers", False),
                    round_int=pr.get("round_int", True),
                    drop_letters=None, add_desc=False
                )
                
                name = f"{base_name} | {sheet_name} / Продукты #{idx}"
                res["meta"] = {
                    "file": base_name, "sheet": sheet_name, "kind": "products",
                    "line": str(line_num),
                    "shift": "День" if idx == 1 else ("Ночь" if idx == 2 else f"#{idx}")
                }
                
                has_data = bool(res["array"]) and any(
                    str(r[0]).strip() not in ("", "0") for r in res["array"]
                )
                
                results.append({
                    "name": name,
                    "data": res,
                    "has_data": has_data
                })
                
            except Exception as e:
                results.append({
                    "name": f"{base_name} | {sheet_name} / Продукты #{idx}",
                    "data": {"array": [], "headers": []},
                    "has_data": False,
                    "error": str(e)
                })
        
        return results
    
    def read_downtimes(self, xls: pd.ExcelFile, sheet_name: str,
                      base_name: str, line_num: str) -> List[Dict]:
        """Чтение данных простоев"""
        results = []
        
        for idx, dt in enumerate(self.ranges["downtimes"], start=1):
            try:
                res = self.read_range_fast(
                    xls=xls, sheet=sheet_name, a1_range=dt["range"],
                    headers=dt.get("headers", False),
                    round_int=dt.get("round_int", False),
                    drop_letters=dt.get("drop_cols", None),
                    add_desc=True
                )
                
                name = f"{base_name} | {sheet_name} / Простои #{idx}"
                res["meta"] = {
                    "file": base_name, "sheet": sheet_name, "kind": "downtimes",
                    "shift": "День" if idx == 1 else ("Ночь" if idx == 2 else f"#{idx}")
                }
                
                has_data = bool(res["array"]) and any(
                    str(r[0]).strip() not in ("", "0") for r in res["array"]
                )
                
                results.append({
                    "name": name,
                    "data": res,
                    "has_data": has_data
                })
                
            except Exception as e:
                results.append({
                    "name": f"{base_name} | {sheet_name} / Простои #{idx}",
                    "data": {"array": [], "headers": []},
                    "has_data": False,
                    "error": str(e)
                })
        
        return results
    
    def read_file(self, file_path: str) -> Dict:
        """Чтение одного Excel файла"""
        if not self.validate_file(file_path):
            return {"error": "Невалидный файл", "data": {}}
        
        try:
            xls = pd.ExcelFile(file_path, engine="openpyxl")
            base_name = os.path.basename(file_path)
            line_num = self.extract_line_number(base_name)
            
            day_sheets = self.get_day_sheets(xls)
            if not day_sheets:
                return {"error": "Нет листов с числовыми днями", "data": {}}
            
            data = {}
            product_blocks = []
            downtime_blocks = []
            all_blocks = []
            
            for sheet_name in day_sheets:
                # Чтение продуктов
                products = self.read_products(xls, sheet_name, base_name, line_num)
                for product in products:
                    data[product["name"]] = product["data"]
                    all_blocks.append(product["name"])
                    if product["has_data"]:
                        product_blocks.append(product["name"])
                
                # Чтение простоев
                downtimes = self.read_downtimes(xls, sheet_name, base_name, line_num)
                for downtime in downtimes:
                    data[downtime["name"]] = downtime["data"]
                    all_blocks.append(downtime["name"])
                    if downtime["has_data"]:
                        downtime_blocks.append(downtime["name"])
            
            return {
                "data": data,
                "product_blocks": product_blocks,
                "downtime_blocks": downtime_blocks,
                "all_blocks": all_blocks
            }
            
        except Exception as e:
            return {"error": str(e), "data": {}}

# Создаем глобальный экземпляр
excel_reader = ExcelReportReader(RANGES)


# === ФОНОВАЯ ЗАГРУЗКА ========================================================
_q = queue.Queue()
_cancel_event = threading.Event()
_worker_thread: Optional[threading.Thread] = None


def poll_queue():
    """Пуллим события из рабочего потока (каждые 50 мс)."""
    try:
        while True:
            kind, payload = _q.get_nowait()
            if kind == "progress_init":
                prg["maximum"] = max(1, int(payload))
                prg["value"] = 0
                lbl_prog.config(text="Загрузка…")
                btn_cancel.config(state="normal")
            elif kind == "progress_step":
                prg.step(1)
            elif kind == "log":
                log(str(payload))
            elif kind == "result":
                apply_loaded_result(payload)
                lbl_prog.config(text="Готово")
                btn_cancel.config(state="disabled")
            elif kind == "error":
                messagebox.showerror("Ошибка чтения", str(payload))
                lbl_prog.config(text="Ошибка")
                btn_cancel.config(state="disabled")
            elif kind == "canceled":
                lbl_prog.config(text="Отменено")
                btn_cancel.config(state="disabled")
            _q.task_done()
    except queue.Empty:
        pass
    root.after(50, poll_queue)


def start_load(paths: List[str]):
    """Запуск рабочего потока (Excel или JSON)."""
    global _worker_thread
    if _worker_thread and _worker_thread.is_alive():
        return
    _cancel_event.clear()
    btn_file.config(state="disabled")
    btn_folder.config(state="disabled")

    # определяем режим по первому пути
    mode = "excel"
    if paths and paths[0].lower().endswith(ACCEPT_JSON):
        mode = "json"

    target = worker_read_json if mode == "json" else worker_read
    _worker_thread = threading.Thread(
        target=target, args=(paths, _q, _cancel_event), daemon=True
    )
    _worker_thread.start()



def cancel_load():
    _cancel_event.set()


btn_cancel.config(command=cancel_load)

# ===== ЧТЕНИЕ JSON-ОТЧЁТА МАСТЕРА ===========================================
def _is_master_json(path: str) -> bool:
    if not str(path).lower().endswith(".json"):
        return False
    try:
        with open(path, "r", encoding="utf-8") as f:
            head = f.read(4096)
        import json as _json
        obj = _json.loads(head) if head.strip().startswith("{") else None
        return bool(isinstance(obj, dict) and str(obj.get("$schema","")).lower().startswith("master_report"))
    except Exception:
        return False

def _safe_int(v, default=0):
    try:
        if v in (None, ""):
            return default
        return int(round(float(str(v).replace(",", "."))))
    except Exception:
        return default

def _coalesce(*vals, default=""):
    for v in vals:
        if v not in (None, ""):
            return v
    return default

def _load_master_json_to_blocks(path: str):
    """Читает JSON по схеме master_report.v1 и возвращает
    (DATA_local, PRODUCT_BLOCKS_local, DOWNTIME_BLOCKS_local, ALL_BLOCKS_local)
    в том же формате, что и Excel-ветка."""
    import json as _json
    with open(path, "r", encoding="utf-8") as f:
        obj = _json.load(f)
    if not isinstance(obj, dict) or not str(obj.get("$schema","")).lower().startswith("master_report"):
        raise ValueError("Файл не выглядит как отчёт мастера (нет $schema: master_report.*).")

    DATA_local: Dict[str, Dict] = {}
    PRODUCT_BLOCKS_local: List[str] = []
    DOWNTIME_BLOCKS_local: List[str] = []
    ALL_BLOCKS_local: List[str] = []

    # Общие мета (если заданы на верхнем уровне)
    meta_day  = str(obj.get("day","")).strip()
    meta_line = str(obj.get("line","")).strip()

    # ---- PRODUCTS ------------------------------------------------------
    by_key_prod: Dict[tuple, list] = {}  # (day, shift, line) -> rows
    for p in (obj.get("products") or []):
        if not isinstance(p, dict): 
            continue
        name  = str(p.get("name","")).strip()
        if not name or name == "0":
            continue
        day   = str(_coalesce(p.get("day"), meta_day)).strip()
        shift = str(_coalesce(p.get("shift"), "День")).strip().capitalize()
        line  = str(_coalesce(p.get("line"), meta_line)).strip()

        beg   = str(p.get("beg",""))
        end   = str(p.get("end",""))
        dur_m = _safe_int(p.get("run_min"), 0)      # длительность запуска, мин
        speed = _coalesce(p.get("speed"), "")       # шт/ч
        fact  = _safe_int(p.get("fact_qty"), 0)     # шт

        # Сформируем строку под текущие ожидания индексов:
        # B1_COL_NAME=0, B1_COL_BEG=1, B1_COL_END=2, B1_COL_DUR=3, скорость — в колонке 4, факт — в «последних» числах
        row = [""] * 12
        row[B1_COL_NAME] = name
        row[B1_COL_BEG]  = beg
        row[B1_COL_END]  = end
        row[B1_COL_DUR]  = dur_m
        # в «продуктах» твой код читает скорость как row[4]
        if len(row) < 5:
            row += [""] * (5 - len(row))
        row[4] = speed if speed != "" else ""
        # а факт вытягивается _extract_fact_qty как «последнее число в строке» — положим ближе к концу
        row[-1] = fact

        by_key_prod.setdefault((day, shift, line), []).append(row)

    for (day, shift, line), arr in by_key_prod.items():
        base = os.path.basename(path)
        block_name = f"{base} | {day or '—'} / Продукты ({shift})"
        DATA_local[block_name] = {
            "array": arr,
            "headers": [index_to_col_letters(j+1) for j in range(max(len(r) for r in arr) if arr else 1)],
            "meta": {
                "file": base, "sheet": day or "", "kind": "products",
                "line": line or "", "shift": shift or "День",
            },
        }
        ALL_BLOCKS_local.append(block_name)
        if arr and any(str(r[B1_COL_NAME]).strip() not in ("","0") for r in arr):
            PRODUCT_BLOCKS_local.append(block_name)

    # ---- DOWNTIMES -----------------------------------------------------
    by_key_dt: Dict[tuple, list] = {}  # (day, shift, line) -> rows
    for d in (obj.get("downtimes") or []):
        if not isinstance(d, dict):
            continue
        name  = str(d.get("product","")).strip()
        if not name:
            continue
        day   = str(_coalesce(d.get("day"), meta_day)).strip()
        shift = str(_coalesce(d.get("shift"), "День")).strip().capitalize()
        line  = str(_coalesce(d.get("line"), meta_line)).strip()
        kind  = str(d.get("kind",""))
        reason= str(d.get("reason",""))
        beg   = str(d.get("beg",""))
        end   = str(d.get("end",""))
        mins  = _safe_int(d.get("minutes"), 0)
        desc  = str(d.get("desc",""))

        # Порядок колонок в коде: NAME(0), REASON(1), KIND(2), BEG(3), END(4), MIN(5), DESC(6)
        row = [""] * 7
        row[D2_COL_NAME]   = name
        row[D2_COL_REASON] = reason
        row[D2_COL_KIND]   = kind
        row[D2_COL_BEG]    = beg
        row[D2_COL_END]    = end
        row[D2_COL_MIN]    = mins
        row[D2_COL_DESC]   = desc

        by_key_dt.setdefault((day, shift, line), []).append(row)

    for (day, shift, line), arr in by_key_dt.items():
        base = os.path.basename(path)
        block_name = f"{base} | {day or '—'} / Простои ({shift})"
        # headers тут не критичны (у тебя они не обязательны), дадим по ширине
        DATA_local[block_name] = {
            "array": arr,
            "headers": [index_to_col_letters(j+1) for j in range(max(len(r) for r in arr) if arr else 1)],
            "meta": {
                "file": base, "sheet": day or "", "kind": "downtimes",
                "line": line or "", "shift": shift or "День",
            },
        }
        ALL_BLOCKS_local.append(block_name)
        if arr and any(str(r[D2_COL_NAME]).strip() not in ("","0") for r in arr):
            DOWNTIME_BLOCKS_local.append(block_name)

    if not PRODUCT_BLOCKS_local and not DOWNTIME_BLOCKS_local:
        raise ValueError("В JSON не найдены секции products[] или downtimes[].")

    return DATA_local, PRODUCT_BLOCKS_local, DOWNTIME_BLOCKS_local, ALL_BLOCKS_local

def worker_read(paths: List[str], q: queue.Queue, cancel_evt: threading.Event):
    """
    Фоновая загрузка: поддерживает Excel (*.xlsx/*.xlsm) и JSON отчёты мастера (*.json).
    Для Excel:
      - читаем только листы-дни (имя — число),
      - продукты и простои по прямоугольникам RANGES (ускоренное чтение),
      - безопасно пропускаем временные файлы/битые книги.
    Для JSON:
      - читаем по схеме master_report.* через _load_master_json_to_blocks.
    """
    try:
        # 0) Соберём валидные пути и заранее посчитаем объём работ для прогресса
        valid_paths: List[str] = []
        total_tasks = 0

        for p in paths:
            if cancel_evt.is_set():
                q.put(("canceled", None))
                return

            base = os.path.basename(p)
            low = base.lower()

            # неподдерживаемые расширения
            if not low.endswith(ACCEPT_EXT):
                q.put(("log", f"[skip] {base} — неподдерживаемое расширение"))
                continue

            # --- JSON ветка: считаем как одну задачу ---
            if low.endswith(".json"):
                if not _is_master_json(p):
                    q.put(("log", f"[skip] {base} — JSON не по схеме master_report"))
                    continue
                total_tasks += 1
                valid_paths.append(p)
                continue

            # --- Excel ветка: фильтры и проверка OOXML zip ---
            if base.startswith(SKIP_PREFIXES):
                q.put(("log", f"[skip] {base} — временный файл"))
                continue
            try:
                if not zipfile.is_zipfile(p):
                    q.put(("log", f"[skip] {base} — не OOXML zip (битый/не тот формат)"))
                    continue
            except Exception as e:
                q.put(("log", f"[skip] {base} — ошибка проверки zip: {e}"))
                continue

            # пробуем открыть книгу и посчитать, сколько листов-дней
            try:
                xls = pd.ExcelFile(p, engine="openpyxl")
            except Exception as e:
                q.put(("log", f"[skip] {base} — не открылся: {e}"))
                continue

            day_sheets = [s for s in xls.sheet_names if str(s).strip().isdigit()]
            if not day_sheets:
                q.put(("log", f"[skip] {base} — нет листов с числовыми днями"))
                continue

            # на каждый день — продукты + простои, как и раньше
            total_tasks += len(day_sheets) * (len(RANGES["products"]) + len(RANGES["downtimes"]))
            valid_paths.append(p)

        q.put(("progress_init", total_tasks if total_tasks > 0 else 1))

        # 1) Локальные контейнеры результата
        DATA_local: Dict[str, Dict[str, List[List]]] = {}
        PRODUCT_BLOCKS_local: List[str] = []
        DOWNTIME_BLOCKS_local: List[str] = []
        ALL_BLOCKS_local: List[str] = []

        # 2) Основной цикл чтения
        for p in valid_paths:
            if cancel_evt.is_set():
                q.put(("canceled", None))
                return

            base = os.path.basename(p)
            low  = base.lower()

            # ---- JSON ветка ---------------------------------------------------
            if low.endswith(".json"):
                q.put(("log", f"Читаем JSON: {base}"))
                try:
                    DATA_j, PROD_j, DOWN_j, ALL_j = _load_master_json_to_blocks(p)
                    # вносим в локальные контейнеры
                    for k, v in DATA_j.items():
                        DATA_local[k] = v
                        ALL_BLOCKS_local.append(k)
                    PRODUCT_BLOCKS_local.extend(PROD_j)
                    DOWNTIME_BLOCKS_local.extend(DOWN_j)
                except Exception as e:
                    q.put(("log", f"[skip] {base} — ошибка JSON: {e}"))
                finally:
                    q.put(("progress_step", 1))
                continue  # к следующему файлу

            # ---- Excel ветка (упрощенная с использованием класса) -----------
            q.put(("log", f"Читаем: {base}"))
            
            try:
                result = excel_reader.read_file(p)
                if "error" in result:
                    q.put(("log", f"[skip] {base} — {result['error']}"))
                    continue
                
                # Добавляем данные в локальные контейнеры
                for name, data in result["data"].items():
                    DATA_local[name] = data
                    ALL_BLOCKS_local.append(name)
                
                PRODUCT_BLOCKS_local.extend(result["product_blocks"])
                DOWNTIME_BLOCKS_local.extend(result["downtime_blocks"])
                
                # Обновляем прогресс
                total_sheets = len(result["all_blocks"])
                for _ in range(total_sheets):
                    q.put(("progress_step", 1))
                    
            except Exception as e:
                q.put(("log", f"[skip] {base} — ошибка чтения: {e}"))
                continue

        # 3) Готово — отдаём всё наверх
        q.put(("result", (DATA_local, PRODUCT_BLOCKS_local, DOWNTIME_BLOCKS_local, ALL_BLOCKS_local)))

    except Exception as e:
        q.put(("error", e))

# ======================================================================
#                   ЧТЕНИЕ МАСТЕРСКОГО JSON (ВТОРОЙ ВАРИАНТ)
# ======================================================================

def _as_int(x, default=0):
    try:
        if x is None or str(x).strip() == "":
            return default
        return int(round(float(str(x).replace(",", "."))))
    except Exception:
        return default

def _as_str(x, default=""):
    s = "" if x is None else str(x)
    return s.strip() if s else default

def _guess_shift(v) -> str:
    s = _as_str(v).lower()
    if s in ("1","день","day","day1","shift1","дневная"): return "День"
    if s in ("2","ночь","night","shift2","ночная"): return "Ночь"
    # по времени
    if re.search(r"\b(07|08|09|10|11|12|13|14|15|16)\b", s): return "День"
    if re.search(r"\b(19|20|21|22|23|00|01|02|03|04)\b", s): return "Ночь"
    return "День"  # дефолт

def _guess_day(v) -> str:
    """Ищем день месяца: сначала по паттерну YYYY-MM-DD, затем любые 1–2 цифры в начале/конце."""
    s = _as_str(v)
    if not s:
        return ""
    # 1) ISO/дата внутри строки: 2025-10-20 / 2025/10/20 / 2025.10.20
    m = re.search(r"(\d{4})[-/\.](\d{1,2})[-/\.](\d{1,2})", s)
    if m:
        dd = int(m.group(3))
        return str(dd)
    # 2) чистое число
    if s.isdigit():
        return s
    # 3) иначе берём первую 1–2-значную группу как день
    m = re.search(r"\b(\d{1,2})\b", s)
    return m.group(1) if m else ""


def _guess_line(v) -> str:
    s = _as_str(v)
    m = re.search(r"(\d+)", s)
    return m.group(1) if m else s

def _hms_or_blank(s) -> str:
    s = _as_str(s)
    return s if s else ""

def _minutes_from_hhmm(beg: str, end: str) -> int:
    try:
        def _to_min(t):
            t = t.strip()
            if not t: return None
            hh, mm = re.split(r"[:.]", t)[:2]
            return int(hh) * 60 + int(mm)
        a = _to_min(beg); b = _to_min(end)
        if a is None or b is None:
            return 0
        # переход через полночь
        if b < a:
            b += 24*60
        return max(b - a, 0)
    except Exception:
        return 0

def _ingest_master_json_object(
    obj,
    DATA_local,
    PRODUCT_BLOCKS_local,
    DOWNTIME_BLOCKS_local,
    ALL_BLOCKS_local,
    base_name,
    *,
    day_hint: str = ""
):
    """
    Принимает один объект "job" и добавляет 2 блока:
      1) Продукты: [name, beg, end, run_min, speed, fact_qty]
      2) Простои:  [name, reason, kind, beg, end, minutes, desc]
    Meta: day/shift/line.
    """
    name  = _as_str(obj.get("name") or obj.get("product") or obj.get("title"))
    if not name:
        return

    line  = _guess_line(obj.get("line") or obj.get("line_no") or obj.get("line_num") or "")
    # день: сначала из поля job["day"/"date"], затем из meta.* или внешнего day_hint
    day   = _guess_day(obj.get("day") or obj.get("date") or "")
    if not day:
        meta = obj.get("meta") or {}
        day = _guess_day(meta.get("created_at") or meta.get("updated_at") or day_hint)

    shift = _guess_shift(obj.get("shift"))

    beg = _hms_or_blank(obj.get("start") or obj.get("start_time"))
    end = _hms_or_blank(obj.get("end")   or obj.get("end_time"))

    run_min = _as_int(obj.get("run_min") or obj.get("duration_min"))
    if run_min <= 0 and (beg or end):
        run_min = _minutes_from_hhmm(beg, end)

    speed = None
    try:
        v = obj.get("speed")
        speed = float(str(v).replace(",", ".")) if v not in (None, "") else None
    except Exception:
        speed = None

    fact_qty = obj.get("fact_qty") or obj.get("actual") or obj.get("produced")
    fact_qty = _as_int(fact_qty, default=None)

    # ---------- Продукты ----------
    prod_headers = ["A","B","C","D","E","F"]
    prod_row = [name, beg, end, run_min, speed if speed is not None else "", fact_qty if fact_qty is not None else ""]
    prod_block_name = f"{base_name} | {day or '—'} / Продукты (JSON)"
    if prod_block_name not in DATA_local:
        DATA_local[prod_block_name] = {"headers": prod_headers, "array": [], "meta": {
            "file": base_name, "sheet": day or "", "kind": "products", "line": str(line), "shift": shift
        }}
        ALL_BLOCKS_local.append(prod_block_name)
        PRODUCT_BLOCKS_local.append(prod_block_name)
    DATA_local[prod_block_name]["array"].append(prod_row)

    # ---------- Простои ----------
    dts = obj.get("downtimes") or []
    if isinstance(dts, dict):
        dts = dts.get("items") or []
    if dts:
        dt_headers = ["A","B","C","D","E","F","__DESC__"]
        dt_block_name = f"{base_name} | {day or '—'} / Простои (JSON)"
        if dt_block_name not in DATA_local:
            DATA_local[dt_block_name] = {"headers": dt_headers, "array": [], "meta": {
                "file": base_name, "sheet": day or "", "kind": "downtimes", "shift": shift
            }}
            ALL_BLOCKS_local.append(dt_block_name)
            DOWNTIME_BLOCKS_local.append(dt_block_name)
        for ev in dts:
            reason = _as_str(ev.get("reason") or ev.get("category"))     # reason
            kind   = _as_str(ev.get("kind") or ev.get("type") or ev.get("category"))  # ← category = kind
            dbeg   = _hms_or_blank(ev.get("beg") or ev.get("start"))
            dend   = _hms_or_blank(ev.get("end") or ev.get("stop"))
            mins   = _as_int(ev.get("minutes") or ev.get("dur") or ev.get("duration_min"))
            if mins <= 0 and (dbeg or dend):
                mins = _minutes_from_hhmm(dbeg, dend)
            desc   = _as_str(ev.get("desc") or ev.get("note") or ev.get("comment"))
            DATA_local[dt_block_name]["array"].append([name, reason, kind, dbeg, dend, mins, desc])



def worker_read_json(paths: List[str], q: queue.Queue, cancel_evt: threading.Event):
    """
    Чтение JSON отчёта мастера.
    Поддерживаем корни:
      - список записей
      - {"data":[...]}      ← ваш случай
      - {"jobs":[...]} / {"items":[...]}
      - одиночный объект (как один job)
    """
    try:
        if not paths:
            q.put(("error", "Не указан путь к JSON"))
            return

        path = paths[0]
        base = os.path.basename(path)

        if not path.lower().endswith(ACCEPT_JSON):
            q.put(("error", f"Не JSON: {base}"))
            return

        q.put(("progress_init", 10))
        q.put(("log", f"Читаем JSON: {base}"))

        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            q.put(("error", f"Не удалось прочитать JSON: {e}"))
            return

        DATA_local: Dict[str, Dict[str, List[List]]] = {}
        PRODUCT_BLOCKS_local: List[str] = []
        DOWNTIME_BLOCKS_local: List[str] = []
        ALL_BLOCKS_local: List[str] = []

        # где искать список работ
        jobs = []
        if isinstance(data, list):
            jobs = data
        elif isinstance(data, dict):
            # поддерживаем разные ключи
            if isinstance(data.get("data"), list):
                jobs = data["data"]
            elif isinstance(data.get("jobs"), list):
                jobs = data["jobs"]
            elif isinstance(data.get("items"), list):
                jobs = data["items"]
            else:
                jobs = [data]  # пробуем как одиночный job
        else:
            q.put(("error", "Неподдерживаемый формат JSON"))
            return

        # подсказка дня: saved_at (если date в job пустой)
        day_hint = _guess_day((data.get("saved_at") if isinstance(data, dict) else "") or "")

        total = max(1, len(jobs))
        q.put(("progress_init", total))

        added = 0
        for j in jobs:
            if cancel_evt.is_set():
                q.put(("canceled", None))
                return
            try:
                before_prod = len(PRODUCT_BLOCKS_local)
                _ingest_master_json_object(
                    j, DATA_local, PRODUCT_BLOCKS_local, DOWNTIME_BLOCKS_local, ALL_BLOCKS_local,
                    base_name=base, day_hint=day_hint
                )
                after_prod = len(PRODUCT_BLOCKS_local)
                if after_prod > before_prod or (DOWNTIME_BLOCKS_local and len(DOWNTIME_BLOCKS_local) > 0):
                    added += 1
            except Exception as e:
                q.put(("log", f"[warn] Пропущена запись: {e}"))
            finally:
                q.put(("progress_step", 1))

        if not PRODUCT_BLOCKS_local and not DOWNTIME_BLOCKS_local:
            q.put(("error", "В JSON не найдено ни одного задания/простоя"))
            return

        q.put(("result", (DATA_local, PRODUCT_BLOCKS_local, DOWNTIME_BLOCKS_local, ALL_BLOCKS_local)))

    except Exception as e:
        q.put(("error", e))




def apply_loaded_result(payload):
    global DATA, PRODUCT_BLOCKS, DOWNTIME_BLOCKS, ALL_BLOCKS
    DATA, PRODUCT_BLOCKS, DOWNTIME_BLOCKS, ALL_BLOCKS = payload

    # Индекс простоев
    build_downtime_index()
    inject_unaccounted_time()
    
    # Значения для фильтров — только реальные
    global ALL_LINES, ALL_DAYS
    ALL_LINES = sorted(
        {
            str(DATA[name]["meta"].get("line", "")).strip()
            for name in PRODUCT_BLOCKS
            if "meta" in DATA[name]
            and str(DATA[name]["meta"].get("line", "")).strip() != ""
        },
        key=_natural_key,
    )

    ALL_DAYS = sorted(
        {
            str(DATA[name]["meta"].get("sheet", "")).strip()
            for name in PRODUCT_BLOCKS
            if "meta" in DATA[name]
            and str(DATA[name]["meta"].get("sheet", "")).strip().isdigit()
        },
        key=lambda s: int(s),
    )

    try:
        _rebuild_filter_menus()
    except Exception:
        pass
    try:
        lbl_filters.config(text=_sel_to_human())
    except Exception:
        pass

    # Комбо: «Сводка» + сырые блоки
    try:
        combo_block["values"] = [_summary_option] + ALL_BLOCKS
        combo_block.set(_summary_option)
    except Exception:
        pass



    btn_file.config(state="normal")
    btn_folder.config(state="normal")

    # сразу показать сводку
    show_block(_summary_option)
    try:
        render_oee_matrix()
    except Exception:
        pass
    try:
        render_report_table()
    except Exception:
        pass




# --- кнопки выбора источника -------------------------------------------------
def on_pick_folder():
    folder = filedialog.askdirectory(title="Выберите папку с отчётами (Excel)")
    if not folder:
        return

    files = []
    for f in os.listdir(folder):
        p = os.path.join(folder, f)
        if not os.path.isfile(p):
            continue
        low = f.lower()
        if not low.endswith(ACCEPT_EXT):
            continue
        if f.startswith(SKIP_PREFIXES):
            continue
        files.append(p)

    files.sort()
    if not files:
        messagebox.showinfo("Пусто", "В папке нет файлов *.xlsx или *.xlsm.")
        return

    state["paths"] = files
    state["path"] = None
    lbl_file.config(text=f"{folder}  —  файлов: {len(files)}")
    log(f"Папка выбрана: {folder}")
    for i, p in enumerate(files, 1):
        log(f"  [{i}] {os.path.basename(p)}")
    start_load(files)




def on_pick_file():
    path = filedialog.askopenfilename(
        title="Выберите файл Excel/JSON",
        filetypes=[
            ("Excel", "*.xlsx *.xlsm"),
            ("JSON", "*.json"),
            ("Все файлы", "*.*")
        ],
    )
    if not path:
        return

    state["path"] = path
    state["paths"] = []
    lbl_file.config(text=path)
    log(f"Файл выбран: {path}")
    # если это JSON — помним путь
    if str(path).lower().endswith(".json"):
        _remember_last_json(path)
        try:
            btn_last_json.state(["!disabled"])
        except Exception:
            pass

    start_load([path])   # ← не меняем: start_load сам выберет нужный воркер






btn_file.configure(command=on_pick_file)
btn_folder.configure(command=on_pick_folder)
btn_make_report.configure(command=lambda: export_report_to_excel(tree))




# --- показ таблицы -----------------------------------------------------------
def show_block(name: str):
    for col in tree["columns"]:
        tree.heading(col, text="")
    tree.delete(*tree.get_children())

    if name == _summary_option:
        if not PRODUCT_BLOCKS and not DOWNTIME_BLOCKS:
            messagebox.showinfo("Нет данных", "Сначала прочитайте блоки.")
            return
        headers, rows = build_summary_rows()

        combo_events.configure(state="disabled", values=[])
        combo_events.set("")
        lbl_ev_count.config(text="")

        tree["columns"] = [f"c{i}" for i in range(len(headers))]
        col_widths = [
            220,
            60,
            70,
            70,
            90,
            90,
            90,
            110,
            90,
            70,
            120,
            120,
            110,
            120,
            110,
            80,
        ]
        for i, h in enumerate(headers):
            tree.heading(f"c{i}", text=str(h))
            w = col_widths[i] if i < len(col_widths) else 110
            tree.column(f"c{i}", width=w, anchor="w")

        for r in rows:
            tree.insert("", "end", values=[_fmt_cell(x) for x in r])
                    # сортировка по клику в сводной таблице
        enable_tree_sort(tree)
                # редактирование + коп/вставка в сводной таблице (однократно)
        if not getattr(tree, "_editing_enabled", False):
            enable_treeview_editing(
                tree,
                readonly_cols=(),   # при желании укажи неизменяемые колонки
                on_commit=None
            )
            tree._editing_enabled = True



        try:
            lbl_filters.config(text=_sel_to_human())
        except Exception:
            pass
        try:
            first = tree.get_children()
            if first:
                tree.selection_set(first[0])
                on_row_select()
        except Exception:
            pass
        return

    if name not in DATA:
        messagebox.showinfo("Нет данных", "Сначала прочитайте блоки.")
        return

    combo_events.configure(state="disabled", values=[])
    combo_events.set("")
    lbl_ev_count.config(text="")

    headers = DATA[name]["headers"]
    tree["columns"] = [f"c{i}" for i in range(len(headers))]
    col_widths = [260, 120, 90, 110, 110, 110, 120, 100, 90]
    for i, h in enumerate(headers):
        tree.heading(f"c{i}", text=str(h))
        w = col_widths[i] if i < len(col_widths) else 110
        tree.column(f"c{i}", width=w, anchor="w")

    for row in DATA[name]["array"]:
        tree.insert("", "end", values=[_fmt_cell(x) for x in row])
            # сортировка по клику в таблице блока
    enable_tree_sort(tree)
    tree.bind("<ButtonRelease-1>", lambda e: update_quick_subtotal(), add="+")




def read_range(*args, **kwargs):
    """Старое имя, чтобы ничего не ломать (вызывает быстрый вариант)."""
    return excel_reader.read_range_fast(*args, **kwargs)


def _on_close():
    try:
        save_catalog_json(silent=True)
    except Exception:
        pass
    root.destroy()

if __name__ == "__main__":
        # Загрузим настройки (последний JSON) и активируем кнопку при наличии пути
    _settings_load()
    try:
        if _get_last_json():
            btn_last_json.state(["!disabled"])
        else:
            btn_last_json.state(["disabled"])
    except Exception:
        pass

    root.protocol("WM_DELETE_WINDOW", _on_close)
    root.after(50, poll_queue)  # стартуем поллинг очереди
    root.mainloop()
