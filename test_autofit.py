# -*- coding: utf-8 -*-
"""
Тест автофита
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def test_autofit():
    """Тестируем автофит"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Тест автофита"
    
    # Добавляем тестовые данные
    test_data = [
        ["Время начала", "Время окончания", "Длительность", "Продукт", "Количество"],
        ["08:00", "11:09", "189 мин", "Сироп Миндаль 1,0 л ТМ «Баринофф»", "6000"],
        ["11:09", "14:38", "209 мин", "CIP", "-"],
        ["14:38", "20:00", "322 мин", "Сироп Манго 1,0 л ТМ «Баринофф»", "9551"],
        ["20:00", "20:30", "30 мин", "CIP", "-"],
        ["20:30", "23:30", "180 мин", "Сироп Шоколад 1,0 л ТМ «Баринофф»", "949"]
    ]
    
    # Заполняем данные
    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Выравниваем по центру
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    print("До автофита:")
    for col in range(1, 6):
        col_letter = openpyxl.utils.get_column_letter(col)
        width = ws.column_dimensions[col_letter].width
        print(f"Колонка {col_letter}: {width}")
    
    # Применяем автофит
    autofit_columns(ws)
    
    print("\nПосле автофита:")
    for col in range(1, 6):
        col_letter = openpyxl.utils.get_column_letter(col)
        width = ws.column_dimensions[col_letter].width
        print(f"Колонка {col_letter}: {width}")
    
    # Сохраняем файл
    wb.save("test_autofit.xlsx")
    print("\nФайл сохранен как test_autofit.xlsx")

def autofit_columns(ws):
    """Автофит столбцов по содержимому, как в Excel"""
    # Словарь для хранения максимальной ширины каждой колонки
    column_widths = {}
    
    for row in ws.iter_rows():
        for cell in row:
            # Пропускаем пустые ячейки и объединенные ячейки
            if not cell.value:
                continue
            
            # Проверяем, является ли ячейка частью объединенной области
            if hasattr(ws, 'merged_cells') and ws.merged_cells:
                is_merged = False
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        is_merged = True
                        break
                if is_merged:
                    continue
            
            col_letter = cell.column_letter
            if col_letter not in column_widths:
                column_widths[col_letter] = 0
            
            # Вычисляем ширину текста с учетом шрифта
            cell_value = str(cell.value)
            
            # Базовая ширина на основе количества символов
            # Учитываем, что разные символы имеют разную ширину
            text_width = 0
            for char in cell_value:
                if char.isupper() or char in 'БВГДЖЗКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ':
                    # Заглавные буквы и кириллица шире
                    text_width += 1.2
                elif char in 'ijl!|:;,.':
                    # Узкие символы
                    text_width += 0.5
                elif char in 'mwMWАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ':
                    # Широкие символы
                    text_width += 1.3
                else:
                    text_width += 1.0
            
            # Учитываем размер шрифта
            font_size = 11  # по умолчанию
            if cell.font and cell.font.size:
                font_size = cell.font.size
            
            # Корректируем ширину на размер шрифта
            # Базовая ширина для шрифта 11
            adjusted_width = text_width * (font_size / 11.0)
            
            # Учитываем жирный шрифт (он шире)
            if cell.font and cell.font.bold:
                adjusted_width *= 1.1
            
            # Обновляем максимум для колонки
            if adjusted_width > column_widths[col_letter]:
                column_widths[col_letter] = adjusted_width
    
    # Применяем ширины колонок
    for col_letter, max_width in column_widths.items():
        # Добавляем больше padding для гарантии, что текст влезет
        # Excel использует единицы измерения, где 1 ≈ ширина символа в шрифте Calibri 11
        final_width = min(max_width + 3.5, 80)  # Увеличен padding с 2.5 до 3.5
        ws.column_dimensions[col_letter].width = final_width
        print(f"Устанавливаем ширину колонки {col_letter}: {final_width:.1f} (макс ширина: {max_width:.1f})")

if __name__ == "__main__":
    test_autofit()
