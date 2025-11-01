# -*- coding: utf-8 -*-
"""
schedule_tab.py — Вкладка «Расписание»
--------------------------------------
Простое и мощное планирование с приоритетами
"""

import datetime as dt
import os
import json
import re
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from product_parse import parse_product_name, clear_product_parse_cache
import tkinter.font as tkFont
from schedule_excel_export import show_export_dialog

# CP-SAT импорты
try:
    from ortools.sat.python import cp_model
    CP_SAT_AVAILABLE = True
except ImportError:
    CP_SAT_AVAILABLE = False
    print("CP-SAT не доступен. Установите ortools: pip install ortools")

# ---------------------------------------------------------------------
# КОНСТАНТЫ И КЭШИ
# ---------------------------------------------------------------------

_THIS_DIR = os.path.dirname(__file__)
_PLAN_JSON = os.path.join(_THIS_DIR, "jobs_plan.json")
_SCHEDULE_JSON = os.path.join(_THIS_DIR, "schedule_data.json")
_CIP_THRESHOLDS_JSON = os.path.join(_THIS_DIR, "cip_thresholds.json")
_RULES_JSON = os.path.join(_THIS_DIR, "rules_sets.json")
_EVICTIONS_JSON = os.path.join(_THIS_DIR, "evictions_sets.json")
_NORMS_JSON = os.path.join(_THIS_DIR, "norms_data.json")
_DENSITY_JSON = os.path.join(_THIS_DIR, "product_density.json")

# Глобальные кэши
_TRANSITION_CACHE: dict[tuple, int] = {}  # (line, job_a_id, job_b_id) -> minutes
_RULES_CACHE: dict[str, dict] = None  # Загружается при первой необходимости
_EVICTIONS_CACHE: dict[str, dict] = None  # Кэш правил вытеснений
_EVICTIONS_CACHE_TIME: float = 0  # Время последней загрузки кэша вытеснений

SCHED_COLS = (
    "date", "line", "job_id", "name", "duration", "volume", "qty",
    "flavor", "brand", "type", "start", "end", "note"
)

# ---------------------------------------------------------------------
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ---------------------------------------------------------------------

def _load_json(file_path, default=None):
    """Загрузка JSON файла"""
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return default or []

def _save_json(file_path, data):
    """Сохранение JSON файла"""
    tmp = file_path + ".tmp"
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, file_path)

def _qty_to_int(s: str) -> int:
    """Конвертация количества"""
    try:
        return int(str(s).replace(" ", "").replace(",", ""))
    except:
        return 0

def _fmt_dt_dmy_hm(dt_obj: dt.datetime) -> str:
    """Форматирование даты/времени в DD.MM HH:MM"""
    return dt_obj.strftime("%d.%m %H:%M")

def _job_duration_minutes(job: dict) -> int:
    """Вычисление длительности работы в минутах"""
    try:
        qty = int(str(job.get("quantity", "0")).replace(" ", ""))
    except:
        qty = 0
    try:
        spd = float(str(job.get("speed", "1000")).replace(",", "."))
    except:
        spd = 1000.0
    if qty <= 0 or spd <= 0:
        return 0
    return int((60.0 * qty) / spd + 0.5)

def _load_evictions_cache():
    """Загрузка и кэширование правил вытеснений с автоматической перезагрузкой"""
    global _EVICTIONS_CACHE, _EVICTIONS_CACHE_TIME
    
    # Проверяем время модификации файла
    try:
        file_mtime = os.path.getmtime(_EVICTIONS_JSON)
        # Если файл изменился или кэш пустой - перезагружаем
        if _EVICTIONS_CACHE is None or file_mtime > _EVICTIONS_CACHE_TIME:
            _EVICTIONS_CACHE_TIME = file_mtime
            _EVICTIONS_CACHE = {}
            
            evictions_data = _load_json(_EVICTIONS_JSON, [])
            
            for ruleset in evictions_data:
                # Проверяем, какие линии относятся к этому набору правил
                lines = ruleset.get("lines", [])
                rules = ruleset.get("rules", [])
                
                # Строим словарь правил вытеснений для этого набора
                evictions_by_product = {}
                for rule in rules:
                    from_product = rule.get("from", "").strip().lower()
                    to_products_str = rule.get("to", "").strip().lower()
                    
                    # Разбиваем список целевых продуктов
                    to_products = set(prod.strip() for prod in to_products_str.split(";") if prod.strip())
                    
                    evictions_by_product[from_product] = to_products
                
                # Заполняем правила для всех линий этого набора
                for line_name in lines:
                    if line_name:
                        _EVICTIONS_CACHE[line_name.strip().lower()] = evictions_by_product
            
            print(f"[CACHE] Перезагружен кэш вытеснений из {_EVICTIONS_JSON}")
    
    except (OSError, FileNotFoundError):
        # Если файл недоступен, используем пустой кэш
        if _EVICTIONS_CACHE is None:
            _EVICTIONS_CACHE = {}
    
    return _EVICTIONS_CACHE


def _load_rules_cache():
    """Загрузка и кэширование правил CIP переходов"""
    global _RULES_CACHE
    if _RULES_CACHE is not None:
        return _RULES_CACHE
    
    _RULES_CACHE = {}
    rules_data = _load_json(_RULES_JSON, [])
    
    for ruleset in rules_data:
        # Проверяем, какие линии относятся к этому набору правил
        lines = ruleset.get("lines", [])
        rules = ruleset.get("rules", [])
        
        # Строим словарь правил для этого набора
        rules_by_product = {}
        for rule in rules:
            product = rule.get("product", "").strip().lower()
            cip1 = rule.get("CIP1", "").strip().lower()
            cip2 = rule.get("CIP2", "").strip().lower()
            cip3 = rule.get("CIP3", "").strip().lower()
            
            rules_by_product[product] = {
                "cip1": set(prod.strip().lower() for prod in cip1.split(";") if prod.strip()),
                "cip2": set(prod.strip().lower() for prod in cip2.split(";") if prod.strip()),
                "cip3": set(prod.strip().lower() for prod in cip3.split(";") if prod.strip()),
            }
        
        # Заполняем правила для всех линий этого набора
        for line_name in lines:
            if line_name:
                _RULES_CACHE[line_name] = rules_by_product
    
    return _RULES_CACHE


def _get_product_name(job: dict) -> str:
    """Извлечение названия продукта из задания с использованием product_parse"""
    name = job.get("name", "").strip()
    
    if name:
        # Используем parse_product_name для корректного парсинга
        parsed = parse_product_name(name)
        
        # Извлекаем flavor (вкус) и тип (сироп, нектар, сок и т.д.)
        flavor = parsed.get("flavor", "").strip()
        product_type = parsed.get("type", "").strip()
        
        if flavor and product_type:
            # В rules_sets.json формат: "нектар Манго", "сок Яблоко"
            # Формируем полное название: "нектар манго", "сок яблоко"
            return f"{product_type} {flavor}".lower()
        elif flavor:
            # Если есть только вкус, добавляем предположение о типе
            # Проверяем по исходному названию
            if "сироп" in name.lower():
                return f"сироп {flavor}".lower()
            elif "база" in name.lower():
                return f"база {flavor}".lower()
            else:
                return flavor.lower()
        else:
            # Если flavor не найден, берем базовое название без бренда
            parts = name.split(",")[0].strip()
            # Убираем ТМ и бренды
            parts = re.sub(r'\bТМ\b.*', '', parts).strip()
            parts = re.sub(r'\bTM\b.*', '', parts).strip()
            return parts.lower()
    
    return ""


def _check_eviction(line: str, product_a: str, product_b: str) -> bool:
    """
    Проверка наличия правила вытеснения для перехода между продуктами
    
    Args:
        line: название линии
        product_a: продукт "from" (в нижнем регистре)
        product_b: продукт "to" (в нижнем регистре)
    
    Returns:
        True если есть правило вытеснения, False иначе
    """
    evictions_cache = _load_evictions_cache()
    
    # Находим правила для данной линии
    evictions_for_line = None
    for line_name, evictions in evictions_cache.items():
        if line_name in line.lower() or line.lower() in line_name:
            evictions_for_line = evictions
            break
    
    if not evictions_for_line:
        return False
    
    # Проверяем есть ли правило: product_a → product_b
    if product_a in evictions_for_line:
        target_products = evictions_for_line[product_a]
        if product_b in target_products:
            return True
    
    return False


def _get_eviction_time(line: str) -> int:
    """
    Получение времени вытеснения для линии из norms_data.json
    
    Args:
        line: название линии
    
    Returns:
        Время вытеснения в минутах (по умолчанию 30)
    """
    norms_data = _load_json(_NORMS_JSON, [])
    
    for norm in norms_data:
        event = norm.get("event", "").strip().lower()
        if event == "выт" or event == "вытеснение":
            # Извлекаем номер линии
            line_num = None
            for i in range(1, 11):
                if str(i) in line:
                    line_num = i
                    break
            
            if line_num:
                line_key = f"line{line_num}"
                time_str = norm.get(line_key, "30").strip()
                try:
                    return int(time_str) if time_str else 30
                except:
                    return 30
    
    return 30  # Дефолтное значение


def _get_format_change_time(line: str) -> int:
    """
    Получение времени переналадки формата для линии из norms_data.json
    
    Приоритет: ищем запись с event="П", если нет - берем первую с "переналадка формата"
    
    Args:
        line: название линии
    
    Returns:
        Время переналадки в минутах (по умолчанию 120)
    """
    norms_data = _load_json(_NORMS_JSON, [])
    
    # Извлекаем номер линии (поддержка линий 1-10)
    line_num = None
    for i in range(1, 11):
        if str(i) in line or f"линия {i}" in line.lower() or f"line {i}" in line.lower():
            line_num = i
            break
    
    if not line_num:
        return 120  # Не смогли определить линию
    
    line_key = f"line{line_num}"
    
    # ПРИОРИТЕТ 1: Ищем запись с event="П"
    for norm in norms_data:
        event = norm.get("event", "").strip()
        category = norm.get("category", "").strip().lower()
        if event == "П" and "переналадка" in category:
            time_str = str(norm.get(line_key, "")).strip()
            if time_str:
                try:
                    return int(time_str)
                except:
                    pass
    
    # ПРИОРИТЕТ 2: Ищем любую "переналадка формата"
    for norm in norms_data:
        category = norm.get("category", "").strip().lower()
        if "переналадка" in category and "формат" in category:
            time_str = str(norm.get(line_key, "")).strip()
            if time_str:
                try:
                    return int(time_str)
                except:
                    pass
    
    return 120  # Дефолтное значение


def _get_volume_from_job(job: dict) -> str:
    """
    Извлекает объем тары из задания
    
    Args:
        job: словарь с данными задания
    
    Returns:
        Объем в виде строки (например "0,25 л" или "1,0 л")
    """
    volume = job.get("volume", "").strip()
    return volume.lower() if volume else ""


def _transition_time_estimate(line: str, job_a: dict, job_b: dict) -> tuple[int, str]:
    """Оценка времени перехода между работами на основе правил вытеснений и CIP
    
    Приоритет:
    0. ПЕРЕНАЛАДКА ФОРМАТА - если меняется объем тары (обязательно!)
    1. Проверяем правила ВЫТЕСНЕНИЙ (быстрее, заменяет CIP)
    2. Если нет вытеснения - используем правила CIP
    
    Returns:
        (время_в_минутах, тип_перехода) - например (30, "ВЫТ"), (240, "CIP2"), (120, "П")
    """
    # ПРИОРИТЕТ 0: Проверяем изменение объема тары (ПЕРЕНАЛАДКА ФОРМАТА)
    volume_a = _get_volume_from_job(job_a)
    volume_b = _get_volume_from_job(job_b)
    
    if volume_a and volume_b and volume_a != volume_b:
        format_time = _get_format_change_time(line)
        print(f"[FORMAT] Переналадка формата: {volume_a} -> {volume_b} на {line}: {format_time} мин")
        return (format_time, "П")
    
    # Получаем названия продуктов
    product_a = _get_product_name(job_a)
    product_b = _get_product_name(job_b)
    
    # ПРИОРИТЕТ 1: Проверяем ВЫТЕСНЕНИЕ (заменяет CIP!)
    if _check_eviction(line, product_a, product_b):
        eviction_time = _get_eviction_time(line)
        print(f"[EVICTION] Вытеснение: {product_a} -> {product_b} на {line}: {eviction_time} мин")
        return (eviction_time, "ВЫТ")
    
    # ПРИОРИТЕТ 2: Используем обычные правила CIP
    # Загружаем правила
    rules_cache = _load_rules_cache()
    
    # Находим соответствующее правило для линии
    rules_for_line = None
    for ruleset_name, rules in rules_cache.items():
        if ruleset_name.lower() in line.lower() or line.lower() in ruleset_name.lower():
            rules_for_line = rules
            break
    
    if not rules_for_line:
        return (40, "DEFAULT")  # Дефолтное значение
    
    # Определяем уровень CIP
    cip_level = None
    
    if product_a in rules_for_line:
        rules_a = rules_for_line[product_a]
        
        # Шаг 1: Находим БАЗОВЫЙ уровень (где указана "База")
        base_level = None
        if "база" in rules_a["cip1"]:
            base_level = "CIP1"
        elif "база" in rules_a["cip2"]:
            base_level = "CIP2"
        elif "база" in rules_a["cip3"]:
            base_level = "CIP3"
        
        # Шаг 2: Проверяем ИСКЛЮЧЕНИЯ - переходы, которые требуют другой уровень
        # Проверяем уровни, где НЕ указана "База"
        if "база" not in rules_a["cip1"] and product_b in rules_a["cip1"]:
            cip_level = "CIP1"  # Исключение в CIP1
        elif "база" not in rules_a["cip2"] and product_b in rules_a["cip2"]:
            cip_level = "CIP2"  # Исключение в CIP2
        elif "база" not in rules_a["cip3"] and product_b in rules_a["cip3"]:
            cip_level = "CIP3"  # Исключение в CIP3
        
        # Шаг 3: Если нет исключения, используем БАЗОВЫЙ уровень
        if not cip_level:
            cip_level = base_level
    
    # Если не нашлись правила - используем дефолт
    if not cip_level:
        return (40, "DEFAULT")
    
    # Загружаем нормы времени
    norms_data = _load_json(_NORMS_JSON, [])
    
    for norm in norms_data:
        event = norm.get("event", "").strip().lower()
        # Проверяем разные варианты названия CIP
        if event == cip_level.lower() or (cip_level == "CIP1" and event in ["сип1", "cip1"]) or \
           (cip_level == "CIP2" and event in ["сип2", "cip2"]) or \
           (cip_level == "CIP3" and event in ["сип3", "cip3"]):
            # Ищем время для конкретной линии
            # Извлекаем номер линии (поддержка линий 1-10)
            line_num = None
            for i in range(1, 11):
                if str(i) in line or f"линия {i}" in line.lower() or f"line {i}" in line.lower():
                    line_num = i
                    break
            
            if line_num:
                line_key = f"line{line_num}"
                if line_key in norm:
                    try:
                        time_val = int(str(norm[line_key]).strip())
                        return (time_val, cip_level)
                    except:
                        pass
    
    # Дефолтные значения в зависимости от уровня (из norms_data.json для Line5)
    if cip_level == "CIP1":
        return (40, "CIP1")
    elif cip_level == "CIP2":
        return (240, "CIP2")
    elif cip_level == "CIP3":
        return (300, "CIP3")
    
    return (40, "DEFAULT")


def _load_cip_thresholds():
    """Загрузка настроек порогов для автоматических CIP"""
    thresholds = _load_json(_CIP_THRESHOLDS_JSON, [])
    result = {}
    for threshold in thresholds:
        line = threshold.get("line", "")
        if line:
            # Сохраняем как по исходному ключу, так и по нормализованному названию линии
            result[line] = threshold
            try:
                m = re.search(r"(\d+)", str(line))
                if m:
                    norm_key = f"Линия {int(m.group(1))}"
                    result.setdefault(norm_key, threshold)
            except Exception:
                pass
    return result


def _load_product_density():
    """Загрузка справочника плотности продуктов"""
    return _load_json(_DENSITY_JSON, {})


def _calculate_mass(job: dict, density_map: dict) -> float:
    """Вычисляет массу продукта в кг
    
    Масса = количество × объем_единицы × плотность
    Например: 6000 шт × 1.0 л × 1.3 кг/л = 7800 кг
    """
    try:
        qty = _qty_to_int(job.get("quantity", "0"))
        if qty <= 0:
            return 0.0
        
        # Извлекаем объем из поля volume (например "1,0 л" -> 1.0)
        volume_str = job.get("volume", "")
        volume = 0.0
        if volume_str:
            # Ищем число в строке
            match = re.search(r'(\d+[,.]?\d*)', volume_str)
            if match:
                volume = float(match.group(1).replace(',', '.'))
        
        if volume <= 0:
            return float(qty)  # Если объем неизвестен, возвращаем штуки
        
        # Получаем плотность по типу продукта
        product_type = job.get("type", "")
        density = density_map.get(product_type, 1.0)  # По умолчанию 1.0 кг/л
        
        # Масса в кг
        mass_kg = qty * volume * density
        return mass_kg
        
    except:
        # В случае ошибки возвращаем количество штук
        return float(_qty_to_int(job.get("quantity", "0")))


def _get_cip_duration_for_type(line: str, cip_type: str) -> int:
    """Получить длительность CIP по типу и линии"""
    norms_data = _load_json(_NORMS_JSON, [])
    
    for norm in norms_data:
        event = norm.get("event", "").strip().lower()
        if (cip_type.lower() == event) or \
           (cip_type == "CIP1" and event in ["сип1", "cip1"]) or \
           (cip_type == "CIP2" and event in ["сип2", "cip2"]) or \
           (cip_type == "CIP3" and event in ["сип3", "cip3"]):
            
            line_num = None
            for i in range(1, 11):
                if str(i) in line:
                    line_num = i
                    break
            
            if line_num:
                line_key = f"line{line_num}"
                if line_key in norm:
                    try:
                        return int(str(norm[line_key]).strip())
                    except:
                        pass
    
    # Дефолтные значения
    if cip_type == "CIP1":
        return 40
    elif cip_type == "CIP2":
        return 240
    elif cip_type == "CIP3":
        return 300
    
    return 240


def _check_cip_conflict(results: list[dict], start_time: int, duration: int, line: str) -> bool:
    """Проверяет, есть ли конфликт с существующими CIP в заданном интервале"""
    end_time = start_time + duration

    for record in results:
        if record.get("line") != line:
            continue

        # Проверяем только CIP записи (не рабочие задания)
        if record.get("type") not in ["CIP1", "CIP2", "CIP3", "CIP", "ВЫТ", "П"]:
            continue

        # Получаем время начала и окончания CIP
        start_str = record.get("start", "")
        end_str = record.get("end", "")

        try:
            # Парсим время (формат: "25.10 14:30")
            start_match = re.search(r"(\d{1,2})\.(\d{1,2})\s+(\d{2}):(\d{2})", start_str)
            end_match = re.search(r"(\d{1,2})\.(\d{1,2})\s+(\d{2}):(\d{2})", end_str)

            if start_match and end_match:
                start_minutes = int(start_match.group(3)) * 60 + int(start_match.group(4))
                end_minutes = int(end_match.group(3)) * 60 + int(end_match.group(4))

                # Проверяем пересечение интервалов
                if not (end_time <= start_minutes or start_time >= end_minutes):
                    return True  # Есть конфликт
        except:
            continue

    return False  # Нет конфликта


def _find_next_free_slot(results: list[dict], base_dt: dt.datetime, start_time: int,
                        duration: int, line: str, max_shift: int = 24*60) -> int:
    """Находит следующий свободный слот для CIP, пропуская существующие CIP"""
    current_time = start_time

    while current_time < max_shift:
        if not _check_cip_conflict(results, current_time, duration, line):
            return current_time  # Найден свободный слот
        current_time += duration  # Пропускаем занятый слот

    return start_time  # Если не нашли свободный слот, возвращаем исходное время


def _create_auto_cip_record(line: str, cip_type: str, base_dt: dt.datetime, t: int,
                             prev_job: dict, next_job: dict, accumulated_volume: float, unit: str = "шт") -> dict:
    """Создать запись автоматического CIP"""
    transition = _get_cip_duration_for_type(line, cip_type)
    
    prev_name_short = prev_job.get('name', '')[:30].strip() if prev_job else "Начало"
    next_name_short = next_job.get('name', '')[:30].strip() if next_job else "Конец"
    cip_name = f"{cip_type} по объему ({accumulated_volume:.0f} {unit})"
    
    cip_start_dt = base_dt + dt.timedelta(minutes=t)
    cip_end_dt = base_dt + dt.timedelta(minutes=t + transition)
    
    job_id = next_job.get('job_id', '') if next_job else prev_job.get('job_id', '')
    
    return {
        "date": cip_start_dt.date().isoformat(),
        "line": line,
        "job_id": f"AUTO-CIP-{job_id}",
        "name": cip_name,
        "duration": str(transition),
        "volume": "",
        "qty": "",
        "flavor": "",
        "brand": "",
        "type": cip_type,
        "start": _fmt_dt_dmy_hm(cip_start_dt),
        "end": _fmt_dt_dmy_hm(cip_end_dt),
        "note": f"Автоматический {cip_type} по объему ({accumulated_volume:.0f} {unit})",
        "_auto_cip": True
    }


def _group_by_shifts(schedule_data: list[dict]) -> dict[str, dict[str, list[dict]]]:
    """Группирует данные расписания по сменам"""
    grouped = {}
    
    for event in schedule_data:
        line = event.get("line", "—")
        if line not in grouped:
            grouped[line] = {}
        
        # Определяем смену по времени начала
        start_str = event.get("start", "")
        date_str = event.get("date", "")
        
        # Парсим время
        m = re.search(r"(\d{1,2})\.(\d{1,2})\s+(\d{2}):(\d{2})", start_str)
        if not m:
            continue

        day, month, hour, minute = map(int, m.groups())
        
        # Определяем дату
        if date_str:
            date_obj = dt.datetime.strptime(date_str, "%Y-%m-%d").date()
        else:
            year = dt.date.today().year
            if month == 1 and dt.date.today().month == 12:
                year += 1
            date_obj = dt.date(year, month, day)
        
        start_dt = dt.datetime.combine(date_obj, dt.time(hour, minute))
        
        # Определяем смену
        if 8 <= hour < 20:
            shift_name = "Дневная"
        else:
            shift_name = "Ночная"
        
        # Для ночной смены корректируем дату
        if hour >= 20:
            shift_date = (date_obj + dt.timedelta(days=1)).isoformat()
        elif hour < 8:
            shift_date = (date_obj - dt.timedelta(days=1)).isoformat()
        else:
            shift_date = date_obj.isoformat()
        
        shift_key = f"{shift_date}_{shift_name}"
        
        if shift_key not in grouped[line]:
            grouped[line][shift_key] = []
        
        grouped[line][shift_key].append(event)
    
    return grouped

# ---------------------------------------------------------------------
# ОСНОВНАЯ ФУНКЦИЯ ПОСТРОЕНИЯ РАСПИСАНИЯ
# ---------------------------------------------------------------------

def build_schedule_from_plan(shift_date=None, shift_start="08:00", fix_priorities: list[int] = [1], 
                            use_cp_sat: bool = False, line_bindings: dict = None) -> list[dict]:
    """
    Построение расписания из jobs_plan.json
    
    Args:
        shift_date: Дата смены (по умолчанию сегодня)
        shift_start: Время начала смены
        fix_priorities: Список приоритетов для фиксации порядка
    
    Returns:
        Список расписаний для всех линий
    """
    # Загружаем план
    plan = _load_json(_PLAN_JSON, [])
    if not plan:
        messagebox.showwarning("Расписание", "Файл jobs_plan.json пустой или не найден")
        return []

    # Предварительная обработка
    jobs_by_line = _preprocess_plan_data(plan)
    if not jobs_by_line:
        messagebox.showinfo("Расписание", "Нет заданий для расписания")
        return []

    # Строим расписание для каждой линии с учетом привязок
    line_schedules = {}
    all_results = []
    
    for line, jobs in jobs_by_line.items():
        result = _build_schedule_for_line(line, jobs, shift_date, shift_start, fix_priorities, use_cp_sat)
        if result:
            line_schedules[line] = result
    
    # Применяем привязки линий
    if line_bindings:
        for target_line, source_line in line_bindings.items():
            if source_line in line_schedules and target_line in line_schedules:
                source_schedule = line_schedules[source_line]
                target_schedule = line_schedules[target_line]
                if source_schedule and target_schedule:
                    # Находим время окончания исходной линии
                    last_end = source_schedule[-1].get("end", "")
                    m = re.search(r"(\d{1,2})\.(\d{1,2})\s+(\d{2}):(\d{2})", last_end)
                    if m:
                        day, month, hour, minute = map(int, m.groups())
                        year = dt.date.today().year
                        if month == 1 and dt.date.today().month == 12:
                            year += 1
                        source_end_dt = dt.datetime(year, month, day, hour, minute)
                        
                        # Находим время начала целевой линии
                        first_start = target_schedule[0].get("start", "")
                        n = re.search(r"(\d{1,2})\.(\d{1,2})\s+(\d{2}):(\d{2})", first_start)
                        if n:
                            t_day, t_month, t_hour, t_minute = map(int, n.groups())
                            t_year = dt.date.today().year
                            if t_month == 1 and dt.date.today().month == 12:
                                t_year += 1
                            target_start_dt = dt.datetime(t_year, t_month, t_day, t_hour, t_minute)
                            
                            # Вычисляем разницу
                            time_diff = (source_end_dt - target_start_dt).total_seconds() / 60
                            
                            # Сдвигаем целевую линию
                            if time_diff != 0:
                                for record in target_schedule:
                                    for time_key in ["start", "end"]:
                                        time_str = record.get(time_key, "")
                                        s = re.search(r"(\d{1,2})\.(\d{1,2})\s+(\d{2}):(\d{2})", time_str)
                                        if s:
                                            d, m, h, mi = map(int, s.groups())
                                            y = dt.date.today().year
                                            if m == 1 and dt.date.today().month == 12:
                                                y += 1
                                            dt_obj = dt.datetime(y, m, d, h, mi)
                                            dt_obj += dt.timedelta(minutes=time_diff)
                                            record[time_key] = _fmt_dt_dmy_hm(dt_obj)
                                            if time_key == "start":
                                                record["date"] = dt_obj.date().isoformat()
                                
                                print(f"Привязано: {source_line} -> {target_line}, сдвиг {time_diff:.0f} мин")
    
    # Объединяем все расписания
    for line_schedule in line_schedules.values():
        all_results.extend(line_schedule)
    
    print(f"\nПостроено {len(all_results)} записей расписания")
    return all_results


def _preprocess_plan_data(plan: list[dict]) -> dict[str, list[dict]]:
    """Предварительная обработка плана - группировка по линиям"""
    jobs_by_line = {}
    
    for job in plan:
        # Нормализуем название линии к виду "Линия N" для единообразия
        raw_line = str(job.get("line", "")).strip()
        m = re.search(r"(\d+)", raw_line)
        line = f"Линия {int(m.group(1))}" if m else raw_line
        qty = _qty_to_int(job.get("quantity", ""))
        status = job.get("status", "").strip().lower()

        # Фильтруем задания с некорректными данными
        if not line or qty <= 0:
                continue

        # Отфильтровываем отложенные задания (не включаем в расписание)
        if status in ("postponed", "отложено", "отложен"):
            continue

        # Отфильтровываем завершенные (если есть факт)
        fact_qty = _qty_to_int(job.get("fact_qty", ""))
        if fact_qty >= qty:
            continue  # Задание уже выполнено
        
        # Остаток к выполнению
        remaining = qty - fact_qty
        job_copy = job.copy()
        job_copy["quantity"] = str(remaining)
        
        if line not in jobs_by_line:
            jobs_by_line[line] = []
        jobs_by_line[line].append(job_copy)
    
    return jobs_by_line


def _build_schedule_for_line(line: str, jobs: list[dict], shift_date=None, 
                             shift_start="08:00", fix_priorities: list[int] = None, 
                             use_cp_sat: bool = False) -> list[dict]:
    """Построение расписания для одной линии"""
    if not jobs:
        return []
    
    # Сохраняем исходный порядок
    for idx, job in enumerate(jobs):
        job["_original_index"] = idx
    
    # Сортируем по приоритету
    jobs = _sort_by_priority(jobs)
    
    # Применяем CP-SAT оптимизацию если включена
    if use_cp_sat and CP_SAT_AVAILABLE:
        print(f"Применяем CP-SAT оптимизацию для линии {line}")
        print(f"Заблокированные приоритеты: {fix_priorities}")
        jobs = _optimize_with_cp_sat(line, jobs, fix_priorities)
    
    # Загружаем пороги для автоматических CIP
    cip_thresholds = _load_cip_thresholds()
    threshold_config = cip_thresholds.get(line, {})
    auto_cip_enabled = threshold_config.get("enabled", False)
    volume_threshold = threshold_config.get("volume_threshold", 50000)
    product_threshold = threshold_config.get("product_threshold", 30000)
    min_remainder = threshold_config.get("min_remainder", 2000)  # Минимальный остаток
    auto_cip_type = threshold_config.get("cip_type", "CIP2")
    auto_cip_mode = threshold_config.get("mode", "штуки")  # "штуки" или "масса"
    
    # Загружаем справочник плотности если режим по массе
    density_map = {}
    if auto_cip_mode == "масса":
        density_map = _load_product_density()
    
    unit = "шт" if auto_cip_mode == "штуки" else "кг"
    
    if auto_cip_enabled:
        print(f"[АВТО-CIP] {line}: включено | Режим: {auto_cip_mode} | Порог общий: {volume_threshold} {unit} | Порог продукта: {product_threshold} {unit} | Буфер: {min_remainder} {unit} (применяется при риске частых CIP) | Тип: {auto_cip_type}")
    
    # Строим расписание последовательно
    date_str = shift_date or dt.date.today().isoformat()
    h0, m0 = map(int, (shift_start or "08:00").split(":"))
    base_dt = dt.datetime.strptime(date_str, "%Y-%m-%d").replace(hour=h0, minute=m0)

    t = 0
    results = []
    prev_job = None
    
    # Счетчики для автоматических CIP
    total_volume_since_cip = 0
    current_product_volume = 0
    current_product_name = None
    last_auto_cip_index = -1
    
    for idx, job in enumerate(jobs):
        # Добавляем CIP переход между работами (не перед первой работой)
        if prev_job is not None:
            # Проверяем, не был ли последней записью автосип - если да, обычный переход не нужен
            skip_normal_transition = False
            if results and len(results) > 0:
                last_record = results[-1]
                last_job_id = last_record.get("job_id", "")
                # Если последняя запись - автосип, пропускаем обычный переход
                if last_job_id.startswith("AUTO-CIP-"):
                    skip_normal_transition = True
                    print(f"[АВТО-CIP] Пропуск обычного перехода после автосипа: {last_job_id}")
            
            if not skip_normal_transition:
                transition, cip_type = _transition_time_estimate(line, prev_job, job)
                
                if transition > 0:
                    # Добавляем CIP как отдельную запись в расписание
                    prev_name_short = prev_job.get('name', '')[:30].strip()
                    next_name_short = job.get('name', '')[:30].strip()
                    cip_name = f"{prev_name_short} -> {next_name_short}"
                    
                    # Времена CIP
                    cip_start_dt = base_dt + dt.timedelta(minutes=t)
                    cip_end_dt = base_dt + dt.timedelta(minutes=t + transition)
                    
                    # Определяем префикс ID в зависимости от типа перехода
                    if cip_type == "ВЫТ":
                        id_prefix = "ВЫТ"
                    elif cip_type == "П":
                        id_prefix = "П"
                    else:
                        id_prefix = "CIP"
                    
                    cip_record = {
                        "date": cip_start_dt.date().isoformat(),
                        "line": line,
                        "job_id": f"{id_prefix}-{job.get('job_id', '')}",
                        "name": cip_name,
                        "duration": str(transition),
                        "volume": "",
                        "qty": "",
                        "flavor": "",
                        "brand": "",
                        "type": cip_type,
                        "start": _fmt_dt_dmy_hm(cip_start_dt),
                        "end": _fmt_dt_dmy_hm(cip_end_dt),
                        "note": f"{cip_type} ({transition} мин)",
                    }
                    
                    results.append(cip_record)
                    t += transition
                    
                    # Сбрасываем счетчики после CIP и переналадки (не вытеснения)
                    if cip_type not in ("ВЫТ",):
                        if cip_type == "П":
                            print(f"[АВТО-CIP] Сброс счетчиков после переналадки формата на {line}")
                        total_volume_since_cip = 0
                        current_product_volume = 0
                        current_product_name = None
                        last_auto_cip_index = len(results) - 1
        
        # Получаем количество текущей работы (в штуках или кг)
        job_qty_pieces = _qty_to_int(job.get("quantity", "0"))
        job_name = _get_product_name(job)
        
        # Вычисляем значение для сравнения с порогом (штуки или масса)
        if auto_cip_mode == "масса":
            job_qty = _calculate_mass(job, density_map)  # в кг
        else:
            job_qty = float(job_qty_pieces)  # в штуках
        
        # Разбиваем работу на части если нужно
        remaining_qty = job_qty
        remaining_qty_pieces = job_qty_pieces  # Для расчета длительности
        part_number = 1
        
        if auto_cip_enabled and job_qty > product_threshold:
            print(f"[АВТО-CIP] Работа '{job_name}' {job_qty:.0f} {unit} > порог {product_threshold} {unit} - будет разбита")
        
        while remaining_qty > 0:
            # Определяем размер текущей части
            need_auto_cip = False
            auto_cip_reason = ""
            current_part_qty = remaining_qty
            
            if auto_cip_enabled:
                # Проверка 1: Превышен общий объем
                if prev_job is not None and total_volume_since_cip + remaining_qty > volume_threshold:
                    # Рассчитываем остаток после CIP на обычном пороге
                    remainder_after_cip = total_volume_since_cip + remaining_qty - volume_threshold

                    # Если остаток слишком мал, применяем буфер
                    if remainder_after_cip < min_remainder:
                        # Сдвигаем CIP вперед, чтобы после него осталось min_remainder
                        buffer_needed = min_remainder - remainder_after_cip
                        effective_threshold = volume_threshold + buffer_needed
                        current_part_qty = effective_threshold - total_volume_since_cip
                        auto_cip_reason = f"Общий объем достиг {effective_threshold} {unit} (буфер {buffer_needed} для предотвращения частых CIP)"
                    else:
                        # Обычный случай - остаток достаточный
                        current_part_qty = volume_threshold - total_volume_since_cip
                        auto_cip_reason = f"Общий объем достиг {volume_threshold} {unit}"

                    if current_part_qty <= 0:
                        current_part_qty = min(remaining_qty, volume_threshold)
                    need_auto_cip = True
                
                # Проверка 2: Превышен объем по продукту
                if not need_auto_cip:
                    if job_name == current_product_name:
                        # Продолжаем тот же продукт
                        if current_product_volume + remaining_qty > product_threshold:
                            # Рассчитываем остаток после CIP на обычном пороге
                            remainder_after_cip = current_product_volume + remaining_qty - product_threshold

                            # Если остаток слишком мал, применяем буфер
                            if remainder_after_cip < min_remainder:
                                buffer_needed = min_remainder - remainder_after_cip
                                effective_threshold = product_threshold + buffer_needed
                                current_part_qty = effective_threshold - current_product_volume
                                auto_cip_reason = f"Объем продукта '{job_name}' достиг {effective_threshold} {unit} (буфер {buffer_needed} для предотвращения частых CIP)"
                            else:
                                # Обычный случай
                                current_part_qty = product_threshold - current_product_volume
                                auto_cip_reason = f"Объем продукта '{job_name}' достиг {product_threshold} {unit}"

                            if current_part_qty <= 0:
                                current_part_qty = min(remaining_qty, product_threshold)
                            need_auto_cip = True
                    else:
                        # Новый продукт - проверяем обычный порог
                        if remaining_qty > product_threshold:
                            # Рассчитываем остаток после CIP
                            remainder_after_cip = remaining_qty - product_threshold

                            if remainder_after_cip < min_remainder:
                                # Применяем буфер
                                buffer_needed = min_remainder - remainder_after_cip
                                effective_threshold = product_threshold + buffer_needed
                                current_part_qty = effective_threshold
                                auto_cip_reason = f"Работа '{job_name}' требует буфера {buffer_needed} для предотвращения частых CIP"
                            else:
                                # Обычный случай
                                current_part_qty = product_threshold
                                auto_cip_reason = f"Работа '{job_name}' превышает порог {product_threshold} {unit}"

                            need_auto_cip = True
            
            # Пересчитываем части в штуки для отображения
            if auto_cip_mode == "масса" and job_qty_pieces > 0:
                # Пропорция: сколько штук в текущей части
                current_part_qty_pieces = int((current_part_qty / job_qty) * job_qty_pieces)
                if current_part_qty_pieces <= 0:
                    current_part_qty_pieces = max(1, remaining_qty_pieces)
            else:
                current_part_qty_pieces = int(current_part_qty)
            
            # Вычисляем длительность части (по штукам)
            if current_part_qty_pieces > 0:
                try:
                    spd = float(str(job.get("speed", "1000")).replace(",", "."))
                except:
                    spd = 1000.0
                if spd <= 0:
                    spd = 1000.0
                part_dur = int((60.0 * current_part_qty_pieces) / spd + 0.5)
            else:
                part_dur = 0
            
            # Добавляем часть работы
            if current_part_qty > 0:
                start_dt = base_dt + dt.timedelta(minutes=t)
                end_dt = base_dt + dt.timedelta(minutes=t + part_dur)
                
                # Формируем название с номером части если работа разбита
                job_note = job.get("note", "")
                if job_qty > current_part_qty or part_number > 1:
                    job_note = f"Часть {part_number} из {int((job_qty + product_threshold - 1) // product_threshold)}" + (f" | {job_note}" if job_note else "")
                
                results.append({
                    "date": start_dt.date().isoformat(),
                    "line": line,
                    "job_id": job.get("job_id", "") + (f"-P{part_number}" if part_number > 1 else ""),
                    "name": job.get("name", ""),
                    "duration": str(part_dur),
                    "volume": job.get("volume", ""),
                    "qty": str(current_part_qty_pieces),  # Количество в штуках
                    "flavor": job.get("flavor", ""),
                    "brand": job.get("brand", ""),
                    "type": job.get("type", ""),
                    "start": _fmt_dt_dmy_hm(start_dt),
                    "end": _fmt_dt_dmy_hm(end_dt),
                    "note": job_note,
                })
                
                # Обновляем счетчики
                total_volume_since_cip += current_part_qty
                if job_name == current_product_name:
                    current_product_volume += current_part_qty
                else:
                    current_product_name = job_name
                    current_product_volume = current_part_qty
                
                t += part_dur
                remaining_qty -= current_part_qty
                remaining_qty_pieces -= current_part_qty_pieces
                part_number += 1
                prev_job = job
            
            # Вставляем автоматический CIP если нужно
            if need_auto_cip and remaining_qty > 0 and last_auto_cip_index != len(results) - 1:
                # Проверяем обязательный CIP перед вставкой автосипа
                skip_auto_cip_due_to_required = False
                
                # Проверяем, есть ли следующая работа и какой CIP требуется по правилам
                if idx + 1 < len(jobs):
                    next_job = jobs[idx + 1]
                    # Проверяем переход от ТЕКУЩЕЙ работы (job) к следующей
                    _, required_cip_type_from_rules = _transition_time_estimate(line, job, next_job)
                    
                    job_name_short = job.get("name", "")[:30]
                    next_job_name_short = next_job.get("name", "")[:30]
                    print(f"[АВТО-CIP] Проверка обязательного CIP для перехода {job_name_short} -> {next_job_name_short}: {required_cip_type_from_rules}")

                    # Определяем жесткость CIP (CIP3 > CIP2 > CIP1)
                    def get_cip_priority(cip_type):
                        if cip_type == "CIP3":
                            return 3
                        elif cip_type == "CIP2":
                            return 2
                        elif cip_type == "CIP1":
                            return 1
                        else:
                            return 0  # ВЫТ, П - не учитываются

                    required_priority = get_cip_priority(required_cip_type_from_rules)
                    auto_cip_priority = get_cip_priority(auto_cip_type)

                    print(f"[АВТО-CIP] Сравнение приоритетов (внутри разбиения): обязательный CIP = {required_cip_type_from_rules} (приоритет {required_priority}), автосип = {auto_cip_type} (приоритет {auto_cip_priority})")

                    # Если обязательный CIP >= автосипа по жесткости - пропускаем автосип
                    if required_priority >= auto_cip_priority and required_priority > 0:
                        skip_auto_cip_due_to_required = True
                        print(f"[АВТО-CIP] Пропуск автосипа {auto_cip_type}: обязательный табличный CIP {required_cip_type_from_rules} имеет приоритет (переход {job_name_short} -> {next_job_name_short})")
                
                if not skip_auto_cip_due_to_required:
                    # Получаем длительность автосопровождения
                    auto_cip_duration = _get_cip_duration_for_type(line, auto_cip_type)

                    # Ищем свободный слот, пропуская существующие CIP
                    free_slot_time = _find_next_free_slot(results, base_dt, t, auto_cip_duration, line)

                    if free_slot_time != t:
                        print(f"[АВТО-CIP] Вставка {auto_cip_type} на {line}: {auto_cip_reason}")
                        print(f"[АВТО-CIP] Конфликт с существующим CIP, сдвигаем с {t} мин на {free_slot_time} мин")
                    else:
                        print(f"[АВТО-CIP] Вставка {auto_cip_type} на {line}: {auto_cip_reason}")

                    auto_cip = _create_auto_cip_record(
                        line, auto_cip_type, base_dt, free_slot_time, prev_job, job, total_volume_since_cip, unit
                    )
                    results.append(auto_cip)

                    # Обновляем время и счетчики
                    t = free_slot_time + auto_cip_duration  # Устанавливаем время после вставленного CIP
                    total_volume_since_cip = 0
                    current_product_volume = 0
                    current_product_name = None
                    last_auto_cip_index = len(results) - 1
                else:
                    # Отменяем автосип, так как обязательный CIP имеет приоритет
                    need_auto_cip = False
            
            # Если не нужен CIP и остаток есть - выходим (ошибка логики)
            if not need_auto_cip:
                break
        
        # Проверяем автосип ПОСЛЕ завершения работы, если накопилось достаточно
        if auto_cip_enabled and remaining_qty == 0:
            # Работа полностью обработана - проверяем накопленный объем
            if total_volume_since_cip >= volume_threshold or (current_product_volume >= product_threshold):
                print(f"[АВТО-CIP] Работа завершена, проверка автосипа: общий объем = {total_volume_since_cip} {unit}, объем продукта = {current_product_volume} {unit}, порог объема = {volume_threshold} {unit}, порог продукта = {product_threshold} {unit}")
                
                # Проверяем, есть ли следующая работа и какой CIP требуется по правилам
                skip_auto_cip = False
                required_cip_type = None
                
                if idx + 1 < len(jobs):
                    next_job = jobs[idx + 1]
                    # Проверяем переход от ТЕКУЩЕЙ работы (job) к следующей, а не от prev_job!
                    _, required_cip_type_from_rules = _transition_time_estimate(line, job, next_job)
                    
                    job_name_short = job.get("name", "")[:30]
                    next_job_name_short = next_job.get("name", "")[:30]
                    print(f"[АВТО-CIP] Проверка обязательного CIP для перехода {job_name_short} -> {next_job_name_short}: {required_cip_type_from_rules}")

                    # Определяем жесткость CIP (CIP3 > CIP2 > CIP1)
                    def get_cip_priority(cip_type):
                        if cip_type == "CIP3":
                            return 3
                        elif cip_type == "CIP2":
                            return 2
                        elif cip_type == "CIP1":
                            return 1
                        else:
                            return 0  # ВЫТ, П - не учитываются

                    required_priority = get_cip_priority(required_cip_type_from_rules)
                    auto_cip_priority = get_cip_priority(auto_cip_type)

                    print(f"[АВТО-CIP] Сравнение приоритетов (после завершения): обязательный CIP = {required_cip_type_from_rules} (приоритет {required_priority}), автосип = {auto_cip_type} (приоритет {auto_cip_priority})")

                    # Если обязательный CIP >= автосипа по жесткости - пропускаем автосип
                    if required_priority >= auto_cip_priority and required_priority > 0:
                        skip_auto_cip = True
                        required_cip_type = required_cip_type_from_rules
                        print(f"[АВТО-CIP] ✓ ПРОПУСК автосипа {auto_cip_type}: обязательный табличный CIP {required_cip_type_from_rules} имеет приоритет (переход {job_name_short} -> {next_job_name_short})")
                    else:
                        print(f"[АВТО-CIP] ✗ НЕ пропускаем автосип: обязательный CIP {required_cip_type_from_rules} (приоритет {required_priority}) < автосип {auto_cip_type} (приоритет {auto_cip_priority})")
                else:
                    print(f"[АВТО-CIP] Нет следующей работы (idx={idx}, len(jobs)={len(jobs)})")
                
                if not skip_auto_cip:
                    print(f"[АВТО-CIP] ✓ ВСТАВКА автосипа {auto_cip_type} после завершения работы")
                    auto_cip_duration = _get_cip_duration_for_type(line, auto_cip_type)
                    free_slot_time = _find_next_free_slot(results, base_dt, t, auto_cip_duration, line)
                    
                    if free_slot_time != t:
                        print(f"[АВТО-CIP] Конфликт с существующим CIP, сдвигаем с {t} мин на {free_slot_time} мин")
                    
                    auto_cip_reason = f"Объем достиг порога (общий: {total_volume_since_cip} {unit}, продукт: {current_product_volume} {unit})"
                    auto_cip = _create_auto_cip_record(
                        line, auto_cip_type, base_dt, free_slot_time, prev_job, None, total_volume_since_cip, unit
                    )
                    results.append(auto_cip)
                    print(f"[АВТО-CIP] Автосип {auto_cip_type} добавлен в расписание: {auto_cip.get('name', '')}")
                    
                    # Обновляем время и счетчики только если автосип был вставлен
                    t = free_slot_time + auto_cip_duration
                    total_volume_since_cip = 0
                    current_product_volume = 0
                    current_product_name = None
                    last_auto_cip_index = len(results) - 1
                else:
                    print(f"[АВТО-CIP] ✓ АВТОСИП ПРОПУЩЕН, будет использован обязательный CIP {required_cip_type}")

    return results


def _split_jobs_across_shifts(schedule: list[dict]) -> list[dict]:
    """Разрезает выпуски, которые переходят через смены"""
    split_schedule = []
    
    for record in schedule:
        start_str = record.get("start", "")
        m = re.search(r"(\d{1,2})\.(\d{1,2})\s+(\d{2}):(\d{2})", start_str)
        if not m:
            split_schedule.append(record)
            continue
            
        day, month, hour, minute = map(int, m.groups())
        year = dt.date.today().year
        if month == 1 and dt.date.today().month == 12:
            year += 1
            
        start_dt = dt.datetime(year, month, day, hour, minute)
        duration = int(record.get("duration", "0") or "0")
        end_dt = start_dt + dt.timedelta(minutes=duration)
        
        # Определяем границы смен
        if 8 <= start_dt.hour < 20:
            # Дневная смена: 8:00-20:00
            shift_end = start_dt.replace(hour=20, minute=0)
        else:
            # Ночная смена: 20:00-8:00
            if start_dt.hour >= 20:
                shift_end = start_dt.replace(hour=23, minute=59) + dt.timedelta(days=1)
                shift_end = shift_end.replace(hour=8, minute=0)
            else:  # hour < 8
                shift_end = start_dt.replace(hour=8, minute=0)
        
        # Если работа не выходит за границы смены
        if end_dt <= shift_end:
            split_schedule.append(record)
        else:
            # Работа выходит за границы смены - разрезаем
            remaining_duration = duration
            
            # Первая часть (в текущей смене)
            first_part_duration = int((shift_end - start_dt).total_seconds() / 60)
            if first_part_duration > 0:
                first_part = record.copy()
                first_part["duration"] = str(first_part_duration)
                first_part["end"] = _fmt_dt_dmy_hm(shift_end)
                
                # Пропорционально распределяем количество
                if record.get("qty") and record.get("qty").strip():
                    try:
                        total_qty = int(record.get("qty", "0").replace(",", "").replace(" ", ""))
                        first_part_qty = int((first_part_duration / duration) * total_qty)
                        first_part["qty"] = str(first_part_qty)
                    except (ValueError, ZeroDivisionError):
                        pass
                
                first_part["note"] = (first_part.get("note", "") + " [часть 1]").strip()
                split_schedule.append(first_part)
                remaining_duration -= first_part_duration
            
            # Вторая часть (в следующей смене)
            if remaining_duration > 0:
                next_shift_start = shift_end
                second_part = record.copy()
                second_part["duration"] = str(remaining_duration)
                second_part["start"] = _fmt_dt_dmy_hm(next_shift_start)
                second_part_end = next_shift_start + dt.timedelta(minutes=remaining_duration)
                second_part["end"] = _fmt_dt_dmy_hm(second_part_end)
                second_part["date"] = next_shift_start.date().isoformat()
                
                # Пропорционально распределяем количество
                if record.get("qty") and record.get("qty").strip():
                    try:
                        total_qty = int(record.get("qty", "0").replace(",", "").replace(" ", ""))
                        second_part_qty = total_qty - int((first_part_duration / duration) * total_qty)
                        second_part["qty"] = str(second_part_qty)
                    except (ValueError, ZeroDivisionError):
                        pass
                
                second_part["note"] = (second_part.get("note", "") + " [часть 2]").strip()
                split_schedule.append(second_part)
    
    return split_schedule


def _group_schedule_by_shifts(schedule: list[dict]) -> dict[str, list[dict]]:
    """Группирует расписание по сменам"""
    shifts = {}
    
    for record in schedule:
        start_str = record.get("start", "")
        m = re.search(r"(\d{1,2})\.(\d{1,2})\s+(\d{2}):(\d{2})", start_str)
        if not m:
            continue
            
        day, month, hour, minute = map(int, m.groups())
        year = dt.date.today().year
        if month == 1 and dt.date.today().month == 12:
            year += 1
            
        start_dt = dt.datetime(year, month, day, hour, minute)
        
        # Определяем смену (с 8:00 до 8:00)
        if 8 <= hour < 20:
            shift_name = "Дневная"
            shift_date = start_dt.date()
        else:
            shift_name = "Ночная"
            if hour >= 20:
                # Ночная смена начинается в 20:00 того же дня
                shift_date = start_dt.date()
            else:  # hour < 8
                # Ночная смена продолжается до 8:00 следующего дня
                shift_date = start_dt.date() - dt.timedelta(days=1)
        
        shift_key = f"{shift_date.isoformat()}_{shift_name}"
        
        if shift_key not in shifts:
            shifts[shift_key] = []
        shifts[shift_key].append(record)
    
    return shifts


def _optimize_with_cp_sat(line: str, jobs: list[dict], fix_priorities: list[int] = None) -> list[dict]:
    """
    CP-SAT оптимизация порядка работ с учетом приоритетов.
    
    Стратегия:
    1. Работы сортируются по приоритетам (жесткое ограничение)
    2. Внутри каждой группы приоритетов оптимизируется порядок по времени переходов CIP
    3. Для ЗАБЛОКИРОВАННЫХ приоритетов (fix_priorities) порядок НЕ меняется
    4. Используется взвешенная целевая функция для балансировки приоритетов и времени
    """
    if not CP_SAT_AVAILABLE or len(jobs) < 2:
        return jobs
    
    if fix_priorities is None:
        fix_priorities = []
    
    try:
        model = cp_model.CpModel()
        n_jobs = len(jobs)
        
        # Извлекаем приоритеты работ и исходные индексы
        priorities = []
        original_indices = []
        for i, job in enumerate(jobs):
            p_str = (job.get("priority", "") or "").strip()
            try:
                priority = int(p_str) if p_str else 999
            except:
                priority = 999
            priorities.append(priority)
            original_indices.append(job.get("_original_index", i))
        
        print(f"CP-SAT оптимизация для {line}: {n_jobs} работ")
        print(f"Приоритеты: {priorities}")
        print(f"Заблокированные приоритеты: {fix_priorities}")
        
        # === ПЕРЕМЕННЫЕ ===
        # Позиция каждой работы в итоговом расписании (0..n_jobs-1)
        positions = []
        for i in range(n_jobs):
            positions.append(model.NewIntVar(0, n_jobs - 1, f'pos_{i}'))
        
        # Все работы должны быть на разных позициях
        model.AddAllDifferent(positions)
        
        # === ОГРАНИЧЕНИЯ ПРИОРИТЕТОВ ===
        # Работы с более высоким приоритетом (меньшее число) должны идти раньше
        priority_constraints = 0
        for i in range(n_jobs):
            for j in range(n_jobs):
                if i != j and priorities[i] < priorities[j]:
                    # Если приоритет работы i выше, она должна быть раньше работы j
                    model.Add(positions[i] < positions[j])
                    priority_constraints += 1
        
        print(f"Добавлено ограничений по приоритетам: {priority_constraints}")
        
        # === ОГРАНИЧЕНИЯ ДЛЯ ЗАБЛОКИРОВАННЫХ ПРИОРИТЕТОВ ===
        # Для заблокированных приоритетов сохраняем исходный порядок внутри группы
        locked_constraints = 0
        for i in range(n_jobs):
            for j in range(n_jobs):
                if i != j and priorities[i] == priorities[j]:
                    # Работы с одинаковым приоритетом
                    if priorities[i] in fix_priorities:
                        # Приоритет заблокирован - сохраняем исходный порядок
                        if original_indices[i] < original_indices[j]:
                            model.Add(positions[i] < positions[j])
                            locked_constraints += 1
        
        if locked_constraints > 0:
            print(f"🔒 Добавлено ограничений для заблокированных приоритетов: {locked_constraints}")
        else:
            print("Нет ограничений для заблокированных приоритетов (все приоритеты могут оптимизироваться)")
        
        # === ОЦЕНКА ВРЕМЕНИ ПЕРЕХОДОВ ===
        # Для каждой пары (i, j) вычисляем, будет ли работа j следовать сразу после работы i
        follows = {}
        for i in range(n_jobs):
            for j in range(n_jobs):
                if i != j:
                    # Булева переменная: работа j идет сразу после работы i
                    follows[(i, j)] = model.NewBoolVar(f'follows_{i}_{j}')
                    # follows[i,j] = 1 тогда и только тогда, когда positions[j] = positions[i] + 1
                    model.Add(positions[j] == positions[i] + 1).OnlyEnforceIf(follows[(i, j)])
                    model.Add(positions[j] != positions[i] + 1).OnlyEnforceIf(follows[(i, j)].Not())
        
        # === ЦЕЛЕВАЯ ФУНКЦИЯ ===
        # Минимизируем суммарное время переходов CIP
        transition_cost = []
        
        # Кэшируем расчет времени переходов (значительно ускоряет для больших задач)
        transition_cache = {}
        
        for i in range(n_jobs):
            for j in range(n_jobs):
                if i != j:
                    # Проверяем кэш
                    if (i, j) not in transition_cache:
                        transition_time, _ = _transition_time_estimate(line, jobs[i], jobs[j])
                        transition_cache[(i, j)] = transition_time
                    else:
                        transition_time = transition_cache[(i, j)]
                    
                    # Если работа j следует за работой i, добавляем время перехода к стоимости
                    # Используем промежуточную переменную для произведения
                    # Лимит 1500: CIP3=300, Переналадка=300, с запасом
                    cost_var = model.NewIntVar(0, 1500, f'cost_{i}_{j}')
                    model.Add(cost_var == transition_time).OnlyEnforceIf(follows[(i, j)])
                    model.Add(cost_var == 0).OnlyEnforceIf(follows[(i, j)].Not())
                    transition_cost.append(cost_var)
        
        # Суммарная стоимость переходов
        total_cost = model.NewIntVar(0, 100000, 'total_cost')
        model.Add(total_cost == sum(transition_cost))
        
        # Минимизируем общую стоимость переходов (при соблюдении приоритетов)
        model.Minimize(total_cost)
        
        # === РЕШЕНИЕ ===
        solver = cp_model.CpSolver()
        
        # === ОПТИМИЗАЦИЯ ПРОИЗВОДИТЕЛЬНОСТИ ===
        # 1. Параллелизация - используем все ядра CPU
        solver.parameters.num_search_workers = 8
        
        # 2. Таймаут адаптивный: меньше работ = меньше времени
        if n_jobs < 10:
            solver.parameters.max_time_in_seconds = 5.0
        elif n_jobs < 20:
            solver.parameters.max_time_in_seconds = 10.0
        elif n_jobs < 50:
            solver.parameters.max_time_in_seconds = 20.0
        else:
            solver.parameters.max_time_in_seconds = 30.0
        
        # 3. Hint - даем солверу текущий порядок как начальное решение
        # Это позволяет быстро найти базовое решение и улучшать его
        for i in range(n_jobs):
            model.AddHint(positions[i], i)
        
        # 4. Стратегия поиска - фокус на переменных с большим влиянием
        solver.parameters.search_branching = cp_model.PORTFOLIO_SEARCH
        
        # 5. Отключаем verbose лог
        solver.parameters.log_search_progress = False
        
        # Засекаем время
        import time
        start_time = time.time()
        
        status = solver.Solve(model)
        
        solve_time = time.time() - start_time
        
        if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
            # Извлекаем оптимизированный порядок
            job_order = [(i, solver.Value(positions[i])) for i in range(n_jobs)]
            job_order.sort(key=lambda x: x[1])  # Сортируем по позиции
            
            optimized_jobs = [jobs[job_idx] for job_idx, _ in job_order]
            
            total_transition = solver.Value(total_cost)
            status_text = "OPTIMAL" if status == cp_model.OPTIMAL else "FEASIBLE"
            print(f"[CP-SAT] OK: {status_text} за {solve_time:.2f}с: переходы = {total_transition} мин")
            
            # Проверяем, не нарушен ли порядок в заблокированных группах
            if locked_constraints > 0:
                violations = 0
                for i, (job_a_idx, _) in enumerate(job_order):
                    for j, (job_b_idx, _) in enumerate(job_order):
                        if i < j:
                            p_a = priorities[job_a_idx]
                            p_b = priorities[job_b_idx]
                            
                            if p_a == p_b and p_a in fix_priorities:
                                orig_a = original_indices[job_a_idx]
                                orig_b = original_indices[job_b_idx]
                                if orig_a > orig_b:
                                    violations += 1
                                    job_a_name = jobs[job_a_idx].get("name", "")[:30]
                                    job_b_name = jobs[job_b_idx].get("name", "")[:30]
                                    print(f"[WARNING] НАРУШЕНИЕ блокировки P{p_a}: '{job_a_name}' (orig {orig_a}) перед '{job_b_name}' (orig {orig_b})")
                
                if violations == 0:
                    print(f"[CP-SAT] Порядок в заблокированных группах сохранен ({locked_constraints} ограничений)")
                else:
                    print(f"[WARNING] ВНИМАНИЕ: Обнаружено {violations} нарушений блокировок!")
            
            return optimized_jobs
        else:
            print(f"[CP-SAT] Не нашел решение (статус {status}), используем исходный порядок")
            return jobs
            
    except Exception as e:
        print(f"[CP-SAT ERROR] Ошибка CP-SAT: {e}")
        import traceback
        traceback.print_exc()
        return jobs


def _sort_by_priority(jobs: list[dict]) -> list[dict]:
    """Сортировка по приоритету с сохранением исходного порядка внутри приоритета"""
    def priority_key(job):
        # 1. Приоритет (меньше = выше приоритет)
        p = (job.get("priority", "") or "").strip()
        try:
            user_priority = int(p) if p else 999
        except:
            user_priority = 999
        
        # 2. Исходный порядок для сохранения порядка внутри приоритета
        orig_idx = job.get("_original_index", 999)
        
        # 3. Количество (больше = выше)
        qty = _qty_to_int(job.get("quantity", ""))
        
        return (user_priority, orig_idx, -qty)
    
    return sorted(jobs, key=priority_key)

# ---------------------------------------------------------------------
# ОКНО НАСТРОЕК АВТОМАТИЧЕСКИХ CIP
# ---------------------------------------------------------------------

class DensitySettingsWindow:
    """Окно для настройки плотности продуктов"""
    
    def __init__(self, parent):
        self.window = tk.Toplevel(parent)
        self.window.title("Плотность продуктов")
        self.window.geometry("600x500")
        self.window.transient(parent)
        self.window.grab_set()
        
        # Загружаем текущие настройки плотности
        self.densities = _load_product_density()
        self.entries = {}
        
        # Создаем UI
        self._create_ui()
        
        # Центрируем окно
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (self.window.winfo_width() // 2)
        y = (self.window.winfo_screenheight() // 2) - (self.window.winfo_height() // 2)
        self.window.geometry(f"+{x}+{y}")
    
    def _create_ui(self):
        """Создание интерфейса"""
        # Заголовок
        header_frame = ttk.Frame(self.window)
        header_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(header_frame, text="Плотность продуктов (кг/л)", 
                 font=("TkDefaultFont", 12, "bold")).pack(anchor="w")
        ttk.Label(header_frame, text="Используется для расчета массы при режиме 'по массе'", 
                 font=("TkDefaultFont", 9)).pack(anchor="w", pady=(5, 0))
        
        ttk.Separator(self.window, orient="horizontal").pack(fill="x", padx=10, pady=5)
        
        # Область прокрутки
        canvas_frame = ttk.Frame(self.window)
        canvas_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        canvas = tk.Canvas(canvas_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Создаем поля для каждого типа продукта
        for product_type in sorted(self.densities.keys()):
            self._create_density_row(scrollable_frame, product_type)
        
        # Кнопка добавления нового типа
        add_frame = ttk.Frame(scrollable_frame)
        add_frame.pack(fill="x", padx=5, pady=10)
        ttk.Button(add_frame, text="+ Добавить тип продукта", 
                  command=lambda: self._add_new_type(scrollable_frame)).pack()
        
        # Кнопки внизу
        button_frame = ttk.Frame(self.window)
        button_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        ttk.Button(button_frame, text="Сохранить", 
                  command=self._save_densities, 
                  style="Accent.TButton").pack(side="right", padx=(5, 0))
        ttk.Button(button_frame, text="Отмена", 
                  command=self.window.destroy).pack(side="right")
    
    def _create_density_row(self, parent, product_type):
        """Создать строку для одного типа продукта"""
        row = ttk.Frame(parent)
        row.pack(fill="x", padx=5, pady=2)
        
        ttk.Label(row, text=f"{product_type}:", width=20).pack(side="left", padx=(0, 10))
        
        entry = ttk.Entry(row, width=10)
        entry.insert(0, str(self.densities.get(product_type, 1.0)))
        entry.pack(side="left", padx=(0, 5))
        
        ttk.Label(row, text="кг/л").pack(side="left", padx=(0, 10))
        
        # Кнопка удаления
        ttk.Button(row, text="X", width=3, 
                  command=lambda: self._remove_row(row, product_type)).pack(side="left")
        
        self.entries[product_type] = entry
    
    def _add_new_type(self, parent):
        """Добавить новый тип продукта"""
        # Диалог ввода имени
        new_type = tk.simpledialog.askstring("Новый тип", "Введите название типа продукта:")
        if new_type and new_type.strip():
            new_type = new_type.strip()
            if new_type not in self.entries:
                self.densities[new_type] = 1.0
                self._create_density_row(parent, new_type)
    
    def _remove_row(self, row, product_type):
        """Удалить строку"""
        if messagebox.askyesno("Подтверждение", f"Удалить тип '{product_type}'?"):
            row.destroy()
            if product_type in self.entries:
                del self.entries[product_type]
            if product_type in self.densities:
                del self.densities[product_type]
    
    def _save_densities(self):
        """Сохранить настройки плотности"""
        try:
            new_densities = {}
            for product_type, entry in self.entries.items():
                try:
                    density = float(entry.get().replace(",", "."))
                    if density <= 0:
                        raise ValueError()
                    new_densities[product_type] = density
                except ValueError:
                    messagebox.showerror("Ошибка", f"Неверное значение плотности для '{product_type}'")
                    return
            
            # Сохраняем в файл
            _save_json(_DENSITY_JSON, new_densities)
            
            messagebox.showinfo("Успех", "Настройки плотности сохранены!")
            self.window.destroy()
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить настройки: {e}")


class AutoCipSettingsWindow:
    """Окно для настройки автоматических CIP"""
    
    def __init__(self, parent):
        self.window = tk.Toplevel(parent)
        self.window.title("Настройки автоматических CIP")
        self.window.geometry("800x600")
        self.window.transient(parent)
        self.window.grab_set()
        
        # Загружаем текущие настройки
        self.thresholds = _load_cip_thresholds()
        self.entries = {}
        
        # Создаем UI
        self._create_ui()
        
        # Центрируем окно
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (self.window.winfo_width() // 2)
        y = (self.window.winfo_screenheight() // 2) - (self.window.winfo_height() // 2)
        self.window.geometry(f"+{x}+{y}")
    
    def _create_ui(self):
        """Создание интерфейса"""
        # Заголовок
        header_frame = ttk.Frame(self.window)
        header_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(header_frame, text="⚙️ Настройки автоматических CIP", 
                 font=("TkDefaultFont", 12, "bold")).pack(anchor="w")
        ttk.Label(header_frame, text="Автоматические CIP вставляются при достижении порогов объема", 
                 font=("TkDefaultFont", 9)).pack(anchor="w", pady=(5, 0))
        
        ttk.Separator(self.window, orient="horizontal").pack(fill="x", padx=10, pady=5)
        
        # Область прокрутки
        canvas_frame = ttk.Frame(self.window)
        canvas_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        canvas = tk.Canvas(canvas_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Создаем настройки для каждой линии
        for line_name in sorted([f"линия {i}" for i in range(1, 11)]):
            self._create_line_settings(scrollable_frame, line_name)
        
        # Кнопки внизу
        button_frame = ttk.Frame(self.window)
        button_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        ttk.Button(button_frame, text="💾 Сохранить", 
                  command=self._save_settings, 
                  style="Accent.TButton").pack(side="right", padx=(5, 0))
        ttk.Button(button_frame, text="❌ Отмена", 
                  command=self.window.destroy).pack(side="right")
        ttk.Button(button_frame, text="⚖️ Плотность продуктов", 
                  command=self._open_density_settings).pack(side="left", padx=(0, 5))
        ttk.Button(button_frame, text="🔄 Сбросить по умолчанию", 
                  command=self._reset_defaults).pack(side="left")
    
    def _create_line_settings(self, parent, line_name):
        """Создать настройки для одной линии"""
        # Получаем текущие настройки для линии
        config = self.thresholds.get(line_name, {
            "enabled": True,
            "volume_threshold": 50000,
            "product_threshold": 30000,
            "cip_type": "CIP2"
        })
        
        # Фрейм для линии
        line_frame = ttk.LabelFrame(parent, text=f"📍 {line_name.upper()}", padding=10)
        line_frame.pack(fill="x", padx=5, pady=5)
        
        # Первая строка: включено/выключено
        row1 = ttk.Frame(line_frame)
        row1.pack(fill="x", pady=(0, 5))
        
        enabled_var = tk.BooleanVar(value=config.get("enabled", True))
        ttk.Checkbutton(row1, text="Включить автоматические CIP", 
                       variable=enabled_var).pack(side="left")
        
        # Вторая строка: режим
        row2 = ttk.Frame(line_frame)
        row2.pack(fill="x", pady=(0, 5))
        
        ttk.Label(row2, text="Режим:").pack(side="left", padx=(0, 5))
        mode_var = tk.StringVar(value=config.get("mode", "штуки"))
        mode_combo = ttk.Combobox(row2, textvariable=mode_var, 
                                values=["штуки", "масса"], 
                                width=12, state="readonly")
        mode_combo.pack(side="left", padx=(0, 15))
        
        # Метка для единиц измерения
        unit_label = ttk.Label(row2, text="(штуки)")
        unit_label.pack(side="left", padx=(0, 15))
        
        # Обновляем метку при изменении режима
        def update_unit_label(*args):
            if mode_var.get() == "штуки":
                unit_label.config(text="(штуки)")
            else:
                unit_label.config(text="(кг)")
        mode_var.trace('w', update_unit_label)
        update_unit_label()  # Инициализация
        
        # Третья строка: пороги
        row3 = ttk.Frame(line_frame)
        row3.pack(fill="x", pady=(0, 5))
        
        ttk.Label(row3, text="Порог общего объема:").pack(side="left", padx=(0, 5))
        volume_entry = ttk.Entry(row3, width=10)
        volume_entry.insert(0, str(config.get("volume_threshold", 50000)))
        volume_entry.pack(side="left", padx=(0, 15))
        
        ttk.Label(row3, text="Порог по продукту:").pack(side="left", padx=(0, 5))
        product_entry = ttk.Entry(row3, width=10)
        product_entry.insert(0, str(config.get("product_threshold", 30000)))
        product_entry.pack(side="left", padx=(0, 15))

        ttk.Label(row3, text="Буфер:").pack(side="left", padx=(0, 5))
        min_remainder_entry = ttk.Entry(row3, width=8)
        min_remainder_entry.insert(0, str(config.get("min_remainder", 2000)))
        min_remainder_entry.pack(side="left", padx=(0, 15))

        ttk.Label(row3, text="Тип CIP:").pack(side="left", padx=(0, 5))
        cip_type_var = tk.StringVar(value=config.get("cip_type", "CIP2"))
        cip_combo = ttk.Combobox(row3, textvariable=cip_type_var,
                                values=["CIP1", "CIP2", "CIP3"],
                                width=8, state="readonly")
        cip_combo.pack(side="left")

        # Сохраняем ссылки на виджеты
        self.entries[line_name] = {
            "enabled": enabled_var,
            "mode": mode_var,
            "volume_threshold": volume_entry,
            "product_threshold": product_entry,
            "min_remainder": min_remainder_entry,
            "cip_type": cip_type_var
        }
    
    def _open_density_settings(self):
        """Открыть окно настройки плотности"""
        DensitySettingsWindow(self.window)
    
    def _save_settings(self):
        """Сохранить настройки"""
        try:
            # Собираем данные из полей
            new_thresholds = []
            
            for line_name, widgets in self.entries.items():
                try:
                    volume = int(widgets["volume_threshold"].get().replace(",", "").replace(" ", ""))
                    product = int(widgets["product_threshold"].get().replace(",", "").replace(" ", ""))
                    min_remainder = int(widgets["min_remainder"].get().replace(",", "").replace(" ", ""))
                except ValueError:
                    messagebox.showerror("Ошибка", f"Неверный формат чисел для {line_name}")
                    return

                new_thresholds.append({
                    "line": line_name,
                    "mode": widgets["mode"].get(),
                    "volume_threshold": volume,
                    "product_threshold": product,
                    "min_remainder": min_remainder,
                    "cip_type": widgets["cip_type"].get(),
                    "enabled": widgets["enabled"].get()
                })
            
            # Сохраняем в файл
            _save_json(_CIP_THRESHOLDS_JSON, new_thresholds)
            
            messagebox.showinfo("Успех", "Настройки автоматических CIP сохранены!")
            self.window.destroy()
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить настройки: {e}")
    
    def _reset_defaults(self):
        """Сбросить настройки по умолчанию"""
        if messagebox.askyesno("Подтверждение", 
                              "Сбросить все настройки на значения по умолчанию?"):
            for line_name, widgets in self.entries.items():
                widgets["enabled"].set(True)
                widgets["volume_threshold"].delete(0, tk.END)
                widgets["volume_threshold"].insert(0, "50000")
                widgets["product_threshold"].delete(0, tk.END)
                widgets["product_threshold"].insert(0, "30000")
                widgets["min_remainder"].delete(0, tk.END)
                widgets["min_remainder"].insert(0, "2000")
                widgets["cip_type"].set("CIP2")


# ---------------------------------------------------------------------
# UI КЛАСС
# ---------------------------------------------------------------------

class ScheduleTab:
    def __init__(self, parent):
        self.parent = parent

        # === ВЕРХНЯЯ ПАНЕЛЬ УПРАВЛЕНИЯ ===
        control_frame = ttk.Frame(parent)
        control_frame.pack(fill="x", padx=8, pady=(8, 4))
        
        # Левая группа - основные действия
        left_group = ttk.LabelFrame(control_frame, text="Основные действия", padding=8)
        left_group.pack(side="left", fill="x", expand=True, padx=(0, 8))
        
        # Чекбокс оптимизации (отключен - всегда используется порядок по приоритету)
        self.var_optimize = tk.BooleanVar(value=False)
        opt_check = ttk.Checkbutton(left_group, text="Использовать оптимизацию CP-SAT", 
                                   variable=self.var_optimize, state="disabled")
        opt_check.pack(side="left", padx=(0, 12))
        
        # Поле ввода приоритетов для фиксации
        ttk.Label(left_group, text="🔒 Фиксировать для приоритетов:").pack(side="left", padx=(0, 4))
        self.fix_priorities_entry = ttk.Entry(left_group, width=15)
        self.fix_priorities_entry.pack(side="left", padx=(0, 12))
        # По умолчанию поле пустое - без блокировок
        
        # Кнопка построения
        build_btn = ttk.Button(left_group, text="📅 Построить расписание", 
                              command=self.build_schedule, style="Accent.TButton")
        build_btn.pack(side="left", padx=(0, 8))
        
        # Кнопка автофита
        autofit_btn = ttk.Button(left_group, text="📏 Автофит колонок", 
                               command=self._autofit_schedule)
        autofit_btn.pack(side="left", padx=(0, 8))
        
        # Кнопка открытия окна привязок линий
        bindings_btn = ttk.Button(left_group, text="🔗 Привязка линий", 
                                 command=self.open_line_bindings_window)
        bindings_btn.pack(side="left", padx=(0, 8))
        
        # Хранилище привязок линий
        self.line_bindings = {}
        self._load_line_bindings()
        
        # Правая группа - дополнительные действия
        right_group = ttk.LabelFrame(control_frame, text="Дополнительно", padding=8)
        right_group.pack(side="right", fill="x")
        
        # Кнопки и настройки
        btn_frame = ttk.Frame(right_group)
        btn_frame.pack(fill="x")
        
        # Чекбокс для CP-SAT оптимизации
        self.use_cp_sat_var = tk.BooleanVar()
        cp_sat_checkbox = ttk.Checkbutton(
            btn_frame, 
            text="🔧 CP-SAT оптимизация", 
            variable=self.use_cp_sat_var,
            state="normal" if CP_SAT_AVAILABLE else "disabled"
        )
        cp_sat_checkbox.pack(side="left", padx=(0, 10))
        
        if not CP_SAT_AVAILABLE:
            cp_sat_checkbox.config(text="🔧 CP-SAT недоступен")
        
        ttk.Button(btn_frame, text="⚙️ Авто-CIP", 
                  command=self.open_auto_cip_settings).pack(side="left", padx=(0, 6))
        
        ttk.Button(btn_frame, text="📋 Упрощенный вид", 
                  command=self.show_simple_view).pack(side="left", padx=(0, 6))
        
        ttk.Button(btn_frame, text="📤 Экспорт Excel", 
                  command=self.export_excel).pack(side="left", padx=(0, 6))
        
        # === ОСНОВНАЯ ТАБЛИЦА ===
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        
        # Создаем Treeview с группировкой по сменам
        self.tree = ttk.Treeview(table_frame, columns=SCHED_COLS, show="tree headings", height=20)
        
        # Настройка заголовков
        headers = {
            "date": "Дата",
            "line": "Линия", 
            "job_id": "ID задания",
            "name": "Наименование",
            "duration": "Длительность",
            "volume": "Объём",
            "qty": "Кол-во",
            "flavor": "Вкус",
            "brand": "Бренд",
            "type": "Тип",
            "start": "Начало",
            "end": "Окончание",
            "note": "Примечание"
        }
        
        column_widths = {
            "date": 80,
            "line": 80,
            "job_id": 120,
            "name": 300,
            "duration": 80,
            "volume": 80,
            "qty": 60,
            "flavor": 200,
            "brand": 120,
            "type": 100,
            "start": 100,
            "end": 100,
            "note": 200
        }
        
        for col in SCHED_COLS:
            self.tree.heading(col, text=headers.get(col, col))
            width = column_widths.get(col, 120)
            anchor = "e" if col in ["duration", "qty"] else "w"
            self.tree.column(col, width=width, anchor=anchor, minwidth=60)
        
        # Настройка тегов для автоматических CIP (без выделения)
        # self.tree.tag_configure("auto_cip", foreground="#8B0000")

        # Скроллбары
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # === СТАТУС И ИТОГИ ===
        status_frame = ttk.Frame(parent)
        status_frame.pack(fill="x", padx=8, pady=(0, 8))
        
        # Заголовок для итогов
        ttk.Label(status_frame, text="📊 ИТОГИ:", font=("TkDefaultFont", 9, "bold")).pack(side="left", padx=(0, 10))
        
        # Информация о статусе
        self.status_label = ttk.Label(status_frame, text="Готов", anchor="w", foreground="blue")
        self.status_label.pack(side="left", fill="x", expand=True)
        
        # Разделитель
        ttk.Separator(parent, orient="horizontal").pack(fill="x", padx=8)
        
    def open_line_bindings_window(self):
        """Открытие окна привязок линий"""
        window = tk.Toplevel(self.parent)
        window.title("Привязка линий")
        window.geometry("500x400")
        window.transient(self.parent)
        window.grab_set()
        
        # Загружаем список линий
        plan = _load_json(_PLAN_JSON, [])
        lines = sorted(set(job.get("line", "").strip() for job in plan if job.get("line")))
        
        # Фрейм для формы
        form_frame = ttk.Frame(window, padding=10)
        form_frame.pack(fill="both", expand=True)
        
        # Выбор исходной линии
        ttk.Label(form_frame, text="Заканчивается:").grid(row=0, column=0, sticky="w", pady=5)
        source_var = tk.StringVar()
        source_combo = ttk.Combobox(form_frame, textvariable=source_var, width=30, state="readonly")
        source_combo['values'] = lines
        source_combo.grid(row=0, column=1, sticky="ew", pady=5, padx=5)
        
        # Выбор целевой линии
        ttk.Label(form_frame, text="Начинается:").grid(row=1, column=0, sticky="w", pady=5)
        target_var = tk.StringVar()
        target_combo = ttk.Combobox(form_frame, textvariable=target_var, width=30, state="readonly")
        target_combo['values'] = lines
        target_combo.grid(row=1, column=1, sticky="ew", pady=5, padx=5)
        
        # Список существующих привязок
        ttk.Label(form_frame, text="Существующие привязки:").grid(row=2, column=0, columnspan=2, sticky="w", pady=10)
        bindings_text = tk.Text(form_frame, height=10, width=50)
        bindings_text.grid(row=3, column=0, columnspan=2, sticky="nsew", pady=5)
        
        # Загружаем существующие привязки
        for target, source in self.line_bindings.items():
            bindings_text.insert("end", f"{source} -> {target}\n")
        
        # Кнопки
        btn_frame = ttk.Frame(form_frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=10)
        
        def add_binding():
            source = source_var.get()
            target = target_var.get()
            
            if not source or not target:
                messagebox.showwarning("Ошибка", "Выберите обе линии")
                return
            
            if source == target:
                messagebox.showwarning("Ошибка", "Нельзя привязать линию к самой себе")
                return
            
            # Добавляем привязку
            self.line_bindings[target] = source
            
            # Обновляем список
            bindings_text.insert("end", f"{source} -> {target}\n")
            
            # Очищаем поля
            source_var.set("")
            target_var.set("")
            
            # Сохраняем изменения
            self._save_line_bindings()
        
        def remove_binding():
            # Получаем выделенную строку
            try:
                current_line = bindings_text.index("insert linestart")
                line_text = bindings_text.get(current_line, current_line + " lineend").strip()
                
                if "->" in line_text:
                    parts = line_text.split("->")
                    if len(parts) == 2:
                        target = parts[1].strip()
                        if target in self.line_bindings:
                            del self.line_bindings[target]
                            bindings_text.delete(current_line, current_line + " lineend")
                            self._save_line_bindings()
            except:
                messagebox.showwarning("Ошибка", "Выберите привязку для удаления")
        
        def clear_all():
            self.line_bindings.clear()
            bindings_text.delete("1.0", tk.END)
            self._save_line_bindings()
        
        def close_window():
            window.destroy()
        
        ttk.Button(btn_frame, text="➕ Добавить", command=add_binding).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="➖ Удалить", command=remove_binding).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="🗑 Очистить все", command=clear_all).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="✓ Готово", command=close_window).pack(side="left", padx=5)
        
        form_frame.grid_columnconfigure(1, weight=1)
        form_frame.grid_rowconfigure(3, weight=1)
    
    def _load_line_bindings(self):
        """Загрузка привязок линий из файла"""
        try:
            bindings_file = "line_bindings.json"
            if os.path.exists(bindings_file):
                with open(bindings_file, "r", encoding="utf-8") as f:
                    self.line_bindings = json.load(f)
                print(f"Загружено {len(self.line_bindings)} привязок линий")
        except Exception as e:
            print(f"Ошибка загрузки привязок линий: {e}")
            self.line_bindings = {}
    
    def _save_line_bindings(self):
        """Сохранение привязок линий в файл"""
        try:
            bindings_file = "line_bindings.json"
            with open(bindings_file, "w", encoding="utf-8") as f:
                json.dump(self.line_bindings, f, ensure_ascii=False, indent=2)
            print(f"Сохранено {len(self.line_bindings)} привязок линий")
        except Exception as e:
            print(f"Ошибка сохранения привязок линий: {e}")
    
    def _load_locked_priorities(self):
        """Загрузка заблокированных приоритетов"""
        try:
            with open("locked_priorities.json", "r", encoding="utf-8") as f:
                data = json.load(f)
                locked = set(data.get("locked", []))
                print(f"Загружено {len(locked)} заблокированных приоритетов: {locked}")
                return locked
        except:
            return set()
    
    def build_schedule(self):
        """Построение расписания"""
        self.status_label.config(text="Построение расписания...")
        
        try:
            # Получаем приоритеты для фиксации
            fix_priorities_str = self.fix_priorities_entry.get().strip()
            fix_priorities = []
            if fix_priorities_str:
                for p in re.split(r'[,\s]+', fix_priorities_str):
                    try:
                        fix_priorities.append(int(p.strip()))
                    except:
                        pass
            
            # Добавляем заблокированные приоритеты из настроек
            locked_priorities = self._load_locked_priorities()
            fix_priorities.extend(locked_priorities)
            fix_priorities = list(set(fix_priorities))  # Убираем дубликаты
            
            # Получаем настройку CP-SAT
            use_cp_sat = self.use_cp_sat_var.get()
            
            # Получаем привязки из хранилища
            line_bindings = self.line_bindings if self.line_bindings else None
            
            # Строим расписание
            schedule = build_schedule_from_plan(fix_priorities=fix_priorities, use_cp_sat=use_cp_sat, line_bindings=line_bindings)
            
            if not schedule:
                self.status_label.config(text="Нет данных для расписания")
                return

            # Отображаем в таблице (включая итоги)
            self._display_schedule(schedule)
            
            # Сохраняем
            _save_json(_SCHEDULE_JSON, schedule)
            
        except Exception as e:
            self.status_label.config(text=f"Ошибка: {e}")
            print(f"Ошибка построения расписания: {e}")
    
    def _display_schedule(self, schedule: list[dict]):
        """Отображение расписания в таблице с группировкой по сменам"""
        # Очищаем таблицу
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Разрезаем выпуски, которые переходят через смены
        split_schedule = _split_jobs_across_shifts(schedule)
        
        # Вычисляем общее затраченное время
        total_duration = sum(int(record.get("duration", "0") or "0") for record in split_schedule)
        total_hours = total_duration // 60
        total_minutes = total_duration % 60
        
        # Вычисляем время только для работ (без CIP)
        job_duration = sum(
            int(record.get("duration", "0") or "0") 
            for record in split_schedule 
            if record.get("type", "") not in ["CIP1", "CIP2", "CIP3"]
        )
        job_hours = job_duration // 60
        job_minutes = job_duration % 60
        
        # Вычисляем время CIP
        cip_duration = sum(
            int(record.get("duration", "0") or "0") 
            for record in split_schedule 
            if record.get("type", "") in ["CIP1", "CIP2", "CIP3"]
        )
        cip_hours = cip_duration // 60
        cip_minutes = cip_duration % 60
        
        # Группируем по сменам
        shifts = _group_schedule_by_shifts(split_schedule)
        
        # Сортируем смены по дате
        sorted_shifts = sorted(shifts.items())
        
        for shift_key, shift_records in sorted_shifts:
            # Парсим ключ смены
            date_str, shift_name = shift_key.split("_")
            shift_date = dt.datetime.strptime(date_str, "%Y-%m-%d").strftime("%d.%m.%Y")
            
            # Создаем заголовок смены
            shift_title = f"📅 {shift_date} - {shift_name} смена ({len(shift_records)} событий)"
            shift_item = self.tree.insert("", "end", text=shift_title, values=[""] * len(SCHED_COLS))
            
            # Группируем события по линиям
            lines_data = {}
            for record in shift_records:
                line = record.get("line", "")
                if line not in lines_data:
                    lines_data[line] = []
                lines_data[line].append(record)
            
            # Сортируем линии по времени первого события
            def get_first_event_time(line_name):
                if lines_data[line_name]:
                    first_record = lines_data[line_name][0]
                    start_str = first_record.get("start", "")
                    m = re.search(r"(\d{1,2})\.(\d{1,2})\s+(\d{2}):(\d{2})", start_str)
                    if m:
                        hour, minute = int(m.group(3)), int(m.group(4))
                        return hour * 60 + minute
                return 0
            
            sorted_lines = sorted(lines_data.keys(), key=get_first_event_time)
            
            # Отображаем линии в отсортированном порядке
            for line in sorted_lines:
                current_line_group = self.tree.insert(
                    shift_item, "end", 
                    text=f"📍 {line}",
                    values=[""] * len(SCHED_COLS)
                )
                self.tree.item(current_line_group, open=True)
                
                for record in lines_data[line]:
                    values = [record.get(col, "") for col in SCHED_COLS]
                    tags = ("auto_cip",) if record.get("_auto_cip") else ()
                    self.tree.insert(current_line_group, "end", values=values, tags=tags)
            
            # Разворачиваем смену по умолчанию
            self.tree.item(shift_item, open=True)
        
        # Вычисляем количество работ и CIP
        job_count = len([r for r in split_schedule if r.get("type", "") not in ["CIP1", "CIP2", "CIP3"]])
        cip_count = len([r for r in split_schedule if r.get("type", "") in ["CIP1", "CIP2", "CIP3"]])
        
        # Вычисляем общее количество продукции
        total_qty = 0
        for record in split_schedule:
            if record.get("type", "") not in ["CIP1", "CIP2", "CIP3"]:
                qty_str = record.get("qty", "").replace(",", "").replace(" ", "")
                try:
                    total_qty += int(qty_str) if qty_str else 0
                except (ValueError, TypeError):
                    pass
        
        # Обновляем статус с полной информацией
        summary = f"📊 ИТОГО: {len(split_schedule)} событий | Работ: {job_count} | CIP: {cip_count}"
        time_info = f"⏱ Время: {total_hours}ч {total_minutes}м | Работы: {job_hours}ч {job_minutes}м | CIP: {cip_hours}ч {cip_minutes}м"
        qty_info = f"📦 Произведено: {total_qty:,} шт"
        
        self.status_label.config(
            text=f"{summary} | {time_info} | {qty_info}"
        )
    
    def _autofit_schedule(self):
        """Автоподгонка ширины колонок"""
        for col in SCHED_COLS:
            col_data = [str(self.tree.set(item, col)) for item in self.tree.get_children()[:50]]
            if col_data:
                max_len = max(len(d) for d in col_data)
                width = min(max(max_len * 8, 80), 400)
                self.tree.column(col, width=width)
    
    def open_auto_cip_settings(self):
        """Открыть окно настроек автоматических CIP"""
        AutoCipSettingsWindow(self.parent)
    
    def show_simple_view(self):
        """Показать упрощенный вид расписания: только продукты и переходы"""
        try:
            schedule = _load_json(_SCHEDULE_JSON, [])
            if not schedule:
                messagebox.showwarning("Упрощенный вид", "Нет данных расписания. Постройте расписание сначала.")
                return
            
            # Создаем окно
            window = tk.Toplevel(self.parent)
            window.title("Упрощенный вид расписания")
            window.geometry("900x700")
            
            # Заголовок
            header_frame = ttk.Frame(window)
            header_frame.pack(fill="x", padx=10, pady=10)
            ttk.Label(header_frame, text="📋 Упрощенный вид: Продукты и переходы", 
                     font=("", 12, "bold")).pack(side="left")
            
            # Фрейм для таблицы
            table_frame = ttk.Frame(window)
            table_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
            
            # Создаем Treeview
            columns = ("product", "duration", "quantity")
            tree = ttk.Treeview(table_frame, columns=columns, show="tree headings", height=30)

            tree.heading("product", text="Наименование продукта")
            tree.heading("duration", text="Длительность")
            tree.heading("quantity", text="Количество")

            tree.column("#0", width=150)  # Для групп линий
            tree.column("product", width=500)
            tree.column("duration", width=120, anchor="center")
            tree.column("quantity", width=120, anchor="center")
            
            # Скроллбары
            vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
            hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            
            tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
            
            table_frame.grid_rowconfigure(0, weight=1)
            table_frame.grid_columnconfigure(0, weight=1)
            
            # Цветовые теги
            tree.tag_configure("production", background="#E8F5E9")
            tree.tag_configure("cip", background="#FFE0E0")
            tree.tag_configure("eviction", background="#FFF9C4")
            tree.tag_configure("format", background="#E3F2FD")
            
            # Группируем по линиям
            lines_data = {}
            for record in schedule:
                line = record.get("line", "Без линии")
                if line not in lines_data:
                    lines_data[line] = []
                lines_data[line].append(record)
            
            # Заполняем дерево
            for line, records in sorted(lines_data.items()):
                # Создаем группу линии
                line_id = tree.insert("", "end", text=f"📍 {line}", values=("", "", ""))
                tree.item(line_id, open=True)

                for record in records:
                    job_id = record.get("job_id", "")
                    name = record.get("name", "")
                    duration = record.get("duration", "")
                    quantity = record.get("qty", "")
                    type_event = record.get("type", "")

                    # Определяем тип записи
                    if job_id.startswith("AUTO-CIP-"):
                        # Для автоматических CIP используем полное название с указанием типа
                        product_display = name if name else (type_event if type_event else "CIP")
                        tag = "cip"
                    elif job_id.startswith("CIP-"):
                        # Для обычных CIP показываем тип (CIP1, CIP2, CIP3)
                        product_display = type_event if type_event else "CIP"
                        tag = "cip"
                    elif job_id.startswith("ВЫТ-"):
                        product_display = "Вытеснение"
                        tag = "eviction"
                    elif job_id.startswith("П-"):
                        product_display = "Переналадка формата"
                        tag = "format"
                    else:
                        product_display = name
                        tag = "production"

                    duration_display = f"{duration} мин" if duration else "—"
                    quantity_display = quantity if quantity else "—"

                    tree.insert(line_id, "end", values=(product_display, duration_display, quantity_display), tags=(tag,))
            
            # Кнопки внизу
            btn_frame = ttk.Frame(window)
            btn_frame.pack(fill="x", padx=10, pady=(0, 10))
            
            ttk.Label(btn_frame, text="💡 Зеленый — производство | Красный — CIP | Желтый — вытеснение | Синий — переналадка", 
                     foreground="#666", font=("", 9)).pack(side="left")
            
            ttk.Button(btn_frame, text="📥 Экспорт в Excel", 
                      command=lambda: self._export_simple_view_to_excel(schedule, lines_data)).pack(side="right", padx=(5, 0))
            ttk.Button(btn_frame, text="Закрыть", command=window.destroy).pack(side="right")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось показать упрощенный вид: {e}")
            import traceback
            traceback.print_exc()
    
    def _export_simple_view_to_excel(self, schedule, lines_data):
        """Экспорт упрощенного вида в Excel"""
        try:
            from tkinter import filedialog
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Запрашиваем имя файла
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")],
                title="Сохранить упрощенное расписание"
            )
            
            if not filename:
                return
            
            # Создаем книгу Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Упрощенное расписание"
            
            # Настройка стилей
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
            
            production_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            cip_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
            eviction_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            format_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Заголовок
            ws['A1'] = "Линия"
            ws['B1'] = "Наименование продукта"
            ws['C1'] = "Длительность"
            ws['D1'] = "Количество"

            for cell in ['A1', 'B1', 'C1', 'D1']:
                ws[cell].font = header_font
                ws[cell].fill = header_fill
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')
                ws[cell].border = border

            # Ширина колонок
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            
            # Заполняем данные
            current_row = 2
            
            for line, records in sorted(lines_data.items()):
                for record in records:
                    job_id = record.get('job_id', '')
                    name = record.get('name', '')
                    duration = record.get('duration', '')
                    quantity = record.get('qty', '')
                    type_event = record.get('type', '')

                    # Определяем тип и заливку
                    if job_id.startswith('AUTO-CIP-'):
                        # Для автоматических CIP используем полное название с указанием типа
                        product_display = name if name else (type_event if type_event else "CIP")
                        fill = cip_fill
                    elif job_id.startswith('CIP-'):
                        product_display = type_event if type_event else "CIP"
                        fill = cip_fill
                    elif job_id.startswith('ВЫТ-'):
                        product_display = "Вытеснение"
                        fill = eviction_fill
                    elif job_id.startswith('П-'):
                        product_display = "Переналадка формата"
                        fill = format_fill
                    else:
                        product_display = name
                        fill = production_fill

                    duration_display = f"{duration} мин" if duration else "—"
                    quantity_display = quantity if quantity else "—"

                    # Записываем строку
                    ws[f'A{current_row}'] = line
                    ws[f'B{current_row}'] = product_display
                    ws[f'C{current_row}'] = duration_display
                    ws[f'D{current_row}'] = quantity_display

                    # Применяем форматирование
                    for col in ['A', 'B', 'C', 'D']:
                        cell = ws[f'{col}{current_row}']
                        cell.fill = fill
                        cell.border = border
                        if col in ['A', 'C', 'D']:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    current_row += 1
            
            # Сохраняем файл
            wb.save(filename)
            messagebox.showinfo("Успех", f"Упрощенное расписание экспортировано в:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Ошибка экспорта", f"Не удалось экспортировать: {e}")
            import traceback
            traceback.print_exc()
    
    def export_excel(self):
        """Экспорт расписания в Excel"""
        try:
            schedule = _load_json(_SCHEDULE_JSON, [])
            if not schedule:
                messagebox.showwarning("Экспорт", "Нет данных для экспорта")
                return
            
            show_export_dialog(self.parent, schedule)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать: {e}")

if __name__ == "__main__":
    # Тест
    result = build_schedule_from_plan(fix_priorities=[1])
    print(f"Построено {len(result)} записей")
    # Сохраняем для проверки
    _save_json(_SCHEDULE_JSON, result)
